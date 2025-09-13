

import os
import sqlite3
import hashlib
import csv
import threading
import time
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk



try:
    import matplotlib
    matplotlib.use('TkAgg')  # Set the backend before importing pyplot
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("Matplotlib not available - charts will be disabled")


# Optional libs
try:
    import ttkbootstrap as tb
    from ttkbootstrap.icons import Icon
    TTB_AVAILABLE = True
except Exception:
    tb = None
    Icon = None
    TTB_AVAILABLE = False
try:
    from tkcalendar import DateEntry as _TKDateEntry
    # wrapper to avoid ttkbootstrap conflicts
    class DateEntry(_TKDateEntry):
        def __init__(self, master=None, **kw):
            if 'bootstyle' in kw:
                kw.pop('bootstyle')
            super().__init__(master, **kw)
    TKCAL_AVAILABLE = True
except Exception:
    DateEntry = None
    TKCAL_AVAILABLE = False

try:
    import matplotlib
    # matplotlib.use('Agg')  # disabled to allow TkAgg
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_AVAILABLE = True
except Exception:
    MATPLOTLIB_AVAILABLE = False


try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as pdf_canvas
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# paths and DB
BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE_DIR, 'pharmacy.db')
BACKUP_FOLDER = os.path.join(BASE_DIR, 'backups')
os.makedirs(BACKUP_FOLDER, exist_ok=True)
RECEIPT_FOLDER = os.path.join(BASE_DIR, 'receipts')
os.makedirs(RECEIPT_FOLDER, exist_ok=True)


# -----------------------
# Utilities & DB Layer
# -----------------------
def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def now_str(): return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    


class DB:

    def __init__(self, path=DB_PATH):
        self.path = path
        self._ensure()

    def connect(self):
        con = sqlite3.connect(self.path, detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
        con.row_factory = sqlite3.Row
        con.execute('PRAGMA foreign_keys = ON;')
        return con

    def _ensure(self):
        con = self.connect(); cur = con.cursor()

        # users - FIXED: Added 'staff' to the CHECK constraint
        cur.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT CHECK(role IN ('admin','staff','cashier')) NOT NULL
        );''')
        cur.execute('SELECT COUNT(*) as c FROM users;')
        if cur.fetchone()['c'] == 0:
            cur.executemany('INSERT INTO users(username,password_hash,role) VALUES(?,?,?);', [
                ('admin', hash_pw('admin123'), 'admin'),
                ('cashier', hash_pw('cashier123'), 'cashier'),
            ])

        # customers
        cur.execute('''CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            phone TEXT UNIQUE,
            notes TEXT
        );''')

        # categories/manufacturers/suppliers/formulas
        cur.execute('''CREATE TABLE IF NOT EXISTS categories (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, notes TEXT);''')
        cur.execute('''CREATE TABLE IF NOT EXISTS manufacturers (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, contact TEXT, notes TEXT);''')
        cur.execute('''CREATE TABLE IF NOT EXISTS suppliers (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, phone TEXT, email TEXT, address TEXT);''')
        cur.execute('''CREATE TABLE IF NOT EXISTS formulas (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, composition TEXT);''')

        # products
        cur.execute('''CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            sku TEXT UNIQUE,
            is_medical INTEGER DEFAULT 1,
            category_id INTEGER,
            manufacturer_id INTEGER,
            formula_id INTEGER,
            unit TEXT,
            sale_price REAL DEFAULT 0,
            notes TEXT
        );''')

        # batches
        cur.execute('''CREATE TABLE IF NOT EXISTS batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            supplier_id INTEGER,
            batch_no TEXT,
            quantity INTEGER NOT NULL,
            expiry_date TEXT,
            cost_price REAL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
            FOREIGN KEY(supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL
        );''')

        # --- Ensure manufacturer_id exists in batches for existing DB ---
        cur.execute("PRAGMA table_info(batches);")
        columns = [c['name'] for c in cur.fetchall()]
        if 'manufacturer_id' not in columns:
            # Add manufacturer_id column
            cur.execute("ALTER TABLE batches ADD COLUMN manufacturer_id INTEGER;")
            # Update table with proper foreign key
            cur.execute("PRAGMA foreign_keys=off;")
            cur.execute("""CREATE TABLE IF NOT EXISTS batches_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                supplier_id INTEGER,
                manufacturer_id INTEGER,
                batch_no TEXT,
                quantity INTEGER NOT NULL,
                expiry_date TEXT,
                cost_price REAL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
                FOREIGN KEY(supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL,
                FOREIGN KEY(manufacturer_id) REFERENCES manufacturers(id) ON DELETE SET NULL
            );""")
            cur.execute("INSERT INTO batches_new SELECT id, product_id, supplier_id, NULL, batch_no, quantity, expiry_date, cost_price, created_at FROM batches;")
            cur.execute("DROP TABLE batches;")
            cur.execute("ALTER TABLE batches_new RENAME TO batches;")
            cur.execute("PRAGMA foreign_keys=on;")

        # sales & items
        cur.execute('''CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            total REAL NOT NULL,
            customer_id INTEGER,
            customer_name TEXT,
            customer_phone TEXT,
            discount REAL DEFAULT 0,
            tax REAL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY(customer_id) REFERENCES customers(id) ON DELETE SET NULL
        );''')
        cur.execute('''CREATE TABLE IF NOT EXISTS sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            price REAL NOT NULL,
            FOREIGN KEY(sale_id) REFERENCES sales(id) ON DELETE CASCADE,
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
        );''')
        cur.execute('''CREATE TABLE IF NOT EXISTS sale_item_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_item_id INTEGER NOT NULL,
            batch_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            FOREIGN KEY(sale_item_id) REFERENCES sale_items(id) ON DELETE CASCADE,
            FOREIGN KEY(batch_id) REFERENCES batches(id) ON DELETE CASCADE
        );''')
        cur.execute('''CREATE TABLE IF NOT EXISTS returns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_item_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            reason TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(sale_item_id) REFERENCES sale_items(id) ON DELETE CASCADE
        );''')

        # settings
        cur.execute('''CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);''')
        def set_if_missing(k,v):
            cur.execute('SELECT value FROM settings WHERE key=?;',(k,))
            if not cur.fetchone(): cur.execute('INSERT INTO settings(key,value) VALUES(?,?);',(k,str(v)))
        set_if_missing('tax_percent','0.0')
        set_if_missing('default_discount','0.0')
        set_if_missing('pharmacy_name','Pharmacy Receipt')
        set_if_missing('pharmacy_address','123 Main Street, City')
        set_if_missing('auto_backup_enabled','0')

        con.commit(); con.close()


    def query(self, sql, params=()):
        with self.connect() as con:
            cur = con.execute(sql, params)
            return [dict(r) for r in cur.fetchall()]

    def execute(self, sql, params=()):
        with self.connect() as con:
            cur = con.execute(sql, params)
            con.commit()
            return cur.lastrowid

db = DB()



class AutocompleteEntry(ttk.Entry):
    def __init__(self, master, suggestions_getter=None, width=30, **kwargs):
        super().__init__(master, width=width, **kwargs)
        self.suggestions_getter = suggestions_getter
        self.var = tk.StringVar()
        self.config(textvariable=self.var)
        self.var.trace_add('write', self._on_change)
        self.listbox = None
        self.bind('<Down>', self._focus_listbox)
        self.bind('<Escape>', lambda e: self._hide())
        self.bind('<Return>', self._select_first)
        style = ttk.Style()
        style.configure("Card.TFrame", background="white", relief="flat", borderwidth=1)
        style.configure("CardHover.TFrame", background="#f1f9ff", relief="flat", borderwidth=1)
        style.configure("CardTitle.TLabel", font=("Segoe UI", 11, "bold"), background="white", foreground="#343a40")
        style.configure("CardValue.TLabel", font=("Segoe UI", 26, "bold"), background="white")
        style.configure("CardSub.TLabel", font=("Segoe UI", 9), background="white", foreground="#6c757d")

    def _on_change(self, *args):
        term = self.var.get().strip()
        if not term:
            self._hide(); return
        try:
            suggestions = self.suggestions_getter(term) if self.suggestions_getter else []
        except Exception:
            suggestions = []
        if not suggestions:
            self._hide(); return
        if not self.listbox:
            self.listbox = tk.Listbox(self.master, height=6)
            self.listbox.bind('<<ListboxSelect>>', self._on_select)
            self.listbox.bind('<Return>', self._on_select)
            self.listbox.bind('<Escape>', lambda e: self._hide())
        self.listbox.delete(0, 'end')
        for s in suggestions:
            self.listbox.insert('end', s)
        x = self.winfo_rootx() - self.master.winfo_rootx()
        y = self.winfo_rooty() - self.master.winfo_rooty() + self.winfo_height()
        self.listbox.place(x=x, y=y, width=self.winfo_width())

    def _on_select(self, event=None):
        if not self.listbox: return
        sel = self.listbox.curselection()
        if not sel: return
        val = self.listbox.get(sel[0])
        self.var.set(val)
        self._hide()

    def _focus_listbox(self, event=None):
        if self.listbox:
            self.listbox.focus_set()
            self.listbox.selection_set(0)

    def _hide(self):
        if self.listbox:
            self.listbox.destroy()
            self.listbox = None

    def _select_first(self, event=None):
        if self.listbox and self.listbox.size() > 0:
            val = self.listbox.get(0)
            self.var.set(val)
            self._hide()


class PlaceholderEntry(ttk.Entry):
    def __init__(self, master, placeholder='', textvariable=None, width=30, **kw):
        if textvariable is None:
            self.var = tk.StringVar()
            kw['textvariable'] = self.var
        else:
            self.var = textvariable
            kw['textvariable'] = self.var
        super().__init__(master, width=width, **kw)
        self.placeholder = placeholder
        self._has_placeholder = False
        self._placeholder_color = '#6c757d'
        try:
            self._normal_color = self.cget('foreground') or 'black'
        except Exception:
            self._normal_color = 'black'
        self.bind('<FocusIn>', self._clear_placeholder)
        self.bind('<FocusOut>', self._set_placeholder)
        self._set_placeholder()

    def _set_placeholder(self, event=None):
        if not self.var.get():
            self._has_placeholder = True
            try:
                self.config(foreground=self._placeholder_color)
            except Exception:
                pass
            self.var.set(self.placeholder)

    def _clear_placeholder(self, event=None):
        if self._has_placeholder:
            self._has_placeholder = False
            try:
                self.config(foreground=self._normal_color)
            except Exception:
                pass
            self.var.set('')





class FormDialog(tk.Toplevel):
    def __init__(self, master, title, fields, on_submit=None, initial=None):
        super().__init__(master)
        self.title(title)
        self.on_submit = on_submit
        self.result = None
        self.widgets = {}

        frm = ttk.Frame(self, padding=8)
        frm.pack(fill='both', expand=True)

        for i, f in enumerate(fields):
            ttk.Label(frm, text=f.get('label', f['key'])).grid(row=i, column=0, sticky='w', pady=4)

            widget_type = f.get('widget', 'entry')
            if widget_type == 'autocomplete':
                w = AutocompleteEntry(frm, suggestions_getter=f.get('suggestions_getter'))
            elif widget_type == 'combobox':
                state = f.get('state', 'normal')
                w = ttk.Combobox(frm, values=f.get('values', []), state=state)
            elif widget_type == 'spinbox':
                w = ttk.Spinbox(frm, from_=f.get('from', 0), to=f.get('to', 999999), increment=f.get('inc', 1))
            elif widget_type == 'text':
                w = tk.Text(frm, height=f.get('height', 3), width=f.get('width', 40))
            else:
                w = ttk.Entry(frm)

            w.grid(row=i, column=1, sticky='we', pady=4)

            if initial and f['key'] in initial and initial[f['key']] is not None:
                val = initial[f['key']]
                if widget_type == 'text':
                    try:
                        w.insert('1.0', str(val))
                    except Exception:
                        pass
                else:
                    try:
                        w.insert(0, str(val))
                    except Exception:
                        pass

            self.widgets[f['key']] = (w, f)

        btns = ttk.Frame(frm)
        btns.grid(row=len(fields), column=0, columnspan=2, pady=8)
        ttk.Button(btns, text='Save', command=self._save).pack(side='left', padx=6)
        ttk.Button(btns, text='Cancel', command=self.destroy).pack(side='left')
        self.bind('<Return>', lambda e: self._save())
        self.bind('<Escape>', lambda e: self.destroy())

    def _save(self):
        data = {}
        for key, (w, f) in self.widgets.items():
            widget = f.get('widget', 'entry')
            if widget == 'text':
                try:
                    data[key] = w.get('1.0', 'end').strip()
                except Exception:
                    data[key] = ''
            else:
                try:
                    data[key] = w.get().strip()
                except Exception:
                    data[key] = ''
        self.result = data
        if self.on_submit:
            self.on_submit(data)
        self.destroy()

class NewSaleTab(ttk.Frame):
    def __init__(self, master, db, user):
        super().__init__(master)
        self.db, self.user = db, user
        self.cart = []
        self.selected_product = None
        self._build()
        self.app = app 
        
    def show_receipt_preview(self, sale_id, total, cust_name, cust_phone):
        import tkinter as tk
        from tkinter import messagebox

        # Generate receipt PDF first
        filepath = self.generate_receipt(sale_id, total, cust_name, cust_phone, preview_only=True)

        # Fetch items for showing as text
        items = self.db.query(
            "SELECT si.quantity, si.price, p.name FROM sale_items si "
            "JOIN products p ON si.product_id=p.id WHERE si.sale_id=?;",
            (sale_id,)
        )

        # ===== Preview Window =====
        win = tk.Toplevel(self)
        win.title(f"Receipt Preview - Sale {sale_id}")
        win.geometry("350x500")

        canvas = tk.Canvas(win, bg="white", width=330, height=420, bd=2, relief="groove")
        canvas.pack(pady=10)

        y = 20
        canvas.create_text(165, y, text="Pharmacy Receipt", font=("Arial", 12, "bold"))
        y += 20
        canvas.create_text(10, y, anchor="w", text=f"Sale ID: {sale_id}", font=("Arial", 9))
        y += 15
        canvas.create_text(10, y, anchor="w", text=f"Customer: {cust_name}", font=("Arial", 9))
        y += 15
        canvas.create_text(10, y, anchor="w", text=f"Phone: {cust_phone}", font=("Arial", 9))
        y += 20
        canvas.create_line(5, y, 325, y); y += 10
        canvas.create_text(10, y, anchor="w", text="Item", font=("Arial", 9, "bold"))
        canvas.create_text(200, y, anchor="w", text="Qty", font=("Arial", 9, "bold"))
        canvas.create_text(240, y, anchor="w", text="Price", font=("Arial", 9, "bold"))
        canvas.create_text(290, y, anchor="w", text="Total", font=("Arial", 9, "bold"))
        y += 15

        # Items
        for it in items:
            canvas.create_text(10, y, anchor="w", text=it["name"][:15], font=("Arial", 9))
            canvas.create_text(200, y, anchor="w", text=str(it["quantity"]), font=("Arial", 9))
            canvas.create_text(240, y, anchor="w", text=f"{it['price']:.2f}", font=("Arial", 9))
            canvas.create_text(290, y, anchor="w", text=f"{it['price']*it['quantity']:.2f}", font=("Arial", 9))
            y += 15

        y += 10
        canvas.create_line(5, y, 325, y); y += 20
        canvas.create_text(290, y, anchor="e", text=f"TOTAL: {total:.2f}", font=("Arial", 10, "bold"))
        y += 30
        canvas.create_text(165, y, text="Thank you for shopping!", font=("Arial", 9))

        # Buttons
        def do_print():
            try:
                os.startfile(filepath, "print")  # direct print
            except Exception as e:
                messagebox.showerror("Print Error", f"Could not print receipt:\n{e}")
            win.destroy()

        def do_save():
            messagebox.showinfo("Saved", f"Receipt saved at {filepath}")
            win.destroy()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="🖨️ Print Receipt", command=do_print,
                width=15, bg="lightgreen").grid(row=0, column=0, padx=5)
        tk.Button(btn_frame, text="💾 Save Receipt", command=do_save,
                width=15, bg="lightblue").grid(row=0, column=1, padx=5)


    def _build(self):
        cust = ttk.Frame(self); cust.pack(fill='x', padx=10, pady=6)
        ttk.Label(cust, text="Customer Name").pack(side='left')
        self.customer_name_e = ttk.Entry(cust, width=30); self.customer_name_e.pack(side='left', padx=6)
        ttk.Label(cust, text="Phone").pack(side='left')
        self.customer_phone_e = ttk.Entry(cust, width=20); self.customer_phone_e.pack(side='left', padx=6)

        top = ttk.Frame(self); top.pack(fill='x', padx=10, pady=6)
        ttk.Label(top, text="Search by name or ID").pack(side='left')
        self.search_e = AutocompleteEntry(top, suggestions_getter=lambda term: [f"{r['id']} - {r['name']} - {r['sale_price']}" for r in self.db.query("SELECT id,name,sale_price FROM products WHERE name LIKE ? OR CAST(id AS TEXT) LIKE ? ORDER BY name LIMIT 50;", (f"%{term}%",f"%{term}%"))]); self.search_e.pack(side='left', padx=6)
        # keep update_suggestions compatibility
        self.search_e.var.trace_add('write', lambda *a: self.update_suggestions())
        ttk.Label(top, text="Qty").pack(side='left', padx=(10,0))
        self.qty_e = ttk.Entry(top, width=6); self.qty_e.pack(side='left', padx=6)
        ttk.Button(top, text="Add", command=self.add_to_cart).pack(side='left')

        self.suggestions = tk.Listbox(self, height=6)
        self.suggestions.pack(fill='x', padx=10)
        self.suggestions.bind("<Double-Button-1>", self._on_suggestion_double)

        self.tree = ttk.Treeview(self, columns=['product','qty','price','subtotal'], show='headings')
        for c in ['product','qty','price','subtotal']:
            self.tree.heading(c, text=c.capitalize())
            self.tree.column(c, anchor='center')
        self.tree.pack(fill='both', expand=True, padx=10, pady=6)

        btns = ttk.Frame(self); btns.pack(fill='x', padx=10, pady=4)
        ttk.Button(btns, text="Remove Selected", command=self.remove_selected).pack(side='left')

        self.lbl_total = ttk.Label(self, text="Total: 0.00", font=('Segoe UI', 12, 'bold'))
        self.lbl_total.pack(anchor='e', padx=10)
        ttk.Button(self, text="Checkout", command=self.checkout).pack(anchor='e', padx=10, pady=6)

    def update_suggestions(self, event=None):
        term = self.search_e.get().strip()
        self.suggestions.delete(0, 'end')
        if not term:
            return
        rows = self.db.query("SELECT id, name, sale_price FROM products WHERE name LIKE ? OR CAST(id AS TEXT) LIKE ? ORDER BY name LIMIT 50;", (f"%{term}%", f"%{term}%"))
        for r in rows:
            self.suggestions.insert('end', f"{r['id']} - {r['name']} - {r['sale_price']}")

    def _on_suggestion_double(self, event=None):
        sel = self.suggestions.curselection()
        if not sel:
            return
        val = self.suggestions.get(sel[0])
        pid = int(val.split(' - ')[0])
        row = self.db.query("SELECT * FROM products WHERE id=?;", (pid,))
        if row:
            self.selected_product = row[0]
            self.search_e.delete(0, 'end')
            self.search_e.insert(0, f"{self.selected_product['name']}")

    def add_to_cart(self):
        term = self.search_e.get().strip()
        try:
            qty = int(self.qty_e.get() or 0)
        except ValueError:
            qty = 0
        prod = None
        if term.isdigit():
            rows = self.db.query("SELECT * FROM products WHERE id=?;", (int(term),))
            if rows: prod = rows[0]
        if not prod:
            rows = self.db.query("SELECT * FROM products WHERE name=? LIMIT 1;", (term,))
            if rows: prod = rows[0]
        if not prod:
            messagebox.showwarning("Not found", "Product not found. Use search box and double-click a suggestion.")
            return
        if qty <= 0:
            messagebox.showwarning("Invalid qty", "Enter quantity > 0")
            return
        self.cart.append({'id': prod['id'], 'name': prod['name'], 'qty': qty, 'price': prod['sale_price'], 'subtotal': prod['sale_price']*qty})
        self.search_e.delete(0, 'end'); self.qty_e.delete(0, 'end'); self.refresh()

    def remove_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.cart[idx]
        self.refresh()

    def refresh(self):
        self.tree.delete(*self.tree.get_children())
        total = 0
        for item in self.cart:
            self.tree.insert('', 'end', values=[item['name'], item['qty'], item['price'], item['subtotal']])
            total += item['subtotal']
        self.lbl_total.config(text=f"Total: {total:.2f}")
        
    def checkout(self):
        if not self.cart:
            messagebox.showwarning("Empty", "Cart is empty")
            return

        cust_name = self.customer_name_e.get().strip()
        cust_phone = self.customer_phone_e.get().strip()
        total = sum(i['subtotal'] for i in self.cart)

        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        sid = self.db.execute(
            "INSERT INTO sales(user_id,total,customer_name,customer_phone,created_at) VALUES(?,?,?,?,?);",
            (self.user['id'], total, cust_name, cust_phone, created_at)
        )

        for i in self.cart:
            sale_item_id = self.db.execute(
                "INSERT INTO sale_items(sale_id,product_id,quantity,price) VALUES(?,?,?,?);",
                (sid, i['id'], i['qty'], i['price'])
            )
            self._fifo_deduct_with_batch_tracking(i['id'], i['qty'], sale_item_id, i['name'])

        # 🚀 Auto save + open in Microsoft Edge
        filepath = self.generate_receipt(sid, total, cust_name, cust_phone, preview_only=True)
        try:
            os.startfile(filepath)  # opens with default PDF viewer (Edge, if set as default)
        except Exception as e:
            messagebox.showerror("Open Error", f"Could not open receipt:\n{e}")

        messagebox.showinfo("Sale Complete", f"Sale #{sid} completed.")
        self.cart.clear()
        self.refresh()





    # def checkout(self):
    #     cust_name = self.cust_name_entry.get().strip()
    #     cust_phone = self.cust_phone_entry.get().strip()
    #     total = sum(item['subtotal'] for item in self.cart)

    #     if not self.cart:
    #         messagebox.showwarning("Empty Cart", "No items in the cart")
    #         return

    #     # ===== Insert Sale =====
    #     self.db.execute("INSERT INTO sales (customer_name, customer_phone, total, user_id, created_at) VALUES (?,?,?,?,datetime('now'))",
    #                     (cust_name, cust_phone, total, self.app.user['id']))
    #     sale_id = self.db.lastrowid

    #     # ===== Insert Sale Items =====
    #     for item in self.cart:
    #         self.db.execute("INSERT INTO sale_items (sale_id, product_id, quantity, price) VALUES (?,?,?,?)",
    #                         (sale_id, item['id'], item['qty'], item['price']))

    #     self.db.commit()

    #     # ===== Generate Receipt File =====
    #     filepath = self.app.generate_receipt(sale_id, total, cust_name, cust_phone, preview_only=True)

    #     # ===== Preview Window =====
    #     from pdf2image import convert_from_path
    #     from PIL import ImageTk, Image

    #     win = tk.Toplevel(self)
    #     win.title("Receipt Preview")
    #     win.geometry("350x500")

    #     # Convert first page of PDF to image
    #     images = convert_from_path(filepath, dpi=150, first_page=1, last_page=1)
    #     img = images[0]
    #     img.thumbnail((320, 400))  # fit preview window
    #     tk_img = ImageTk.PhotoImage(img)

    #     lbl = tk.Label(win, image=tk_img)
    #     lbl.image = tk_img
    #     lbl.pack(pady=5)

    #     # Buttons
    #     def do_print():
    #         self.app.generate_receipt(sale_id, total, cust_name, cust_phone, direct_print=True)
    #         win.destroy()

    #     def do_save():
    #         messagebox.showinfo("Saved", f"Receipt saved at {filepath}")
    #         win.destroy()

    #     tk.Button(win, text="🖨️ Print Receipt", command=do_print, bg="lightgreen").pack(fill="x", pady=5, padx=20)
    #     tk.Button(win, text="💾 Save Receipt", command=do_save, bg="lightblue").pack(fill="x", pady=5, padx=20)

    #     # Clear cart after checkout
    #     self.cart.clear()
    #     self.refresh_cart()


    def _fifo_deduct_with_batch_tracking(self, product_id, qty_needed, sale_item_id, pname):
        batches = self.db.query(
            "SELECT id, quantity FROM batches WHERE product_id=? AND quantity>0 ORDER BY created_at ASC;", 
            (product_id,)
        )
        remain = qty_needed
        
        for b in batches:
            if remain <= 0:
                break
            take = min(remain, b['quantity'])
            
            # Update batch quantity
            self.db.execute("UPDATE batches SET quantity=quantity-? WHERE id=?;", (take, b['id']))
            
            # Create sale_item_batches entry
            self.db.execute(
                "INSERT INTO sale_item_batches(sale_item_id, batch_id, quantity) VALUES(?,?,?);",
                (sale_item_id, b['id'], take)
            )
            
            remain -= take
        
        if remain > 0:
            messagebox.showwarning("Stock Warning", f"Product {pname} had insufficient stock. Short by {remain}.")
    

   
    
    # def generate_receipt(self, sale_id, total, cust_name, cust_phone, direct_print=False):
    #     try:
    #         from reportlab.pdfgen import canvas
    #         from reportlab.lib.units import mm
    #     except Exception:
    #         messagebox.showerror("Missing Package", "reportlab not installed; cannot generate PDF.")
    #         return

    #     # Fetch pharmacy settings
    #     settings = {r['key']: r['value'] for r in self.db.query("SELECT key,value FROM settings;")}
    #     pharmacy_name = settings.get('pharmacy_name', 'Pharmacy Receipt')
    #     pharmacy_address = settings.get('pharmacy_address', '')
    #     pharmacy_phone = settings.get('pharmacy_phone', '')

    #     # Fetch sale items
    #     items = self.db.query("""
    #         SELECT si.quantity, si.price, p.name 
    #         FROM sale_items si 
    #         JOIN products p ON si.product_id=p.id 
    #         WHERE si.sale_id=?;
    #     """, (sale_id,))

    #     # Dynamic receipt size
    #     line_height = 12
    #     header_height = 120
    #     footer_height = 60
    #     receipt_height = header_height + footer_height + (len(items) * line_height)

    #     receipt_width = 80 * mm  # 80mm thermal paper
    #     folder = os.path.join(os.path.dirname(__file__), "receipts")
    #     os.makedirs(folder, exist_ok=True)
    #     filepath = os.path.join(folder, f"receipt_{sale_id}.pdf")

    #     c = canvas.Canvas(filepath, pagesize=(receipt_width, receipt_height))
    #     width, height = receipt_width, receipt_height
    #     y = height - 10

    #     # Header
    #     c.setFont("Helvetica-Bold", 12)
    #     c.drawCentredString(width / 2, y, pharmacy_name); y -= 14
    #     if pharmacy_address:
    #         c.setFont("Helvetica", 8)
    #         c.drawCentredString(width / 2, y, pharmacy_address); y -= 10
    #     if pharmacy_phone:
    #         c.drawCentredString(width / 2, y, f"Tel: {pharmacy_phone}"); y -= 10

    #     c.line(2, y, width-2, y); y -= 12

    #     # Sale Info
    #     c.setFont("Helvetica", 8)
    #     c.drawString(2, y, f"Sale ID: {sale_id}")
    #     c.drawRightString(width-2, y, datetime.now().strftime("%Y-%m-%d %H:%M")); y -= 10
    #     if cust_name:
    #         c.drawString(2, y, f"Customer: {cust_name}"); y -= 10
    #     if cust_phone:
    #         c.drawString(2, y, f"Phone: {cust_phone}"); y -= 10

    #     c.line(2, y, width-2, y); y -= 12

    #     # Table header
    #     c.setFont("Helvetica-Bold", 8)
    #     c.drawString(2, y, "Item")
    #     c.drawRightString(width-40, y, "Qty")
    #     c.drawRightString(width-20, y, "Price")
    #     c.drawRightString(width-2, y, "Sub")
    #     y -= 12

    #     # Items
    #     c.setFont("Helvetica", 8)
    #     for it in items:
    #         c.drawString(2, y, str(it['name'])[:15])  # truncate long names
    #         c.drawRightString(width-40, y, str(it['quantity']))
    #         c.drawRightString(width-20, y, f"{it['price']:.2f}")
    #         c.drawRightString(width-2, y, f"{it['price']*it['quantity']:.2f}")
    #         y -= line_height

    #     c.line(2, y, width-2, y); y -= 14

    #     # Total
    #     c.setFont("Helvetica-Bold", 10)
    #     c.drawRightString(width-2, y, f"TOTAL: {total:.2f}")
    #     y -= 20

    #     # Footer
    #     c.setFont("Helvetica-Oblique", 8)
    #     c.drawCentredString(width/2, y, "Thank you for shopping with us!"); y -= 10
    #     c.setFont("Helvetica", 6)
    #     c.drawCentredString(width/2, y, "Medicines once sold are not returnable.")

    #     c.save()

    #     # Direct print
    #     if direct_print:
    #         try:
    #             import win32print, win32api
    #             printer_name = win32print.GetDefaultPrinter()
    #             win32api.ShellExecute(0, "print", filepath, f'"{printer_name}"', ".", 0)
    #         except Exception as e:
    #             messagebox.showwarning("Print Error", f"Could not print directly:\n{e}")
    #     else:
    #         os.startfile(filepath)

    #     return filepath

    def generate_receipt(self, sale_id, total, cust_name, cust_phone, direct_print=False, preview_only=False):
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm

        settings = {r['key']: r['value'] for r in self.db.query("SELECT key,value FROM settings;")}
        pharmacy_name = settings.get('pharmacy_name', 'Pharmacy Receipt')
        pharmacy_address = settings.get('pharmacy_address', '')
        pharmacy_phone = settings.get('pharmacy_phone', '')

        items = self.db.query("""
            SELECT si.quantity, si.price, p.name 
            FROM sale_items si 
            JOIN products p ON si.product_id=p.id 
            WHERE si.sale_id=?;
        """, (sale_id,))

        # Dynamic receipt height
        line_height = 12
        header_height = 120
        footer_height = 60
        receipt_height = header_height + footer_height + (len(items) * line_height)
        receipt_width = 80 * mm

        folder = os.path.join(os.path.dirname(__file__), "receipts")
        os.makedirs(folder, exist_ok=True)
        filepath = os.path.join(folder, f"receipt_{sale_id}.pdf")

        c = canvas.Canvas(filepath, pagesize=(receipt_width, receipt_height))
        width, height = receipt_width, receipt_height
        y = height - 10

        # ===== Header =====
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(width / 2, y, pharmacy_name); y -= 14
        if pharmacy_address:
            c.setFont("Helvetica", 8)
            c.drawCentredString(width / 2, y, pharmacy_address); y -= 10
        if pharmacy_phone:
            c.drawCentredString(width / 2, y, f"Tel: {pharmacy_phone}"); y -= 10

        c.line(2, y, width-2, y); y -= 12

        # ===== Sale Info =====
        from datetime import datetime
        c.setFont("Helvetica", 8)
        c.drawString(2, y, f"Sale ID: {sale_id}")
        c.drawRightString(width-2, y, datetime.now().strftime("%Y-%m-%d %H:%M")); y -= 10
        if cust_name:
            c.drawString(2, y, f"Customer: {cust_name}"); y -= 10
        if cust_phone:
            c.drawString(2, y, f"Phone: {cust_phone}"); y -= 10

        c.line(2, y, width-2, y); y -= 12

        # ===== Table Header =====
        c.setFont("Helvetica-Bold", 8)
        c.drawString(2, y, "Item")
        c.drawRightString(width-40, y, "Qty")
        c.drawRightString(width-20, y, "Price")
        c.drawRightString(width-2, y, "Sub")
        y -= 12

        # ===== Items =====
        c.setFont("Helvetica", 8)
        for it in items:
            c.drawString(2, y, str(it['name'])[:15])  # truncate long names
            c.drawRightString(width-40, y, str(it['quantity']))
            c.drawRightString(width-20, y, f"{it['price']:.2f}")
            c.drawRightString(width-2, y, f"{it['price']*it['quantity']:.2f}")
            y -= line_height

        c.line(2, y, width-2, y); y -= 14

        # ===== Total =====
        c.setFont("Helvetica-Bold", 10)
        c.drawRightString(width-2, y, f"TOTAL: {total:.2f}")
        y -= 20

        # ===== Footer =====
        c.setFont("Helvetica-Oblique", 8)
        c.drawCentredString(width/2, y, "Thank you for shopping with us!"); y -= 10
        c.setFont("Helvetica", 6)
        c.drawCentredString(width/2, y, "Medicines once sold are not returnable.")

        c.save()

        if preview_only:
            return filepath

        if direct_print:
            try:
                import win32print, win32api
                printer_name = win32print.GetDefaultPrinter()
                win32api.ShellExecute(0, "print", filepath, f'"{printer_name}"', ".", 0)
            except Exception as e:
                messagebox.showwarning("Print Error", f"Could not print directly:\n{e}")
        else:
            os.startfile(filepath)

        return filepath




class App:
    def __init__(self):
        if TTB_AVAILABLE:
            self.root = tb.Window(themename='flatly')
        else:
            self.root = tk.Tk()
        
        self.root.title('Pharmacy Management System')
        self.root.geometry('1200x780')
        self.root.protocol('WM_DELETE_WINDOW', self._on_close)
        self.root.iconbitmap("icon.ico")

        self.db = db
        self.user = None
        self._auto_job = None
        self._build_login()

    def _on_close(self):
        try:
            if getattr(self, '_auto_job', None):
                self.root.after_cancel(self._auto_job)
        except Exception:
            pass
        self.root.quit()
        self.root.destroy()

    def _logout(self):
        """Log out and return to login screen."""
        self.user = None
        self._build_login()

    def _open_profile(self):
        def save(d):
            pw = d.get('new_password','').strip()
            if pw:
                self.db.execute(
                    'UPDATE users SET password_hash=? WHERE id=?;',
                    (hash_pw(pw), self.user['id'])
                )
                messagebox.showinfo('Profile','Password updated.')
        FormDialog(
            self.root, 'Profile - Change Password',
            [
                {'key':'username','label':'Username','widget':'entry'},
                {'key':'role','label':'Role','widget':'entry'},
                {'key':'new_password','label':'New Password','widget':'entry'},
            ],
            initial={'username':self.user['username'], 'role':self.user['role']},
            on_submit=save
        )
    # put these methods in App class (or another central class that both
# NewSaleTab.checkout and App._print_receipt can call)

    def _format_receipt_text(self, sale_id, total, cust_name, cust_phone, width=40):
        """
        Build a monospaced text receipt. width = number of characters per line.
        Returns bytes encoded for the printer (cp850 is common for thermal printers).
        """
        # load settings and items
        settings = {r['key']: r['value'] for r in self.db.query("SELECT key,value FROM settings;")}
        pharmacy_name = settings.get('pharmacy_name', 'Pharmacy')
        address = settings.get('pharmacy_address', '')
        phone = settings.get('pharmacy_phone', '')

        items = self.db.query(
            "SELECT si.quantity, si.price, p.name "
            "FROM sale_items si JOIN products p ON si.product_id=p.id WHERE si.sale_id=?;",
            (sale_id,)
        )

        def center(s):
            s = s.strip()
            if len(s) >= width:
                return s[:width]
            left = (width - len(s)) // 2
            return " " * left + s

        def right(s, w):
            s = str(s)
            if len(s) >= w:
                return s[-w:]
            return " " * (w - len(s)) + s

        lines = []
        lines.append(center(pharmacy_name))
        if address:
            lines.append(center(address))
        if phone:
            lines.append(center("Tel: " + phone))
        lines.append("-" * width)
        lines.append(f"Sale ID: {sale_id}")
        if cust_name:
            lines.append(f"Customer: {cust_name}")
        if cust_phone:
            lines.append(f"Phone: {cust_phone}")
        # date/time
        from datetime import datetime
        lines.append(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("-" * width)

        # Header for items - choose columns sizes
        # Example layout (width 40):
        # Item (20 chars) | QTY (3) | Price (7) | Subtotal (8) = 20+3+7+8 = 38 + spaces
        name_w = 20
        qty_w = 3
        price_w = 7
        sub_w = width - name_w - qty_w - price_w - 3  # -3 for spacing
        header = f"{'Item'.ljust(name_w)} {'Qty'.rjust(qty_w)} {'Price'.rjust(price_w)} {'Total'.rjust(sub_w)}"
        lines.append(header)
        lines.append("-" * width)

        for it in items:
            name = str(it['name'])[:name_w]
            qty = str(it['quantity'])
            price = f"{it['price']:.2f}"
            subtotal = f"{(it['price'] * it['quantity']):.2f}"
            line = f"{name.ljust(name_w)} {qty.rjust(qty_w)} {price.rjust(price_w)} {subtotal.rjust(sub_w)}"
            lines.append(line)

        lines.append("-" * width)
        tot_line = f"{'TOTAL'.ljust(width - 10)}{format(total, '.2f').rjust(10)}"
        lines.append(tot_line)
        lines.append("")
        lines.append(center("Thank you for shopping!"))
        lines.append(center("Get well soon!"))
        lines.append("\n\n")  # some feed for cutter

        out_text = "\r\n".join(lines)
        # use cp850 or cp437 to support common thermal printers; fallback to utf-8 if needed
        try:
            out_bytes = out_text.encode('cp850')
        except Exception:
            out_bytes = out_text.encode('utf-8', 'replace')
        return out_bytes


    def _send_bytes_to_printer(self, raw_bytes):
        """
        Send raw bytes to the default Windows printer using win32print.
        """
        try:
            import win32print
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Missing dependency", "pywin32 is required for direct printing: pip install pywin32")
            raise

        printer_name = win32print.GetDefaultPrinter()
        if not printer_name:
            raise RuntimeError("No default printer found.")

        hPrinter = win32print.OpenPrinter(printer_name)
        try:
            # Start a RAW print job
            hJob = win32print.StartDocPrinter(hPrinter, 1, ("Receipt", None, "RAW"))
            win32print.StartPagePrinter(hPrinter)
            win32print.WritePrinter(hPrinter, raw_bytes)
            win32print.EndPagePrinter(hPrinter)
            win32print.EndDocPrinter(hPrinter)
        finally:
            win32print.ClosePrinter(hPrinter)


    def print_receipt_direct(self, sale_id, total, cust_name="", cust_phone=""):
        """
        Save PDF (so history stays).
        If a default printer exists, send raw text to it.
        Otherwise, open PDF in Edge.
        """
        # 1) Always save the PDF (history record)
        filepath = self.generate_receipt(sale_id, total, cust_name, cust_phone, preview_only=True)

        # 2) Try direct print
        try:
            import win32print
            printer_name = win32print.GetDefaultPrinter()
            if printer_name:
                raw = self._format_receipt_text(sale_id, total, cust_name, cust_phone)
                self._send_bytes_to_printer(raw)
                return filepath
            else:
                raise RuntimeError("No default printer installed")
        except Exception:
            # 3) If no printer → just open PDF in Edge
            try:
                os.startfile(filepath)  # This will open in default PDF viewer (Edge usually)
            except Exception as e:
                from tkinter import messagebox
                messagebox.showerror("Open Error", f"Could not open receipt PDF:\n{e}")

        return filepath


        
    def generate_receipt(self, sale_id, total, cust_name="", cust_phone="", preview_only=False, direct_print=False):
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.lib.pagesizes import A4

        # Thermal-style width (~3 inch)
        RECEIPT_WIDTH = 250
        RECEIPT_HEIGHT = 600

        # Query sale + items
        items = self.db.query(
            "SELECT si.quantity, si.price, p.name FROM sale_items si "
            "JOIN products p ON si.product_id=p.id WHERE si.sale_id=?;",
            (sale_id,)
        )
        sale = self.db.query("SELECT * FROM sales WHERE id=?;", (sale_id,))
        if not sale:
            return None
        sale = sale[0]

        # Save folder
        folder = os.path.join(os.path.dirname(__file__), "receipts")
        os.makedirs(folder, exist_ok=True)
        filepath = os.path.join(folder, f"receipt_{sale_id}.pdf")

        # Create PDF canvas
        c = pdf_canvas.Canvas(filepath, pagesize=(RECEIPT_WIDTH, RECEIPT_HEIGHT))
        width, height = RECEIPT_WIDTH, RECEIPT_HEIGHT
        y = height - 30

        # Header
        c.setFont("Courier-Bold", 12)
        c.drawCentredString(width/2, y, "Pharmacy Receipt")
        y -= 20

        c.setFont("Courier", 8)
        c.drawString(5, y, f"Sale ID: {sale_id}")
        y -= 12
        if cust_name:
            c.drawString(5, y, f"Customer: {cust_name[:20]}")
            y -= 12
        if cust_phone:
            c.drawString(5, y, f"Phone: {cust_phone}")
            y -= 12
        c.drawString(5, y, f"Date: {sale['created_at']}")
        y -= 16

        # Column headers
        c.setFont("Courier-Bold", 8)
        c.drawString(5, y, "Item")
        c.drawString(120, y, "Qty")
        c.drawString(170, y, "Price")
        c.drawString(220, y, "Total")
        y -= 12
        c.setFont("Courier", 8)
        c.line(5, y+5, width-5, y+5)
        y -= 10

        # Items
        for it in items:
            name = str(it['name'])[:15]  # truncate name
            qty = str(it['quantity'])
            price = f"{it['price']:.2f}"
            subtotal = f"{it['price']*it['quantity']:.2f}"

            c.drawString(5, y, name)
            c.drawRightString(140, y, qty)
            c.drawRightString(190, y, price)
            c.drawRightString(width-10, y, subtotal)

            y -= 12
            if y < 50:
                c.showPage()
                y = height - 30
                c.setFont("Courier", 8)

        # Footer
        c.line(5, y, width-5, y)
        y -= 15
        c.setFont("Courier-Bold", 10)
        c.drawRightString(width-10, y, f"TOTAL: {total:.2f}")
        y -= 20
        c.setFont("Courier", 8)
        c.drawCentredString(width/2, y, "Thank you for shopping!")

        c.save()

        # Print directly if requested
        if direct_print:
            try:
                os.startfile(filepath, "print")
            except Exception as e:
                from tkinter import messagebox
                messagebox.showerror("Print Error", f"Could not print receipt:\n{e}")

        return filepath


    
    def _build_main(self):
        # Clear window
        for w in self.root.winfo_children():
            w.destroy()

        # Main container
        container = ttk.Frame(self.root)
        container.pack(fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(1, weight=1)  # main area expands

        # ---------------- SIDEBAR ----------------
        self.sidebar_expanded = True
        self.sidebar_width = 220

        sidebar = tk.Frame(container, bg="#2c3e50", width=self.sidebar_width)
        sidebar.grid(row=0, column=0, sticky="ns")
        sidebar.grid_propagate(False)
        sidebar.grid_rowconfigure(1, weight=1)  # push profile/logout to bottom

        # Sidebar Header
        header_frame = tk.Frame(sidebar, bg="#2c3e50")
        header_frame.grid(row=0, column=0, sticky="ew")

        self.brand_lbl = tk.Label(
            header_frame,
            text="🏥 Pharmacy",
            font=("Segoe UI", 15, "bold"),
            bg="#2c3e50",
            fg="white",
            pady=15
        )
        self.brand_lbl.pack(side="left", padx=10)

        toggle_btn = tk.Label(
            header_frame,
            text="☰",
            font=("Segoe UI", 14, "bold"),
            bg="#2c3e50",
            fg="white",
            cursor="hand2"
        )
        toggle_btn.pack(side="right", padx=10)

        # ---------------- NAVIGATION AREA ----------------
        nav_frame = tk.Frame(sidebar, bg="#2c3e50")
        nav_frame.grid(row=1, column=0, sticky="nsew")
        nav_items, nav_texts = [], []

        def add_nav_button(text, command, icon="📊"):
            full_text = f"{icon} {text}"
            btn = tk.Label(
                nav_frame,
                text=full_text,
                font=("Segoe UI", 11, "bold"),
                bg="#2c3e50",
                fg="white",
                anchor="w",
                padx=20,
                pady=12,
                cursor="hand2"
            )
            btn.pack(fill="x")

            def on_enter(e): btn.config(bg="#34495e")
            def on_leave(e):
                if btn not in active_btn:
                    btn.config(bg="#2c3e50")
            btn.bind("<Enter>", on_enter)
            btn.bind("<Leave>", on_leave)
            btn.bind("<Button-1>", lambda e: command())

            nav_items.append(btn)
            nav_texts.append(full_text)
            return btn

        # ---------------- PROFILE + LOGOUT ----------------
        bottom_frame = tk.Frame(sidebar, bg="#2c3e50")
        bottom_frame.grid(row=2, column=0, sticky="ew")

        # User Info
        self.user_lbl = tk.Label(
            bottom_frame,
            text=f"👤 {self.user['username'].title()} ({self.user['role']})",
            font=("Segoe UI", 10, "bold"),
            bg="#2c3e50",
            fg="white",
            anchor="w",
            padx=15,
            pady=10
        )
        self.user_lbl.pack(fill="x")

        # Logout button
        logout_btn = tk.Label(
            bottom_frame,
            text="🚪 Logout",
            font=("Segoe UI", 11, "bold"),
            bg="#2c3e50",
            fg="white",
            anchor="w",
            padx=20,
            pady=12,
            cursor="hand2"
        )
        logout_btn.pack(fill="x")

        def on_logout_enter(e): logout_btn.config(bg="#e74c3c")
        def on_logout_leave(e): logout_btn.config(bg="#2c3e50")
        logout_btn.bind("<Enter>", on_logout_enter)
        logout_btn.bind("<Leave>", on_logout_leave)
        logout_btn.bind("<Button-1>", lambda e: self._logout())

        # ---------------- TOGGLE SIDEBAR ----------------
        def toggle_sidebar(event=None):
            if self.sidebar_expanded:
                # Collapse
                self.sidebar_width = 60
                sidebar.config(width=self.sidebar_width)
                self.brand_lbl.config(text="🏥")
                self.user_lbl.config(text="👤")  # only icon
                logout_btn.config(text="🚪")
                for b in nav_items:
                    icon = b.cget("text").split(" ")[0]
                    b.config(text=icon)
            else:
                # Expand
                self.sidebar_width = 220
                sidebar.config(width=self.sidebar_width)
                self.brand_lbl.config(text="🏥 Pharmacy")
                self.user_lbl.config(text=f"👤 {self.user['username'].title()} ({self.user['role']})")
                logout_btn.config(text="🚪 Logout")
                for b, full_text in zip(nav_items, nav_texts):
                    b.config(text=full_text)
            self.sidebar_expanded = not self.sidebar_expanded

        toggle_btn.bind("<Button-1>", toggle_sidebar)

        # ---------------- MAIN CONTENT ----------------
        self.main_area = ttk.Frame(container)
        self.main_area.grid(row=0, column=1, sticky="nsew")

        # Content switcher
        def clear_main():
            for w in self.main_area.winfo_children():
                w.destroy()

        def show_dashboard():
            clear_main()
            self._build_dashboard_tab()

        def show_inventory():
            clear_main()
            self._build_inventory_tab()

        def show_pos():
            clear_main()
            self._build_pos_tab()

        def show_staff():
            clear_main()
            self._build_manage_staff_tab()

        def show_settings():
            clear_main()
            self._build_settings_tab()

        # Navigation buttons
   # Navigation buttons with PNG icons
        btn_dashboard = add_nav_button("Dashboard", show_dashboard, self.icon_dashboard)
        btn_inventory = add_nav_button("Inventory", show_inventory, self.icon_inventory)
        btn_pos = add_nav_button("POS", show_pos, self.icon_pos)

        if self.user["role"] == "admin":
            btn_staff = add_nav_button("Staff", show_staff, self.icon_staff)
            btn_settings = add_nav_button("Settings", show_settings, self.icon_settings)


        # Active highlight
        active_btn = []

        def activate(btn):
            for b in nav_items:
                b.config(bg="#2c3e50")
            btn.config(bg="#1abc9c")
            active_btn.clear()
            active_btn.append(btn)

        for btn in nav_items:
            btn.bind("<Button-1>", lambda e, b=btn: (activate(b), b.event_generate("<<CustomClick>>")))

        btn_dashboard.bind("<<CustomClick>>", lambda e: show_dashboard())
        btn_inventory.bind("<<CustomClick>>", lambda e: show_inventory())
        btn_pos.bind("<<CustomClick>>", lambda e: show_pos())
        if self.user["role"] == "admin":
            btn_staff.bind("<<CustomClick>>", lambda e: show_staff())
            btn_settings.bind("<<CustomClick>>", lambda e: show_settings())

        # Load dashboard by default
        activate(btn_dashboard)
        show_dashboard()
        
    def load_icon(path, size=(20, 20)):
        """Load and resize a PNG icon to fit tabs"""
        img = Image.open(path).resize(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    
        def resize_icons(self, event=None):
                w = self.root.winfo_width()
                size = (18, 18) if w < 800 else (24, 24)
                self.icon_dashboard = load_icon("icons/dashboard.png", size=size)
                self.icon_inventory = load_icon("icons/inventory.png", size=size)
                self.icon_pos = load_icon("icons/pos.png", size=size)
                self.icon_staff = load_icon("icons/staff.png", size=size)
                self.icon_settings = load_icon("icons/settings.png", size=size)

                # Update notebook tabs
                self.nb.tab(0, image=self.icon_dashboard)
                self.nb.tab(1, image=self.icon_inventory)
                self.nb.tab(2, image=self.icon_pos)
                if self.user["role"] == "admin":
                    self.nb.tab(3, image=self.icon_staff)
                    self.nb.tab(4, image=self.icon_settings)

            # Bind to window resize
                self.root.bind("<Configure>", self.resize_icons)
        
    def add_nav_button(text, command, icon_img):
        btn = tk.Label(
            nav_frame,
            text=f"  {text}",  # little padding for icon space
            image=icon_img,
            compound="left",  # place icon on the left of text
            font=("Segoe UI", 11, "bold"),
            bg="#2c3e50",
            fg="white",
            anchor="w",
            padx=15,
            pady=12,
            cursor="hand2"
        )
        btn.image = icon_img  # keep reference!
        btn.pack(fill="x")

        def on_enter(e): btn.config(bg="#34495e")
        def on_leave(e):
            if btn not in active_btn:
                btn.config(bg="#2c3e50")
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        btn.bind("<Button-1>", lambda e: command())

        nav_items.append(btn)
        nav_texts.append(text)
        return btn

    # ---------------- Staff Add Fix ----------------
    def _add_staff(self):
        def save(d):
            if not d.get('username') or not d.get('password'):
                return messagebox.showerror('Error','Username and password required')
            existing = self.db.query("SELECT id FROM users WHERE username=?;", (d['username'],))
            if existing:
                return messagebox.showerror('Error','Username already exists')
            self.db.execute(
                "INSERT INTO users(username,password_hash,role) VALUES(?,?,?)",
                (d['username'], hash_pw(d['password']), d['role'])
            )
            messagebox.showinfo('Saved','Staff added successfully')
            self._build_manage_staff_tab()

        FormDialog(self.root, 'Add Staff', [
            {'key':'username','label':'Username'},
            {'key':'password','label':'Password'},
            {'key':'role','label':'Role','widget':'combobox','values':['staff','cashier']}
        ], on_submit=save)


    # ---------------- Login ----------------
    def _build_login(self):
        for w in self.root.winfo_children(): w.destroy()
        frm = ttk.Frame(self.root, padding=20); frm.pack(expand=True)

        # Pharmacy Name + Logo
        top = ttk.Frame(frm); top.grid(row=0, column=0, columnspan=2, pady=(0,20))
        try:
            logo_img = tk.PhotoImage(file=os.path.join(BASE_DIR, "logo.png"))
            logo_lbl = ttk.Label(top, image=logo_img)
            logo_lbl.image = logo_img
            logo_lbl.pack()
        except Exception:
            ttk.Label(top, text='🏥', font=('Segoe UI', 40)).pack()

        ttk.Label(top, text='بایو فینِکس فارمیسی', font=('Nori Nastaleeq', 22, 'bold')).pack()
        ttk.Label(top, text='[-Login Portal-]', font=('Segoe UI', 9,'bold')).pack()
        from tkinter import font
    
        bold_font = font.Font(weight="bold")
        # Login Form
        ttk.Label(frm, text='Login As:', font=bold_font).grid(row=1, column=0, sticky='e')
        role_cb = ttk.Combobox(frm, values=['admin','staff','cashier'], state='readonly')
        role_cb.set('admin')
        role_cb.grid(row=1, column=1, sticky='w', pady=4)

        ttk.Label(frm, text='Username', font=bold_font).grid(row=2, column=0, sticky='e')
        user_e = ttk.Entry(frm); user_e.grid(row=2, column=1, sticky='w', pady=4)

        ttk.Label(frm, text='Password', font=bold_font).grid(row=3, column=0, sticky='e')
        pw_e = ttk.Entry(frm, show='•'); pw_e.grid(row=3, column=1, sticky='w', pady=4)


        def try_login():

            u = user_e.get().strip(); p = pw_e.get().strip(); r = role_cb.get().strip()
            if not u or not p: return messagebox.showerror('Error','Enter username & password')
            rows = self.db.query('SELECT * FROM users WHERE username=?;',(u,))
            if not rows or rows[0]['password_hash'] != hash_pw(p) or rows[0]['role'] != r:
                return messagebox.showerror('Error','Invalid credentials or role')
            self.user = {'id':rows[0]['id'],'username':u,'role':rows[0]['role']}
            self._build_main()
        ttk.Button(frm, text='Login', command=try_login).grid(row=4, column=0, columnspan=2, pady=8)
        user_e.focus_set()
        self.root.bind('<Return>', lambda e: try_login())

    # ---------------- Main ----------------
    def _build_main(self):
        # Clear
        for w in self.root.winfo_children():
            w.destroy()

        # Top bar
        top = ttk.Frame(self.root); top.pack(fill='x')
        ttk.Label(top, text=f"Welcome, {self.user['username'].title()}", font=('Segoe UI',14,'bold')).pack(side='left', padx=10, pady=8)
        ttk.Button(top, text='Profile', command=self._open_profile).pack(side='right', padx=6)
        ttk.Button(top, text='Logout', command=self._logout).pack(side='right')

        # Notebook (tabs)
        self.nb = ttk.Notebook(self.root); self.nb.pack(fill='both', expand=True, padx=8, pady=8)

        role = self.user.get('role')
        # Create tabs based on role
        if role == 'admin':
            self.tab_dashboard = ttk.Frame(self.nb); self.nb.add(self.tab_dashboard, text='Dashboard')
            self.tab_inventory = ttk.Frame(self.nb); self.nb.add(self.tab_inventory, text='Inventory')
            self.tab_pos = ttk.Frame(self.nb); self.nb.add(self.tab_pos, text='POS')
            self.tab_manage_staff = ttk.Frame(self.nb); self.nb.add(self.tab_manage_staff, text='Manage Staff')
            self.tab_settings = ttk.Frame(self.nb); self.nb.add(self.tab_settings, text='Settings')
        elif role == 'staff':
            self.tab_inventory = ttk.Frame(self.nb); self.nb.add(self.tab_inventory, text='Inventory')
            self.tab_pos = ttk.Frame(self.nb); self.nb.add(self.tab_pos, text='POS')
        elif role == 'cashier':
            self.tab_pos = ttk.Frame(self.nb); self.nb.add(self.tab_pos, text='POS')

        # Build tab contents
        if role == 'admin':
            self._build_dashboard_tab()
            self._build_inventory_tab()
            self._build_pos_tab()
            self._build_sale_history_tab()
            self._build_return_history_tab()
            self._build_manage_staff_tab()
            self._build_settings_tab()
        elif role == 'staff':
            self._build_inventory_tab()
            self._build_pos_tab()
        elif role == 'cashier':
            self._build_pos_tab()

        # ---------------- Dashboard ----------------
    def _build_dashboard_tab(self):
        """Magical, modern dashboard with animated stats, hover effects, charts, and alerts."""
        for w in self.tab_dashboard.winfo_children():
            w.destroy()

        main_container = ttk.Frame(self.tab_dashboard, padding=20)
        main_container.pack(fill="both", expand=True)

        # ---------------- HEADER ----------------
        header = ttk.Frame(main_container)
        header.pack(fill="x", pady=(0, 15))
        ttk.Label(
            header,
            text=f"✨ Pharmacy Dashboard — Welcome {self.user['username'].title()}",
            font=("Segoe UI", 22, "bold"),
            foreground="#2c3e50"
        ).pack(anchor="w")
        ttk.Label(
            header,
            text="Your pharmacy performance, reimagined",
            font=("Segoe UI", 12),
            foreground="#7f8c8d"
        ).pack(anchor="w")

        # ---------------- STATS CARDS ----------------
        stats_frame = ttk.Frame(main_container)
        stats_frame.pack(fill="x", pady=10)
        stats_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        # Fetch stats
        total_sales = int(self.db.query(
            "SELECT COALESCE(SUM(total),0) AS s FROM sales WHERE strftime('%Y-%m',created_at)=strftime('%Y-%m','now');"
        )[0]['s'])
        total_products = int(self.db.query("SELECT COUNT(*) AS c FROM products;")[0]['c'])
        near_expiry = int(self.db.query(
            "SELECT COUNT(*) AS c FROM batches WHERE expiry_date IS NOT NULL "
            "AND expiry_date <= date('now','+30 day') AND quantity > 0;"
        )[0]['c'])
        low_stock = int(self.db.query(
            "SELECT COUNT(*) AS c FROM ("
            " SELECT product_id, SUM(quantity) AS q FROM batches GROUP BY product_id HAVING q <= 5"
            ");"
        )[0]['c'])

        stats = [
            ("💰 Monthly Sales", total_sales, "#4e73df", lambda: self.nb.select(self.tab_pos)),
            ("📦 Products", total_products, "#1cc88a", lambda: self.nb.select(self.tab_inventory)),
            ("⏳ Near Expiry", near_expiry, "#f6c23e", lambda: self.nb.select(self.tab_inventory)),
            ("⚠️ Low Stock", low_stock, "#e74a3b", lambda: self.nb.select(self.tab_inventory)),
        ]

        def make_card(parent, title, value, color, onclick):
            card = tk.Frame(parent, bg=color, height=110, bd=0, relief="flat", cursor="hand2")
            card.pack_propagate(False)

            title_lbl = tk.Label(card, text=title, font=("Segoe UI", 12, "bold"),
                                bg=color, fg="white")
            title_lbl.pack(anchor="w", padx=10, pady=(10, 0))

            val_lbl = tk.Label(card, text="0", font=("Segoe UI", 28, "bold"),
                            bg=color, fg="white")
            val_lbl.pack(expand=True)

            # Animated counter
            def animate_number(lbl, target, delay=20):
                current = 0
                increment = max(1, target // 40)
                def update():
                    nonlocal current
                    if current < target:
                        current = min(current + increment, target)
                        lbl.config(text=f"{current:,}")
                        lbl.after(delay, update)
                    else:
                        lbl.config(text=f"{target:,}")
                update()
            animate_number(val_lbl, value)

            # Hover effects
            def on_enter(e):
                card.config(bg="#34495e")
                title_lbl.config(bg="#34495e")
                val_lbl.config(bg="#34495e")
            def on_leave(e):
                card.config(bg=color)
                title_lbl.config(bg=color)
                val_lbl.config(bg=color)
            card.bind("<Enter>", on_enter)
            card.bind("<Leave>", on_leave)

            # Click action
            card.bind("<Button-1>", lambda e: onclick())

            return card

        # Render cards in grid
        for i, (title, value, color, onclick) in enumerate(stats):
            card = make_card(stats_frame, title, value, color, onclick)
            card.grid(row=0, column=i, sticky="nsew", padx=8, pady=5)

        # ---------------- CHARTS ----------------
        charts_frame = ttk.Frame(main_container)
        charts_frame.pack(fill="both", expand=True, pady=15)
        charts_frame.grid_rowconfigure(0, weight=1)
        charts_frame.grid_columnconfigure((0, 1), weight=1)

        if MATPLOTLIB_AVAILABLE:
            import matplotlib.pyplot as plt
            plt.style.use("seaborn-v0_8")

            # Sales trend chart
            sales_fig = Figure(figsize=(5, 3), dpi=90)
            ax1 = sales_fig.add_subplot(111)
            days, totals = [], []
            for i in range(6, -1, -1):
                d = (datetime.now().date() - timedelta(days=i)).strftime('%Y-%m-%d')
                days.append(d[5:])
                r = self.db.query("SELECT COALESCE(SUM(total),0) AS s FROM sales WHERE date(created_at)=?;", (d,))
                totals.append(float(r[0]['s']))
            ax1.fill_between(days, totals, color="#4e73df", alpha=0.3)
            ax1.plot(days, totals, marker="o", color="#4e73df", linewidth=2)
            ax1.set_title("Sales - Last 7 Days")
            ax1.grid(True, linestyle="--", alpha=0.5)
            canvas1 = FigureCanvasTkAgg(sales_fig, master=charts_frame)
            canvas1.draw()
            canvas1.get_tk_widget().grid(row=0, column=0, sticky="nsew", padx=8)

            # Inventory chart
            inv_fig = Figure(figsize=(5, 3), dpi=90)
            ax2 = inv_fig.add_subplot(111)
            labels = ["Products", "Near Expiry", "Low Stock"]
            values = [total_products, near_expiry, low_stock]
            colors = ["#1cc88a", "#f6c23e", "#e74a3b"]
            bars = ax2.bar(labels, values, color=colors, alpha=0.9)
            for bar in bars:
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2, height + 0.3,
                        str(int(height)), ha="center", va="bottom", fontsize=9)
            ax2.set_title("Inventory Status")
            ax2.grid(True, linestyle="--", alpha=0.5)
            canvas2 = FigureCanvasTkAgg(inv_fig, master=charts_frame)
            canvas2.draw()
            canvas2.get_tk_widget().grid(row=0, column=1, sticky="nsew", padx=8)

        # ---------------- RECENT SALES & ALERTS ----------------
        bottom_frame = ttk.Frame(main_container)
        bottom_frame.pack(fill="both", expand=True, pady=10)
        bottom_frame.grid_columnconfigure((0, 1), weight=1)

        # Recent sales
        sales_frame = ttk.LabelFrame(bottom_frame, text="🕒 Recent Sales", padding=8)
        sales_frame.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)
        tree = ttk.Treeview(sales_frame, columns=("id", "customer", "total", "date"),
                            show="headings", height=6)
        for c in ("id", "customer", "total", "date"):
            tree.heading(c, text=c.capitalize())
        tree.column("id", width=50, anchor="center")
        tree.column("customer", width=150, anchor="w")
        tree.column("total", width=80, anchor="e")
        tree.column("date", width=120, anchor="center")
        tree.pack(fill="both", expand=True)
        rows = self.db.query(
            "SELECT id, COALESCE(customer_name,'-') AS customer, total, substr(created_at,1,16) AS date "
            "FROM sales ORDER BY id DESC LIMIT 6;"
        )
        for r in rows:
            tree.insert("", "end", values=(r['id'], r['customer'], f"${r['total']:.2f}", r['date']))

        # Alerts
        alerts_frame = ttk.LabelFrame(bottom_frame, text="⚠️ Alerts", padding=8)
        alerts_frame.grid(row=0, column=1, sticky="nsew", padx=6, pady=6)
        alert_tree = ttk.Treeview(alerts_frame, columns=("type", "item", "detail"), show="headings", height=6)
        for c in ("type", "item", "detail"):
            alert_tree.heading(c, text=c.capitalize())
        alert_tree.pack(fill="both", expand=True)
        low_rows = self.db.query(
            "SELECT p.name, COALESCE(SUM(b.quantity),0) AS qty "
            "FROM products p LEFT JOIN batches b ON b.product_id=p.id "
            "GROUP BY p.id HAVING qty <= 5 ORDER BY qty ASC LIMIT 4;"
        )
        for r in low_rows:
            alert_tree.insert("", "end", values=("Low Stock", r['name'], f"{r['qty']} units left"))
        exp_rows = self.db.query(
            "SELECT p.name, b.expiry_date FROM batches b "
            "JOIN products p ON p.id=b.product_id "
            "WHERE b.expiry_date IS NOT NULL "
            "AND b.expiry_date <= date('now','+30 day') "
            "AND b.quantity > 0 ORDER BY b.expiry_date ASC LIMIT 4;"
        )
        for r in exp_rows:
            alert_tree.insert("", "end", values=("Near Expiry", r['name'], f"Expires {r['expiry_date']}"))



    # ---------------- Inventory with nested tabs ----------------
    
    def _import_inventory(self, inv_type):
        path = filedialog.askopenfilename(filetypes=[('CSV Files','*.csv'),('Excel Files','*.xlsx')])
        if not path:
            return
        rows = []
        try:
            if path.lower().endswith('.csv'):
                with open(path, newline='', encoding='utf-8') as f:
                    rows = list(csv.DictReader(f))
            elif path.lower().endswith('.xlsx') and OPENPYXL_AVAILABLE:
                from openpyxl import load_workbook
                wb = load_workbook(path); ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                messagebox.showerror('Error','Unsupported file type')
                return

            for r in rows:
                if inv_type in ('medical','nonmedical'):
                    self.db.execute(
                        """INSERT OR IGNORE INTO products(name,is_medical,sku,unit,sale_price)
                        VALUES(?,?,?,?,?)""",
                        (r.get('name'), 1 if inv_type=='medical' else 0, r.get('sku'), r.get('unit'), float(r.get('price') or 0))
                    )
                elif inv_type == 'suppliers':
                    self.db.execute("INSERT OR IGNORE INTO suppliers(name,phone,email,address) VALUES(?,?,?,?)",
                        (r.get('name'), r.get('phone'), r.get('email'), r.get('address')))
                elif inv_type == 'manufacturers':
                    self.db.execute("INSERT OR IGNORE INTO manufacturers(name,contact,notes) VALUES(?,?,?)",
                        (r.get('name'), r.get('contact'), r.get('notes')))
                elif inv_type == 'categories':
                    self.db.execute("INSERT OR IGNORE INTO categories(name,notes) VALUES(?,?)",
                        (r.get('name'), r.get('notes')))
                elif inv_type == 'formulas':
                    self.db.execute("INSERT OR IGNORE INTO formulas(name,composition) VALUES(?,?)",
                        (r.get('name'), r.get('composition')))
                elif inv_type == 'batches':
                    pid = self.db.query("SELECT id FROM products WHERE name=?", (r.get('product'),))
                    sid = self.db.query("SELECT id FROM suppliers WHERE name=?", (r.get('supplier'),))
                    pid = pid[0]['id'] if pid else None
                    sid = sid[0]['id'] if sid else None
                    if pid:
                        self.db.execute(
                            """INSERT INTO batches(product_id,supplier_id,batch_no,quantity,expiry_date,cost_price,created_at)
                            VALUES(?,?,?,?,?,?,?)""",
                            (pid, sid, r.get('batch_no'), int(r.get('quantity') or 0), r.get('expiry'), float(r.get('cost_price') or 0), now_str())
                        )

            messagebox.showinfo('Import','Data imported successfully!')
            try:
                self._inv_refresh_all()
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror('Import Error', str(e))
    
    def _build_inventory_tab(self):
        for w in self.tab_inventory.winfo_children(): w.destroy()
        frame = self.tab_inventory
        ttk.Label(frame, text='Inventory', font=('Segoe UI',14,'bold')).pack(anchor='w', padx=10, pady=(6,0))
        inv_nb = ttk.Notebook(frame); inv_nb.pack(fill='both', expand=True, padx=8, pady=8)

        med_tab = ttk.Frame(inv_nb); inv_nb.add(med_tab, text='Medical Products')
        # search frame for Medical Products
        med_tab_search_fr = ttk.Frame(med_tab)
        med_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(med_tab_search_fr, text='🔍').pack(side='left')
        self.med_tab_search_var = tk.StringVar()
        med_tab_search_entry = PlaceholderEntry(med_tab_search_fr, placeholder='Search...', textvariable=self.med_tab_search_var, width=36)
        med_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.med_tab_search_var.trace_add('write', lambda *a, t='Medical Products': self._filter_tab(t, self.med_tab_search_var.get()))

        nonmed_tab = ttk.Frame(inv_nb); inv_nb.add(nonmed_tab, text='Non-Medical Products')
        # search frame for Non-Medical Products
        nonmed_tab_search_fr = ttk.Frame(nonmed_tab)
        nonmed_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(nonmed_tab_search_fr, text='🔍').pack(side='left')
        self.nonmed_tab_search_var = tk.StringVar()
        nonmed_tab_search_entry = PlaceholderEntry(nonmed_tab_search_fr, placeholder='Search...', textvariable=self.nonmed_tab_search_var, width=36)
        nonmed_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.nonmed_tab_search_var.trace_add('write', lambda *a, t='Non-Medical Products': self._filter_tab(t, self.nonmed_tab_search_var.get()))

        suppliers_tab = ttk.Frame(inv_nb); inv_nb.add(suppliers_tab, text='Suppliers')
        # search frame for Suppliers
        suppliers_tab_search_fr = ttk.Frame(suppliers_tab)
        suppliers_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(suppliers_tab_search_fr, text='🔍').pack(side='left')
        self.suppliers_tab_search_var = tk.StringVar()
        suppliers_tab_search_entry = PlaceholderEntry(suppliers_tab_search_fr, placeholder='Search...', textvariable=self.suppliers_tab_search_var, width=36)
        suppliers_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.suppliers_tab_search_var.trace_add('write', lambda *a, t='Suppliers': self._filter_tab(t, self.suppliers_tab_search_var.get()))

        manufacturers_tab = ttk.Frame(inv_nb); inv_nb.add(manufacturers_tab, text='Manufacturers')
        # search frame for Manufacturers
        manufacturers_tab_search_fr = ttk.Frame(manufacturers_tab)
        manufacturers_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(manufacturers_tab_search_fr, text='🔍').pack(side='left')
        self.manufacturers_tab_search_var = tk.StringVar()
        manufacturers_tab_search_entry = PlaceholderEntry(manufacturers_tab_search_fr, placeholder='Search...', textvariable=self.manufacturers_tab_search_var, width=36)
        manufacturers_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.manufacturers_tab_search_var.trace_add('write', lambda *a, t='Manufacturers': self._filter_tab(t, self.manufacturers_tab_search_var.get()))

        categories_tab = ttk.Frame(inv_nb); inv_nb.add(categories_tab, text='Categories')
        # search frame for Categories
        categories_tab_search_fr = ttk.Frame(categories_tab)
        categories_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(categories_tab_search_fr, text='🔍').pack(side='left')
        self.categories_tab_search_var = tk.StringVar()
        categories_tab_search_entry = PlaceholderEntry(categories_tab_search_fr, placeholder='Search...', textvariable=self.categories_tab_search_var, width=36)
        categories_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.categories_tab_search_var.trace_add('write', lambda *a, t='Categories': self._filter_tab(t, self.categories_tab_search_var.get()))

        formulas_tab = ttk.Frame(inv_nb); inv_nb.add(formulas_tab, text='Formulas')
        # search frame for Formulas
        formulas_tab_search_fr = ttk.Frame(formulas_tab)
        formulas_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(formulas_tab_search_fr, text='🔍').pack(side='left')
        self.formulas_tab_search_var = tk.StringVar()
        formulas_tab_search_entry = PlaceholderEntry(formulas_tab_search_fr, placeholder='Search...', textvariable=self.formulas_tab_search_var, width=36)
        formulas_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.formulas_tab_search_var.trace_add('write', lambda *a, t='Formulas': self._filter_tab(t, self.formulas_tab_search_var.get()))

        batches_tab = ttk.Frame(inv_nb); inv_nb.add(batches_tab, text='Batches')
        # search frame for Batches
        batches_tab_search_fr = ttk.Frame(batches_tab)
        batches_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(batches_tab_search_fr, text='🔍').pack(side='left')
        self.batches_tab_search_var = tk.StringVar()
        batches_tab_search_entry = PlaceholderEntry(batches_tab_search_fr, placeholder='Search...', textvariable=self.batches_tab_search_var, width=36)
        batches_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.batches_tab_search_var.trace_add('write', lambda *a, t='Batches': self._filter_tab(t, self.batches_tab_search_var.get()))


        cols = ('id','name','sku','unit','category','manufacturer','price','stock')

        def make_prod_tree(parent):
            tree = ttk.Treeview(parent, columns=cols, show='headings', height=18)
            # headings
            for c in cols:
                tree.heading(c, text=c.capitalize())
            # column alignments & widths
            tree.column('id', width=60, anchor='center')
            tree.column('name', width=180, anchor='w')
            tree.column('sku', width=100, anchor='center')
            tree.column('unit', width=100, anchor='center')
            tree.column('category', width=140, anchor='w')
            tree.column('manufacturer', width=140, anchor='w')
            tree.column('price', width=100, anchor='e')
            tree.column('stock', width=80, anchor='center')
            tree.pack(fill='both', expand=True, padx=8, pady=8)
            return tree

        self._med_tree = make_prod_tree(med_tab)
        btns_med = ttk.Frame(med_tab); btns_med.pack(fill='x', padx=8)
        ttk.Button(btns_med, text='Add', command=lambda: self._inv_add_product(is_medical=1)).pack(side='left', padx=6)
        ttk.Button(btns_med, text='Edit', command=lambda: self._inv_edit_product(self._med_tree)).pack(side='left', padx=6)
        ttk.Button(btns_med, text='Delete', command=lambda: self._inv_delete_product(self._med_tree)).pack(side='left', padx=6)
        ttk.Button(btns_med, text='Add by CSV/Excel', command=lambda: self._import_inventory('medical')).pack(side='left', padx=6)
        ttk.Button(btns_med, text='Clear Filter', command=self._inv_refresh_all).pack(side='right', padx=6)

        self._nonmed_tree = make_prod_tree(nonmed_tab)
        btns_non = ttk.Frame(nonmed_tab); btns_non.pack(fill='x', padx=8)
        ttk.Button(btns_non, text='Add', command=lambda: self._inv_add_product(is_medical=0)).pack(side='left', padx=6)
        ttk.Button(btns_non, text='Edit', command=lambda: self._inv_edit_product(self._nonmed_tree)).pack(side='left', padx=6)
        ttk.Button(btns_non, text='Delete', command=lambda: self._inv_delete_product(self._nonmed_tree)).pack(side='left', padx=6)
        ttk.Button(btns_non, text='Add by CSV/Excel', command=lambda: self._import_inventory('nonmedical')).pack(side='left', padx=6)
        ttk.Button(btns_non, text='Clear Filter', command=self._inv_refresh_all).pack(side='right', padx=6)

        self._sup_tree = ttk.Treeview(suppliers_tab, columns=('id','name','phone','email','address'), show='headings')
        for c in ('id','name','phone','email','address'):
            self._sup_tree.heading(c, text=c.capitalize())
        self._sup_tree.column('id', width=60, anchor='center')
        self._sup_tree.column('name', width=180, anchor='w')
        self._sup_tree.column('phone', width=120, anchor='center')
        self._sup_tree.column('email', width=180, anchor='w')
        self._sup_tree.column('address', width=240, anchor='w')
        self._sup_tree.pack(fill='both', expand=True, padx=8, pady=8)
        sup_btns = ttk.Frame(suppliers_tab); sup_btns.pack(fill='x', padx=8)
        ttk.Button(sup_btns, text='Add', command=self._add_supplier).pack(side='left', padx=6)
        ttk.Button(sup_btns, text='Edit', command=self._edit_supplier).pack(side='left', padx=6)
        ttk.Button(sup_btns, text='Delete', command=self._delete_supplier).pack(side='left', padx=6)
        ttk.Button(sup_btns, text='Add by CSV/Excel', command=lambda: self._import_inventory('suppliers')).pack(side='left', padx=6)

        self._man_tree = ttk.Treeview(manufacturers_tab, columns=('id','name','contact','notes'), show='headings')
        for c in ('id','name','contact','notes'):
            self._man_tree.heading(c, text=c.capitalize())
        self._man_tree.column('id', width=60, anchor='center')
        self._man_tree.column('name', width=180, anchor='w')
        self._man_tree.column('contact', width=160, anchor='w')
        self._man_tree.column('notes', width=240, anchor='w')
        self._man_tree.pack(fill='both', expand=True, padx=8, pady=8)
        man_btns = ttk.Frame(manufacturers_tab); man_btns.pack(fill='x', padx=8)
        ttk.Button(man_btns, text='Add', command=self._add_manufacturer).pack(side='left', padx=6)
        ttk.Button(man_btns, text='Edit', command=self._edit_manufacturer).pack(side='left', padx=6)
        ttk.Button(man_btns, text='Delete', command=self._delete_manufacturer).pack(side='left', padx=6)
        ttk.Button(man_btns, text='Add by CSV/Excel', command=lambda: self._import_inventory('manufacturers')).pack(side='left', padx=6)

        self._cat_tree = ttk.Treeview(categories_tab, columns=('id','name','notes'), show='headings')
        for c in ('id','name','notes'):
            self._cat_tree.heading(c, text=c.capitalize())
        self._cat_tree.column('id', width=60, anchor='center')
        self._cat_tree.column('name', width=200, anchor='w')
        self._cat_tree.column('notes', width=300, anchor='w')
        self._cat_tree.pack(fill='both', expand=True, padx=8, pady=8)
        cat_btns = ttk.Frame(categories_tab); cat_btns.pack(fill='x', padx=8)
        ttk.Button(cat_btns, text='Add', command=self._add_category).pack(side='left', padx=6)
        ttk.Button(cat_btns, text='Edit', command=self._edit_category).pack(side='left', padx=6)
        ttk.Button(cat_btns, text='Delete', command=self._delete_category).pack(side='left', padx=6)
        ttk.Button(cat_btns, text='Add by CSV/Excel', command=lambda: self._import_inventory('categories')).pack(side='left', padx=6)

        self._form_tree = ttk.Treeview(formulas_tab, columns=('id','name','composition'), show='headings')
        for c in ('id','name','composition'):
            self._form_tree.heading(c, text=c.capitalize())
        self._form_tree.column('id', width=60, anchor='center')
        self._form_tree.column('name', width=200, anchor='w')
        self._form_tree.column('composition', width=320, anchor='w')
        self._form_tree.pack(fill='both', expand=True, padx=8, pady=8)
        form_btns = ttk.Frame(formulas_tab); form_btns.pack(fill='x', padx=8)
        ttk.Button(form_btns, text='Add', command=self._add_formula).pack(side='left', padx=6)
        ttk.Button(form_btns, text='Edit', command=self._edit_formula).pack(side='left', padx=6)
        ttk.Button(form_btns, text='Delete', command=self._delete_formula).pack(side='left', padx=6)
        ttk.Button(form_btns, text='Add by CSV/Excel', command=lambda: self._import_inventory('formulas')).pack(side='left', padx=6)

        self._batch_tree = ttk.Treeview(batches_tab, columns=('id','product','batch_no','quantity','expiry','supplier'), show='headings')
        for c in ('id','product','batch_no','quantity','expiry','supplier'): self._batch_tree.heading(c, text=c.capitalize()); self._batch_tree.column(c, width=140, anchor='w')
        self._batch_tree.pack(fill='both', expand=True, padx=8, pady=8)
        batch_btns = ttk.Frame(batches_tab); batch_btns.pack(fill='x', padx=8)
        ttk.Button(batch_btns, text='Add', command=self._add_batch).pack(side='left', padx=6)
        ttk.Button(batch_btns, text='Edit', command=self._edit_batch).pack(side='left', padx=6)
        ttk.Button(batch_btns, text='Delete', command=self._delete_batch).pack(side='left', padx=6)
        ttk.Button(batch_btns, text='Clear Filter', command=self._inv_refresh_all).pack(side='right', padx=6)

        self._inv_refresh_all()

    
    def _filter_tab(self, tab_name, term):
        try:
            mapping = {
                'Medical Products': ('_med_tree', self._get_medical_data),
                'Non-Medical Products': ('_nonmed_tree', self._get_nonmedical_data),
                'Suppliers': ('_sup_tree', self._get_suppliers_data),
                'Manufacturers': ('_man_tree', self._get_manufacturers_data),
                'Categories': ('_cat_tree', self._get_categories_data),
                'Formulas': ('_form_tree', self._get_formulas_data),
                'Batches': ('_batch_tree', self._get_batches_data)
            }
            
            term = (term or '').strip().lower()
            tree_attr, data_getter = mapping.get(tab_name, (None, None))
            
            if not tree_attr or not data_getter:
                return
                
            tree = getattr(self, tree_attr, None)
            if not tree:
                return
            
            # Clear the tree
            tree.delete(*tree.get_children())
            
            # If empty term, load all data
            if not term:
                data = data_getter()
                for item_data in data:
                    tree.insert('', 'end', values=item_data)
                return
            
            # Filter: show only matching items
            data = data_getter()
            for item_data in data:
                if any(term in str(v).lower() for v in item_data):
                    tree.insert('', 'end', values=item_data)
                    
        except Exception as e:
            print(f"Filter error in {tab_name}: {e}")

    def _get_medical_data(self):
        rows = self.db.query('''SELECT p.id,p.name,p.sku,p.unit,c.name as category,m.name as manufacturer,p.sale_price as price,
            COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p
            LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id WHERE p.is_medical=1 ORDER BY p.name;''')
        return [(r['id'], r['name'], r['sku'] or '', r.get('unit','') or '', r.get('category') or '', r.get('manufacturer') or '', f"{r['price']:.2f}", r['stock']) for r in rows]

    def _get_nonmedical_data(self):
        rows = self.db.query('''SELECT p.id,p.name,p.sku,p.unit,c.name as category,m.name as manufacturer,p.sale_price as price,
            COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p
            LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id WHERE p.is_medical=0 ORDER BY p.name;''')
        return [(r['id'], r['name'], r['sku'] or '', r.get('unit','') or '', r.get('category') or '', r.get('manufacturer') or '', f"{r['price']:.2f}", r['stock']) for r in rows]

    def _get_suppliers_data(self):
        rows = self.db.query('SELECT id,name,phone,email,address FROM suppliers ORDER BY name;')
        return [(r['id'], r['name'], r['phone'] or '', r['email'] or '', r['address'] or '') for r in rows]

    def _get_manufacturers_data(self):
        rows = self.db.query('SELECT id,name,contact,notes FROM manufacturers ORDER BY name;')
        return [(r['id'], r['name'], r['contact'] or '', r['notes'] or '') for r in rows]

    def _get_categories_data(self):
        rows = self.db.query('SELECT id,name,notes FROM categories ORDER BY name;')
        return [(r['id'], r['name'], r['notes'] or '') for r in rows]

    def _get_formulas_data(self):
        rows = self.db.query('SELECT id,name,composition FROM formulas ORDER BY name;')
        return [(r['id'], r['name'], r['composition'] or '') for r in rows]

    def _get_batches_data(self):
        rows = self.db.query('SELECT b.id, p.name as product, b.batch_no, b.quantity, b.expiry_date, s.name as supplier FROM batches b LEFT JOIN products p ON p.id=b.product_id LEFT JOIN suppliers s ON s.id=b.supplier_id ORDER BY b.id DESC;')
        return [(r['id'], r['product'], r['batch_no'] or '', r['quantity'], r['expiry_date'] or '', r['supplier'] or '') for r in rows]
    
    def _inv_refresh_all(self):
        # Medical Products
        self._med_tree.delete(*self._med_tree.get_children())
        for item_data in self._get_medical_data():
            self._med_tree.insert('', 'end', iid=item_data[0], values=item_data)

        # Non-Medical Products
        self._nonmed_tree.delete(*self._nonmed_tree.get_children())
        for item_data in self._get_nonmedical_data():
            self._nonmed_tree.insert('', 'end', iid=item_data[0], values=item_data)

        # Suppliers
        self._sup_tree.delete(*self._sup_tree.get_children())
        for item_data in self._get_suppliers_data():
            self._sup_tree.insert('', 'end', values=item_data)

        # Manufacturers
        self._man_tree.delete(*self._man_tree.get_children())
        for item_data in self._get_manufacturers_data():
            self._man_tree.insert('', 'end', values=item_data)

        # Categories
        self._cat_tree.delete(*self._cat_tree.get_children())
        for item_data in self._get_categories_data():
            self._cat_tree.insert('', 'end', values=item_data)

        # Formulas
        self._form_tree.delete(*self._form_tree.get_children())
        for item_data in self._get_formulas_data():
            self._form_tree.insert('', 'end', values=item_data)

        # Batches
        self._batch_tree.delete(*self._batch_tree.get_children())
        for item_data in self._get_batches_data():
            self._batch_tree.insert('', 'end', values=item_data)
            
    def _inv_add_product(self, is_medical=1):
        cats = [r['name'] for r in self.db.query('SELECT name FROM categories ORDER BY name;')]
        mans = [r['name'] for r in self.db.query('SELECT name FROM manufacturers ORDER BY name;')]
        forms = [r['name'] for r in self.db.query('SELECT name FROM formulas ORDER BY name;')]
        units = ['mg','ml','g','IU','tablet','capsule','bottle','strip','box']

        def save(d):
            if not d.get('name'):
                return messagebox.showerror('Error','Name required')
            cid = mid = fid = None
            if d.get('category'):
                row = self.db.query('SELECT id FROM categories WHERE name=?;',(d['category'],))
                if row: cid = row[0]['id']
                else: cid = self.db.execute('INSERT INTO categories(name) VALUES(?);',(d['category'],))
            if d.get('manufacturer'):
                row = self.db.query('SELECT id FROM manufacturers WHERE name=?;',(d['manufacturer'],))
                if row: mid = row[0]['id']
                else: mid = self.db.execute('INSERT INTO manufacturers(name) VALUES(?);',(d['manufacturer'],))
            if d.get('formula'):
                row = self.db.query('SELECT id FROM formulas WHERE name=?;',(d['formula'],))
                if row: fid = row[0]['id']
                else: fid = self.db.execute('INSERT INTO formulas(name) VALUES(?);',(d['formula'],))
            try:
                self.db.execute('INSERT INTO products(name,sku,is_medical,category_id,manufacturer_id,formula_id,unit,sale_price,notes) VALUES(?,?,?,?,?,?,?,?,?);',
                                (d.get('name'), d.get('sku') or None, 1 if is_medical else 0, cid, mid, fid, d.get('unit') or '', float(d.get('price') or 0), d.get('notes') or ''))
                messagebox.showinfo('Saved','Product added'); self._inv_refresh_all()
            except Exception as e:
                messagebox.showerror('Error', str(e))

        fields = [
            {'key':'name','label':'Name'},
            {'key':'sku','label':'SKU'},
            {'key':'unit','label':'Unit','widget':'combobox','values':units,'state':'normal'},
            {'key':'category','label':'Category','widget':'combobox','values':cats,'state':'normal'},
            {'key':'manufacturer','label':'Manufacturer','widget':'combobox','values':mans,'state':'normal'},
            {'key':'formula','label':'Formula','widget':'combobox','values':forms,'state':'normal'},
            {'key':'price','label':'Sale Price'},
            {'key':'notes','label':'Notes','widget':'text'}
        ]

        dlg = FormDialog(self.root, 'Add Product', fields, on_submit=save)

        # Make comboboxes searchable
        self.make_searchable(dlg.widgets['unit'][0], units)
        self.make_searchable(dlg.widgets['category'][0], cats)
        self.make_searchable(dlg.widgets['manufacturer'][0], mans)
        self.make_searchable(dlg.widgets['formula'][0], forms)


   
    def _inv_edit_product(self, tree):
        sel = tree.selection()
        if not sel: return messagebox.showwarning('Select','Select a product to edit')
        values = tree.item(sel[0], 'values')
        pid = int(values[0])
        row = self.db.query('SELECT p.*, c.name as category_name, m.name as manufacturer_name, f.name as formula_name FROM products p LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id LEFT JOIN formulas f ON p.formula_id=f.id WHERE p.id=?;',(pid,))
        if not row: return messagebox.showerror('Error','Not found')
        row = row[0]
        initial = {'name':row['name'],'sku':row['sku'] or '','unit':row['unit'] or '','category':row.get('category_name') or '','manufacturer':row.get('manufacturer_name') or '','formula':row.get('formula_name') or '','price':row['sale_price'],'notes':row['notes']}
        
        cats = [r['name'] for r in self.db.query('SELECT name FROM categories ORDER BY name;')]
        mans = [r['name'] for r in self.db.query('SELECT name FROM manufacturers ORDER BY name;')]
        forms = [r['name'] for r in self.db.query('SELECT name FROM formulas ORDER BY name;')]
        units = ['mg','ml','g','IU','tablet','capsule','bottle','strip','box']

        def save(d):
            cid = mid = fid = None
            if d.get('category'):
                r = self.db.query('SELECT id FROM categories WHERE name=?;',(d['category'],))
                if r: cid = r[0]['id']
                else: cid = self.db.execute('INSERT INTO categories(name) VALUES(?);',(d['category'],))
            if d.get('manufacturer'):
                r = self.db.query('SELECT id FROM manufacturers WHERE name=?;',(d['manufacturer'],))
                if r: mid = r[0]['id']
                else: mid = self.db.execute('INSERT INTO manufacturers(name) VALUES(?);',(d['manufacturer'],))
            if d.get('formula'):
                r = self.db.query('SELECT id FROM formulas WHERE name=?;',(d['formula'],))
                if r: fid = r[0]['id']
                else: fid = self.db.execute('INSERT INTO formulas(name) VALUES(?);',(d['formula'],))
            try:
                self.db.execute('UPDATE products SET name=?,sku=?,category_id=?,manufacturer_id=?,formula_id=?,unit=?,sale_price=?,notes=? WHERE id=?;',
                                (d.get('name'), d.get('sku') or None, cid, mid, fid, d.get('unit') or '', float(d.get('price') or 0), d.get('notes') or '', pid))
                messagebox.showinfo('Saved','Product updated'); self._inv_refresh_all()
            except Exception as e:
                messagebox.showerror('Error', str(e))

        fields = [
            {'key':'name','label':'Name'},
            {'key':'sku','label':'SKU'},
            {'key':'unit','label':'Unit','widget':'combobox','values':units,'state':'normal'},
            {'key':'category','label':'Category','widget':'combobox','values':cats,'state':'normal'},
            {'key':'manufacturer','label':'Manufacturer','widget':'combobox','values':mans,'state':'normal'},
            {'key':'formula','label':'Formula','widget':'combobox','values':forms,'state':'normal'},
            {'key':'price','label':'Sale Price'},
            {'key':'notes','label':'Notes','widget':'text'}
        ]

        dlg = FormDialog(self.root, 'Edit Product', fields, initial=initial, on_submit=save)

        # Make comboboxes searchable
        self.make_searchable(dlg.widgets['unit'][0], units)
        self.make_searchable(dlg.widgets['category'][0], cats)
        self.make_searchable(dlg.widgets['manufacturer'][0], mans)
        self.make_searchable(dlg.widgets['formula'][0], forms)


    def _inv_delete_product(self, tree):
        sel = tree.selection()
        if not sel: return messagebox.showwarning('Select','Select a product to delete')
        values = tree.item(sel[0], 'values')
        pid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete selected product?'): return
        try:
            self.db.execute('DELETE FROM products WHERE id=?;',(pid,))
            messagebox.showinfo('Deleted','Product deleted'); self._inv_refresh_all()
        except Exception as e:
            messagebox.showerror('Error', str(e))

    # Suppliers CRUD
    def _add_supplier(self):
        def save(d):
            if not d.get('name'): return messagebox.showerror('Error','Name required')
            try:
                self.db.execute('INSERT INTO suppliers(name,phone,email,address) VALUES(?,?,?,?);',(d.get('name'), d.get('phone'), d.get('email'), d.get('address')))
                messagebox.showinfo('Saved','Supplier added'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Add Supplier', [{'key':'name','label':'Name'},{'key':'phone','label':'Phone'},{'key':'email','label':'Email'},{'key':'address','label':'Address'}], on_submit=save)

    def _edit_supplier(self):
        sel = self._sup_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select supplier to edit')
        values = self._sup_tree.item(sel[0], 'values')
        sid = int(values[0]); r = self.db.query('SELECT * FROM suppliers WHERE id=?;',(sid,))[0]
        initial = {'name':r['name'],'phone':r['phone'],'email':r['email'],'address':r['address']}
        def save(d):
            try:
                self.db.execute('UPDATE suppliers SET name=?,phone=?,email=?,address=? WHERE id=?;',(d.get('name'),d.get('phone'),d.get('email'),d.get('address'),sid))
                messagebox.showinfo('Saved','Supplier updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Edit Supplier', [{'key':'name','label':'Name'},{'key':'phone','label':'Phone'},{'key':'email','label':'Email'},{'key':'address','label':'Address'}], initial=initial, on_submit=save)

    def _delete_supplier(self):
        sel = self._sup_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select supplier to delete')
        values = self._sup_tree.item(sel[0], 'values')
        sid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete supplier?'): return
        try:
            self.db.execute('DELETE FROM suppliers WHERE id=?;',(sid,)); messagebox.showinfo('Deleted','Supplier deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))

    # Manufacturers CRUD
    def _add_manufacturer(self):
        def save(d):
            if not d.get('name'): return messagebox.showerror('Error','Name required')
            try:
                self.db.execute('INSERT INTO manufacturers(name,contact,notes) VALUES(?,?,?);',(d.get('name'),d.get('contact'),d.get('notes')))
                messagebox.showinfo('Saved','Manufacturer added'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Add Manufacturer', [{'key':'name','label':'Name'},{'key':'contact','label':'Contact'},{'key':'notes','label':'Notes','widget':'text'}], on_submit=save)

    def _edit_manufacturer(self):
        sel = self._man_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select manufacturer to edit')
        values = self._man_tree.item(sel[0], 'values')
        mid = int(values[0]); r = self.db.query('SELECT * FROM manufacturers WHERE id=?;',(mid,))[0]
        initial = {'name':r['name'],'contact':r['contact'],'notes':r['notes']}
        def save(d):
            try:
                self.db.execute('UPDATE manufacturers SET name=?,contact=?,notes=? WHERE id=?;',(d.get('name'),d.get('contact'),d.get('notes'),mid))
                messagebox.showinfo('Saved','Manufacturer updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Edit Manufacturer', [{'key':'name','label':'Name'},{'key':'contact','label':'Contact'},{'key':'notes','label':'Notes','widget':'text'}], initial=initial, on_submit=save)

    def _delete_manufacturer(self):
        sel = self._man_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select manufacturer to delete')
        values = self._man_tree.item(sel[0], 'values')
        mid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete manufacturer?'): return
        try:
            self.db.execute('DELETE FROM manufacturers WHERE id=?;',(mid,)); messagebox.showinfo('Deleted','Manufacturer deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))

    # Categories CRUD
    def _add_category(self):
        def save(d):
            if not d.get('name'): return messagebox.showerror('Error','Name required')
            try:
                self.db.execute('INSERT INTO categories(name,notes) VALUES(?,?);',(d.get('name'), d.get('notes'))); messagebox.showinfo('Saved','Category added'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Add Category', [{'key':'name','label':'Name'},{'key':'notes','label':'Notes','widget':'text'}], on_submit=save)

    def _edit_category(self):
        sel = self._cat_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select category to edit')
        values = self._cat_tree.item(sel[0], 'values')
        cid = int(values[0]); r = self.db.query('SELECT * FROM categories WHERE id=?;',(cid,))[0]
        initial = {'name':r['name'],'notes':r['notes']}
        def save(d):
            try:
                self.db.execute('UPDATE categories SET name=?,notes=? WHERE id=?;',(d.get('name'),d.get('notes'),cid)); messagebox.showinfo('Saved','Category updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Edit Category', [{'key':'name','label':'Name'},{'key':'notes','label':'Notes','widget':'text'}], initial=initial, on_submit=save)

    def _delete_category(self):
        sel = self._cat_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select category to delete')
        values = self._cat_tree.item(sel[0], 'values')
        cid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete category?'): return
        try:
            self.db.execute('DELETE FROM categories WHERE id=?;',(cid,)); messagebox.showinfo('Deleted','Category deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))

    # Formulas CRUD
    def _add_formula(self):
        def save(d):
            if not d.get('name'): return messagebox.showerror('Error','Name required')
            try:
                self.db.execute('INSERT INTO formulas(name,composition) VALUES(?,?);',(d.get('name'),d.get('composition'))); messagebox.showinfo('Saved','Formula added'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Add Formula', [{'key':'name','label':'Name'},{'key':'composition','label':'Composition','widget':'text'}], on_submit=save)

    def _edit_formula(self):
        sel = self._form_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select formula to edit')
        values = self._form_tree.item(sel[0], 'values')
        fid = int(values[0]); r = self.db.query('SELECT * FROM formulas WHERE id=?;',(fid,))[0]
        initial = {'name':r['name'],'composition':r['composition']}
        def save(d):
            try:
                self.db.execute('UPDATE formulas SET name=?,composition=? WHERE id=?;',(d.get('name'),d.get('composition'),fid)); messagebox.showinfo('Saved','Formula updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Edit Formula', [{'key':'name','label':'Name'},{'key':'composition','label':'Composition','widget':'text'}], initial=initial, on_submit=save)

    def _delete_formula(self):
        sel = self._form_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select formula to delete')
        values = self._form_tree.item(sel[0], 'values')
        fid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete formula?'): return
        try:
            self.db.execute('DELETE FROM formulas WHERE id=?;',(fid,)); messagebox.showinfo('Deleted','Formula deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))


    def make_searchable(self, cb_widget, original_values):
        def on_keyrelease(event):
            typed = cb_widget.get()
            if typed == '':
                cb_widget['values'] = original_values  # show all if empty
            else:
                # Only items starting with typed string (case-insensitive)
                filtered = [v for v in original_values if v.lower().startswith(typed.lower())]
                cb_widget['values'] = filtered
            if cb_widget['values']:
                cb_widget.event_generate('<Down>')  # open dropdown automatically

        cb_widget.bind('<KeyRelease>', on_keyrelease)


    def _add_batch(self):
        # Fetch data from database
        products = [r['name'] for r in self.db.query('SELECT name FROM products ORDER BY name;')]
        suppliers = [r['name'] for r in self.db.query('SELECT name FROM suppliers ORDER BY name;')]
        manufacturers = [r['name'] for r in self.db.query('SELECT name FROM manufacturers ORDER BY name;')]

        # Function to save batch
        def save(d):
            pid = sid = mid = None

            p = self.db.query('SELECT id FROM products WHERE name=?;', (d.get('product'),))
            if p: pid = p[0]['id']

            if d.get('supplier'):
                s = self.db.query('SELECT id FROM suppliers WHERE name=?;', (d.get('supplier'),))
                if s: sid = s[0]['id']

            if d.get('manufacturer'):
                m = self.db.query('SELECT id FROM manufacturers WHERE name=?;', (d.get('manufacturer'),))
                if m: mid = m[0]['id']

            if not pid:
                return messagebox.showerror('Error', 'Product is required and must exist.')

            try:
                self.db.execute(
                    'INSERT INTO batches(product_id, supplier_id, manufacturer_id, batch_no, quantity, expiry_date, cost_price, created_at) '
                    'VALUES(?,?,?,?,?,?,?,?);',
                    (pid, sid, mid, d.get('batch_no') or '', int(d.get('quantity') or 0),
                    d.get('expiry') or None, float(d.get('cost_price') or 0), now_str())
                )
                messagebox.showinfo('Saved', 'Batch added successfully.')
                self._inv_refresh_all()
            except Exception as e:
                messagebox.showerror('Error', str(e))

        # Form fields — state='normal' is required for typing
        fields = [
            {'key': 'product', 'label': 'Product', 'widget': 'combobox', 'values': products, 'state': 'normal'},
            {'key': 'supplier', 'label': 'Supplier', 'widget': 'combobox', 'values': suppliers, 'state': 'normal'},
            {'key': 'manufacturer', 'label': 'Manufacturer', 'widget': 'combobox', 'values': manufacturers, 'state': 'normal'},
            {'key': 'batch_no', 'label': 'Batch No'},
            {'key': 'quantity', 'label': 'Quantity'},
            {'key': 'expiry', 'label': 'Expiry (YYYY-MM-DD)'},
            {'key': 'cost_price', 'label': 'Cost Price'}
        ]

        # Open form dialog
        dlg = FormDialog(self.root, 'Add Batch', fields, on_submit=save)

        # Make Product, Supplier, Manufacturer comboboxes smooth-searchable
        self.make_searchable(dlg.widgets['product'][0], products)
        self.make_searchable(dlg.widgets['supplier'][0], suppliers)
        self.make_searchable(dlg.widgets['manufacturer'][0], manufacturers)

    # Find the _edit_batch method and update it to include manufacturer:
    def _edit_batch(self):
        sel = self._batch_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select batch to edit')
        values = self._batch_tree.item(sel[0], 'values')
        bid = int(values[0]); r = self.db.query('SELECT * FROM batches WHERE id=?;',(bid,))[0]
        prodname = self.db.query('SELECT name FROM products WHERE id=?;',(r['product_id'],))[0]['name']
        supname = None
        if r['supplier_id']:
            s = self.db.query('SELECT name FROM suppliers WHERE id=?;',(r['supplier_id'],))
            supname = s[0]['name'] if s else None
        
        # Get manufacturer name
        manname = None
        if r['product_id']:
            m = self.db.query('SELECT m.name FROM products p LEFT JOIN manufacturers m ON m.id=p.manufacturer_id WHERE p.id=?;',(r['product_id'],))
            manname = m[0]['name'] if m and m[0]['name'] else None
        
        initial = {'product':prodname,'supplier':supname,'manufacturer':manname,'batch_no':r['batch_no'],'quantity':r['quantity'],'expiry':r['expiry_date'],'cost_price':r['cost_price']}
        
        def save(d):
            p = self.db.query('SELECT id FROM products WHERE name=?;',(d.get('product'),))
            if not p: return messagebox.showerror('Error','Product required and must exist')
            pid = p[0]['id']
            sid = None
            if d.get('supplier'):
                s = self.db.query('SELECT id FROM suppliers WHERE name=?;',(d.get('supplier'),))
                if s: sid = s[0]['id']
            
            # Get manufacturer ID
            mid = None
            if d.get('manufacturer'):
                m = self.db.query('SELECT id FROM manufacturers WHERE name=?;',(d.get('manufacturer'),))
                if m: mid = m[0]['id']
            
            try:
                self.db.execute('UPDATE batches SET product_id=?,supplier_id=?,batch_no=?,quantity=?,expiry_date=?,cost_price=? WHERE id=?;',
                                (pid, sid, d.get('batch_no') or '', int(d.get('quantity') or 0), d.get('expiry') or None, float(d.get('cost_price') or 0), bid))
                messagebox.showinfo('Saved','Batch updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        
        fields = [
            {'key':'product','label':'Product','widget':'entry'},
            {'key':'supplier','label':'Supplier','widget':'entry'},
            {'key':'manufacturer','label':'Manufacturer','widget':'entry'},
            {'key':'batch_no','label':'Batch No'},
            {'key':'quantity','label':'Quantity'},
            {'key':'expiry','label':'Expiry (YYYY-MM-DD)'},
            {'key':'cost_price','label':'Cost Price'}
        ]
        
        dlg = FormDialog(self.root, 'Edit Batch', fields, initial=initial, on_submit=save)
        
        # Add autocomplete functionality after creating the widgets
        if 'product' in dlg.widgets:
            product_entry = dlg.widgets['product'][0]
            product_entry.suggestions_getter = lambda term: [r['name'] for r in self.db.query(
                "SELECT name FROM products WHERE name LIKE ? ORDER BY name LIMIT 10", (f"%{term}%",))]
        
        if 'supplier' in dlg.widgets:
            supplier_entry = dlg.widgets['supplier'][0]
            supplier_entry.suggestions_getter = lambda term: [r['name'] for r in self.db.query(
                "SELECT name FROM suppliers WHERE name LIKE ? ORDER BY name LIMIT 10", (f"%{term}%",))]
        
        if 'manufacturer' in dlg.widgets:
            manufacturer_entry = dlg.widgets['manufacturer'][0]
            manufacturer_entry.suggestions_getter = lambda term: [r['name'] for r in self.db.query(
                "SELECT name FROM manufacturers WHERE name LIKE ? ORDER BY name LIMIT 10", (f"%{term}%",))]
        
        # Add functionality to auto-fill manufacturer when product is selected
        def on_product_select(event=None):
            product_name = dlg.widgets['product'][0].get().strip()
            if product_name:
                # Get manufacturer for this product
                result = self.db.query('''
                    SELECT m.name 
                    FROM products p 
                    LEFT JOIN manufacturers m ON p.manufacturer_id = m.id 
                    WHERE p.name = ? LIMIT 1
                ''', (product_name,))
                
                if result and result[0]['name']:
                    dlg.widgets['manufacturer'][0].delete(0, 'end')
                    dlg.widgets['manufacturer'][0].insert(0, result[0]['name'])
        
        # Bind the product entry to auto-fill manufacturer
        dlg.widgets['product'][0].bind('<FocusOut>', on_product_select)
        dlg.widgets['product'][0].bind('<Return>', on_product_select)

    def _delete_batch(self):
        sel = self._batch_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select batch to delete')
        values = self._batch_tree.item(sel[0], 'values')
        bid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete batch?'): return
        try:
            self.db.execute('DELETE FROM batches WHERE id=?;',(bid,)); messagebox.showinfo('Deleted','Batch deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))

     # ---------------- POS with nested tabs ----------------
    def _build_pos_tab(self):
        # Clear POS tab
        for w in self.tab_pos.winfo_children():
            w.destroy()
        pos_nb = ttk.Notebook(self.tab_pos)
        pos_nb.pack(fill='both', expand=True, padx=8, pady=8)

        # --- New Sale ---
        new_sale_tab = ttk.Frame(pos_nb)
        pos_nb.add(new_sale_tab, text='New Sale')
        NewSaleTab(new_sale_tab, self.db, self.user).pack(fill='both', expand=True)

        # --- Sale History ---
        history_tab = ttk.Frame(pos_nb)
        pos_nb.add(history_tab, text='Sale History')
        self._sale_history_tree = ttk.Treeview(
            history_tab,
            columns=('sale_id','date','customer','product','qty','price','expiry','supplier','manufacturer','subtotal'),
            show='headings', height=18
        )
        for c in ('sale_id','date','customer','product','qty','price','expiry','supplier','manufacturer','subtotal'):
            self._sale_history_tree.heading(c, text=c.capitalize())
            self._sale_history_tree.column(c, width=120, anchor='w')
        self._sale_history_tree.pack(fill='both', expand=True, padx=8, pady=8)
        btns = ttk.Frame(history_tab)
        btns.pack(fill='x')
        ttk.Button(btns, text='Refresh', command=self._sale_history_refresh).pack(side='left', padx=6)
        ttk.Button(btns, text='Print Receipt (Selected)', command=self._sale_history_print_selected).pack(side='left', padx=6)
        try:
            self._sale_history_refresh()
        except Exception:
            pass
        # --- Return Item ---
        return_item_tab = ttk.Frame(pos_nb)
        pos_nb.add(return_item_tab, text='Return Item')
        self._build_return_item_tab(return_item_tab)

        # --- Return History ---
        returns_tab = ttk.Frame(pos_nb)
        pos_nb.add(returns_tab, text='Return History')
        self._return_tree = ttk.Treeview(
            returns_tab,
            columns=('id','sale_item','product','qty','reason','created','expiry'),
            show='headings'
        )
        for c in ('id','sale_item','product','qty','reason','created','expiry'):
            self._return_tree.heading(c, text=c.capitalize())
            self._return_tree.column(c, width=120, anchor='w')
        self._return_tree.pack(fill='both', expand=True, padx=8, pady=8)
        ttk.Button(returns_tab, text='Refresh', command=self._return_refresh).pack(anchor='e', padx=8, pady=6)
        try:
            self._return_refresh()
        except Exception:
            pass

        # --- Sale Reports ---
        reports_tab = ttk.Frame(pos_nb)
        pos_nb.add(reports_tab, text='Sale Reports')
        try:
            # Use the fixed reports implementation
            self._build_reports_in_frame(reports_tab)
        except Exception as e:
            print(f"Error building reports: {e}")
            # Simple fallback
            ttk.Label(reports_tab, text="Reports functionality not available").pack(pady=20)


    def _build_return_item_tab(self, frame):
        # --- Search by Sale ID ---
        search_frame = ttk.Frame(frame)
        search_frame.pack(fill="x", padx=10, pady=6)

        ttk.Label(search_frame, text="Sale ID:").pack(side="left")
        self.return_sale_id = ttk.Entry(search_frame, width=15)
        self.return_sale_id.pack(side="left", padx=6)

        ttk.Button(search_frame, text="Fetch Sale",
            command=lambda: self._load_sale_for_return(self.return_sale_id.get())
        ).pack(side="left")

        # --- Treeview for sale items ---
        self.return_tree = ttk.Treeview(
            frame,
            columns=("id", "product", "qty", "price", "total"),
            show="headings"
        )
        for col in ("id", "product", "qty", "price", "total"):
            self.return_tree.heading(col, text=col.capitalize())
            self.return_tree.column(col, width=120, anchor="w")
        self.return_tree.pack(fill="both", expand=True, padx=10, pady=6)

        # --- Return controls ---
        control_frame = ttk.Frame(frame)
        control_frame.pack(fill="x", padx=10, pady=6)

        ttk.Label(control_frame, text="Quantity to Return:").pack(side="left")
        self.return_qty = ttk.Entry(control_frame, width=10)
        self.return_qty.pack(side="left", padx=6)

        ttk.Label(control_frame, text="Reason:").pack(side="left")
        self.return_reason = ttk.Entry(control_frame, width=30)
        self.return_reason.pack(side="left", padx=6)

        ttk.Button(control_frame, text="Process Return",
            command=self._process_return
        ).pack(side="left", padx=6)

    def _load_sale_for_return(self, sale_id):
        rows = self.db.query("""
            SELECT si.id, p.name AS product, si.quantity, si.price,
                (si.quantity * si.price) as total
            FROM sale_items si
            JOIN products p ON si.product_id = p.id
            WHERE si.sale_id = ?;
        """, (sale_id,))
        self.return_tree.delete(*self.return_tree.get_children())
        for r in rows:
            self.return_tree.insert("", "end", values=(r["id"], r["product"], r["quantity"], r["price"], r["total"]))

    def _process_return(self):
        selected = self.return_tree.selection()
        if not selected:
            messagebox.showerror("Error", "Select an item to return.")
            return

        item_id, product, qty, price, total = self.return_tree.item(selected[0], "values")
        return_qty = int(self.return_qty.get() or 0)
        reason = self.return_reason.get()

        if return_qty <= 0 or return_qty > int(qty):
            messagebox.showerror("Error", "Invalid return quantity.")
            return

        # 1. Log the return
        self.db.execute("""
            INSERT INTO returns (sale_item_id, quantity, reason)
            VALUES (?, ?, ?)
        """, (item_id, return_qty, reason))

        # 2. Restore stock to batches used in this sale
        batches = self.db.query("""
            SELECT sib.batch_id, sib.quantity
            FROM sale_item_batches sib
            WHERE sib.sale_item_id = ?
            ORDER BY sib.id ASC;
        """, (item_id,))

        remain = return_qty
        for b in batches:
            if remain <= 0:
                break
            give_back = min(remain, b["quantity"])
            self.db.execute(
                "UPDATE batches SET quantity = quantity + ? WHERE id = ?;",
                (give_back, b["batch_id"])
            )
            remain -= give_back

        # 3. Update sale_items to reflect reduced sold qty
        self.db.execute("""
            UPDATE sale_items
            SET quantity = quantity - ?
            WHERE id = ?
        """, (return_qty, item_id))

        messagebox.showinfo("Success", f"Returned {return_qty} x {product}")
        # Refresh items for this sale
        self._load_sale_for_return(self.return_sale_id.get())

    def _build_sale_history_tab(self):
        # kept for backward compatibility (not used as top-level)
        pass

    def _sale_history_refresh(self):
        tree = getattr(self, '_sale_history_tree', None)
        if not tree:
            return
        
        tree.delete(*tree.get_children())
        
        # Simplified query that doesn't rely on sale_item_batches
        rows = self.db.query('''
            SELECT 
                s.id AS sale_id, 
                s.created_at AS date, 
                s.customer_name AS customer,
                p.name AS product, 
                si.quantity AS qty, 
                si.price AS price,
                (si.quantity * si.price) AS subtotal
            FROM sales s
            JOIN sale_items si ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            ORDER BY s.created_at DESC LIMIT 200;
        ''')
        
        for r in rows:
            tree.insert('', 'end', values=(
                r['sale_id'], 
                r['date'], 
                r['customer'] or '', 
                r['product'], 
                r['qty'], 
                f"{r['price']:.2f}", 
                "",  # expiry (empty for now)
                "",  # supplier (empty for now)
                "",  # manufacturer (empty for now)
                f"{r['subtotal']:.2f}"
            ))

    def _sale_history_print_selected(self):
        sel = self._sale_history_tree.selection()
        if not sel:
            return messagebox.showwarning('Select','Select a row (sale item) to print its sale receipt')
        item = self._sale_history_tree.item(sel[0])['values']
        sale_id = item[0]
        if REPORTLAB_AVAILABLE:
            self._print_receipt(sale_id)
        else:
            messagebox.showwarning('Missing','reportlab required for PDF receipt')
 
    def _print_receipt(self, sale_id):
        sale = self.db.query('SELECT * FROM sales WHERE id=?;', (sale_id,))
        if not sale:
            return messagebox.showerror('Error','Sale not found')
        sale = sale[0]

        total = float(sale['total'])
        cust_name = sale.get('customer_name') or ""
        cust_phone = sale.get('customer_phone') or ""

        try:
            self.print_receipt_direct(sale_id, total, cust_name, cust_phone)
            messagebox.showinfo("Printing", "Receipt sent to printer.")
        except Exception as e:
            messagebox.showerror("Print Error", f"Could not print receipt directly:\n{e}")



    # ---------------- Return History ----------------
    def _build_return_history_tab(self):
        # kept for backward compatibility (not used as top-level)
        pass

    def _return_refresh(self):
        tree = getattr(self, '_return_tree', None)
        if not tree: return
        tree.delete(*tree.get_children())
        rows = self.db.query('''SELECT r.id, r.sale_item_id AS sale_item, p.name AS product, r.quantity AS qty, r.reason, r.created_at AS created, b.expiry_date AS expiry
            FROM returns r JOIN sale_items si ON si.id=r.sale_item_id JOIN products p ON p.id=si.product_id
            LEFT JOIN sale_item_batches sib ON sib.sale_item_id=si.id LEFT JOIN batches b ON b.id=sib.batch_id
            ORDER BY r.id DESC LIMIT 500;''')
        for r in rows:
            tree.insert('', 'end', values=(r['id'], r['sale_item'], r['product'], r['qty'], r['reason'], r['created'], r['expiry'] or ''))

   
        # ---------------- Reports ----------------
    def _build_reports_in_frame(self, frame):
        for w in frame.winfo_children(): 
            w.destroy()
        
        # ---------------- Filters ----------------
        filter_frame = ttk.Frame(frame, padding=10)
        filter_frame.pack(fill='x', pady=5)
        
        # Date range
        ttk.Label(filter_frame, text="From Date:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        from_date = ttk.Entry(filter_frame, width=12)
        from_date.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(filter_frame, text="To Date:").grid(row=0, column=2, padx=5, pady=5, sticky='e')
        to_date = ttk.Entry(filter_frame, width=12)
        to_date.grid(row=0, column=3, padx=5, pady=5)
        
        # Product filter with autocomplete
        ttk.Label(filter_frame, text="Product:").grid(row=0, column=4, padx=5, pady=5, sticky='e')
        product_filter = AutocompleteEntry(filter_frame, width=20, suggestions_getter=self._product_suggestions)
        product_filter.grid(row=0, column=5, padx=5, pady=5)
        
        # Customer filter with autocomplete
        ttk.Label(filter_frame, text="Customer:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        customer_filter = AutocompleteEntry(filter_frame, width=20, suggestions_getter=self._customer_suggestions)
        customer_filter.grid(row=1, column=1, padx=5, pady=5)
        
        # Supplier filter with autocomplete
        ttk.Label(filter_frame, text="Supplier:").grid(row=1, column=2, padx=5, pady=5, sticky='e')
        supplier_filter = AutocompleteEntry(filter_frame, width=20, suggestions_getter=self._supplier_suggestions)
        supplier_filter.grid(row=1, column=3, padx=5, pady=5)
        
        # Filter button
        filter_btn = ttk.Button(filter_frame, text="Apply Filters", 
                                command=lambda: self._apply_report_filters(
                                    from_date.get(), to_date.get(),
                                    product_filter.get(), customer_filter.get(), supplier_filter.get()
                                ))
        filter_btn.grid(row=2, column=0, columnspan=4, pady=10)
        
        # ---------------- Report Treeview ----------------
        columns = ('sale_id', 'date', 'customer', 'product', 'quantity', 'price', 'subtotal', 'supplier')
        tree = ttk.Treeview(frame, columns=columns, show='headings', height=15)
        
        col_widths = {
            'sale_id': 80, 'date': 150, 'customer': 150, 'product': 200,
            'quantity': 80, 'price': 80, 'subtotal': 100, 'supplier': 150
        }
        for col in columns:
            tree.heading(col, text=col.replace('_', ' ').title())
            tree.column(col, width=col_widths[col], anchor='center')
        
        scrollbar = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        scrollbar.pack(side='right', fill='y', padx=(0,10), pady=10)
        
        # ---------------- Export Buttons ----------------
        export_frame = ttk.Frame(frame)
        export_frame.pack(fill='x', pady=5, padx=10)
        
        ttk.Button(export_frame, text="Refresh", command=self._load_all_sales).pack(side='left', padx=5)
        ttk.Button(export_frame, text="Export CSV", command=lambda: self._export_report(tree, "csv")).pack(side='left', padx=5)
        ttk.Button(export_frame, text="Export Excel", command=lambda: self._export_report(tree, "excel")).pack(side='left', padx=5)
        
        # Save reference
        self._report_tree = tree
        self._load_all_sales()

    def _apply_report_filters(self, from_date, to_date, product_filter, customer_filter, supplier_filter):
        if not hasattr(self, '_report_tree') or self._report_tree is None:
            return
        
        query = '''
            SELECT s.id AS sale_id, s.created_at AS date, s.customer_name AS customer,
                p.name AS product, si.quantity AS quantity, si.price AS price,
                (si.quantity * si.price) AS subtotal, sup.name AS supplier
            FROM sales s
            JOIN sale_items si ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            LEFT JOIN sale_item_batches sib ON sib.sale_item_id = si.id
            LEFT JOIN batches b ON b.id = sib.batch_id
            LEFT JOIN suppliers sup ON sup.id = b.supplier_id
            WHERE 1=1
        '''
        params = []
        if from_date: params += [from_date]; query += " AND date(s.created_at) >= ?"
        if to_date:   params += [to_date];   query += " AND date(s.created_at) <= ?"
        if product_filter:  params += [f'%{product_filter}%']; query += " AND p.name LIKE ?"
        if customer_filter: params += [f'%{customer_filter}%']; query += " AND s.customer_name LIKE ?"
        if supplier_filter: params += [f'%{supplier_filter}%']; query += " AND sup.name LIKE ?"
        
        query += " ORDER BY s.created_at DESC"
        
        try:
            rows = self.db.query(query, tuple(params))
            self._report_tree.delete(*self._report_tree.get_children())
            for r in rows:
                self._report_tree.insert('', 'end', values=(
                    r['sale_id'], r['date'], r['customer'] or 'N/A',
                    r['product'], r['quantity'], f"{r['price']:.2f}",
                    f"{r['subtotal']:.2f}", r['supplier'] or 'N/A'
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load report data: {str(e)}")

    def _product_suggestions(self, term):
            rows = self.db.query('SELECT name FROM products WHERE name LIKE ? ORDER BY name LIMIT 10;', (f'%{term}%',))
            return [r['name'] for r in rows]

    def _customer_suggestions(self, term):
        rows = self.db.query('SELECT DISTINCT customer_name FROM sales WHERE customer_name LIKE ? ORDER BY customer_name LIMIT 10;', (f'%{term}%',))
        return [r['customer_name'] for r in rows if r['customer_name']]

    def _supplier_suggestions(self, term):
        rows = self.db.query('SELECT name FROM suppliers WHERE name LIKE ? ORDER BY name LIMIT 10;', (f'%{term}%',))
        return [r['name'] for r in rows if r['name']]
    
    def _load_all_sales(self):
        if not hasattr(self, '_report_tree') or self._report_tree is None:
            return
        try:
            rows = self.db.query('''
                SELECT s.id AS sale_id, s.created_at AS date, s.customer_name AS customer,
                    p.name AS product, si.quantity AS quantity, si.price AS price,
                    (si.quantity * si.price) AS subtotal,
                    sup.name AS supplier
                FROM sales s
                JOIN sale_items si ON si.sale_id = s.id
                JOIN products p ON p.id = si.product_id
                LEFT JOIN sale_item_batches sib ON sib.sale_item_id = si.id
                LEFT JOIN batches b ON b.id = sib.batch_id
                LEFT JOIN suppliers sup ON sup.id = b.supplier_id
                ORDER BY s.created_at DESC LIMIT 500
            ''')
            self._report_tree.delete(*self._report_tree.get_children())
            for r in rows:
                self._report_tree.insert('', 'end', values=(
                    r['sale_id'], r['date'], r['customer'] or 'N/A',
                    r['product'], r['quantity'], f"{r['price']:.2f}",
                    f"{r['subtotal']:.2f}", r['supplier'] or 'N/A'
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sales data: {str(e)}")

    def _export_report(self, tree, fmt="csv"):
        try:
            rows = [tree.item(i, "values") for i in tree.get_children()]
            if not rows:
                messagebox.showwarning("No Data", "No report data to export.")
                return
            
            filetypes = {"csv": [("CSV Files", "*.csv")], "excel": [("Excel Files", "*.xlsx")]}
            ext = "xlsx" if fmt=="excel" else "csv"
            fpath = filedialog.asksaveasfilename(defaultextension=f".{ext}", filetypes=filetypes[fmt], title=f"Export Report as {fmt.upper()}")
            if not fpath: return
            
            if fmt=="csv":
                import csv
                with open(fpath, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(("Sale ID","Date","Customer","Product","Quantity","Price","Subtotal","Supplier"))
                    writer.writerows(rows)
            else:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Sales Report"
                ws.append(("Sale ID","Date","Customer","Product","Quantity","Price","Subtotal","Supplier"))
                for r in rows: ws.append(r)
                wb.save(fpath)
            
            messagebox.showinfo("Success", f"Report successfully exported to:\n{fpath}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export report: {str(e)}")

    

    # ---------------- Manage Staff ----------------
    def _build_manage_staff_tab(self):
        # Clear old widgets
        for w in self.tab_manage_staff.winfo_children():
            w.destroy()

        # Header
        header = ttk.Frame(self.tab_manage_staff)
        header.pack(fill='x', pady=6)
        ttk.Label(header, text='Manage Staff', font=('Segoe UI',14,'bold')).pack(side='left', padx=8)
        ttk.Button(header, text='Add Staff', command=self._add_staff).pack(side='right', padx=6)

        # Staff Table
        cols = ('id','username','role')
        tree = ttk.Treeview(self.tab_manage_staff, columns=cols, show='headings', height=15)
        tree.heading('id', text='ID'); tree.column('id', width=60, anchor='center')
        tree.heading('username', text='Username'); tree.column('username', width=200, anchor='w')
        tree.heading('role', text='Role'); tree.column('role', width=120, anchor='center')
        tree.pack(fill='both', expand=True, padx=10, pady=10)

        self._staff_tree = tree
        self._refresh_staff()

        # Buttons below table
        btns = ttk.Frame(self.tab_manage_staff)
        btns.pack(fill='x', pady=5)

        def edit_staff():
            sel = tree.selection()
            if not sel: return
            uid = tree.item(sel[0])['values'][0]
            username = tree.item(sel[0])['values'][1]
            role = tree.item(sel[0])['values'][2]
            
            def save(d):
                if not d.get('username'): 
                    return messagebox.showerror('Error','Username required')
                if d.get('password'):
                    self.db.execute(
                        "UPDATE users SET username=?, password_hash=?, role=? WHERE id=?",
                        (d['username'], hash_pw(d['password']), d['role'], uid)
                    )
                else:
                    self.db.execute(
                        "UPDATE users SET username=?, role=? WHERE id=?",
                        (d['username'], d['role'], uid)
                    )
                messagebox.showinfo('Saved','Staff updated successfully')
                self._refresh_staff()

            FormDialog(self.root, 'Edit Staff', [
                {'key':'username','label':'Username'},
                {'key':'password','label':'New Password (leave blank to keep current)'},
                {'key':'role','label':'Role','widget':'combobox','values':['staff','cashier']}
            ], initial={'username': username, 'role': role}, on_submit=save)

        def delete_staff():
            sel = tree.selection()
            if not sel: return
            uid = tree.item(sel[0])['values'][0]
            if uid == self.user['id']:
                return messagebox.showerror('Error', 'Cannot delete your own account')
            if messagebox.askyesno('Confirm','Delete this staff member?'):
                try:
                    self.db.execute("DELETE FROM users WHERE id=?;", (uid,))
                    self._refresh_staff()
                except Exception as e:
                    messagebox.showerror('Error', f'Failed to delete staff: {str(e)}')

        ttk.Button(btns, text='Edit Staff', command=edit_staff).pack(side='left', padx=6)
        ttk.Button(btns, text='Delete Staff', command=delete_staff).pack(side='left', padx=6)

    def _refresh_staff(self):
        tree = getattr(self, '_staff_tree', None)
        if not tree: return
        tree.delete(*tree.get_children())
        rows = self.db.query("SELECT id, username, role FROM users WHERE role IN ('staff', 'cashier') ORDER BY username;")
        for r in rows:
            tree.insert('', 'end', values=(r['id'], r['username'], r['role']))

    def _add_staff(self):
        def save(d):
            if not d.get('username') or not d.get('password'):
                return messagebox.showerror('Error','Username and password required')
            existing = self.db.query("SELECT id FROM users WHERE username=?;", (d['username'],))
            if existing:
                return messagebox.showerror('Error','Username already exists')
            try:
                self.db.execute(
                    "INSERT INTO users(username,password_hash,role) VALUES(?,?,?)",
                    (d['username'], hash_pw(d['password']), d['role'])
                )
                messagebox.showinfo('Saved','Staff added successfully')
                self._refresh_staff()
            except Exception as e:
                messagebox.showerror('Error', f'Failed to add staff: {str(e)}')

        FormDialog(self.root, 'Add Staff', [
            {'key':'username','label':'Username'},
            {'key':'password','label':'Password'},
            {'key':'role','label':'Role','widget':'combobox','values':['staff','cashier']}
        ], on_submit=save)

    # ---------------- Settings ----------------
    def _build_settings_tab(self):
        
        for w in self.tab_settings.winfo_children(): w.destroy()
        f = ttk.Frame(self.tab_settings); f.pack(fill='x', padx=8, pady=8)
        # Pharmacy identity
        ttk.Label(f, text='Pharmacy Name').grid(row=0, column=0, sticky='w', padx=4, pady=4)
        name_e = ttk.Entry(f, width=40); name_e.grid(row=0, column=1, padx=4, pady=4)
        ttk.Label(f, text='Pharmacy Address').grid(row=1, column=0, sticky='w', padx=4, pady=4)
        addr_e = ttk.Entry(f, width=60); addr_e.grid(row=1, column=1, padx=4, pady=4)
        # Existing
        ttk.Label(f, text='Default Tax Percent (%)').grid(row=2, column=0, sticky='w', padx=4, pady=4)
        tax_e = ttk.Entry(f, width=8); tax_e.grid(row=2, column=1, padx=4)
        ttk.Label(f, text='Default Discount (%)').grid(row=3, column=0, sticky='w', padx=4, pady=4)
        disc_e = ttk.Entry(f, width=8); disc_e.grid(row=3, column=1, padx=4)
        def _get(key, default=''):
            rows = self.db.query('SELECT value FROM settings WHERE key=?;', (key,))
            return rows[0]['value'] if rows else default
        name_e.insert(0, _get('pharmacy_name','Pharmacy Receipt'))
        addr_e.insert(0, _get('pharmacy_address','123 Main Street, City'))
        tax_e.insert(0, _get('tax_percent','0'))
        disc_e.insert(0, _get('default_discount','0'))
        def save():
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('pharmacy_name', name_e.get().strip()))
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('pharmacy_address', addr_e.get().strip()))
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('tax_percent', tax_e.get().strip()))
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('default_discount', disc_e.get().strip()))
            messagebox.showinfo('Saved','Settings saved')
        ttk.Button(f, text='Save Settings', command=save).grid(row=4, column=0, columnspan=2, pady=8)

        def save():
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('tax_percent', tax_e.get().strip()))
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('default_discount', disc_e.get().strip()))
            messagebox.showinfo('Saved','Settings saved')
        ttk.Button(f, text='Save Settings', command=save).grid(row=3, column=0, columnspan=2, pady=8)

    # ---------------- Helpers ----------------
    def _open_tab_by_name(self, name):
        for i in range(self.nb.index('end')):
            if self.nb.tab(i, option='text') == name:
                self.nb.select(i); return
        messagebox.showinfo('Info', f'Tab {name} not found')

    def _filter_inventory_low_stock(self):
        med_rows = self.db.query("""SELECT p.id,p.name,p.sku,p.unit,c.name as category,m.name as manufacturer,p.sale_price as price,
            COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id WHERE p.is_medical=1 GROUP BY p.id HAVING stock<=5 ORDER BY p.name;""")
        self._med_tree.delete(*self._med_tree.get_children())
        for r in med_rows: self._med_tree.insert('', 'end', iid=r['id'], values=(r['id'], r['name'], r['sku'] or '', r.get('unit','') or '', r.get('category') or '', r.get('manufacturer') or '', f"{r['price']:.2f}", r['stock']))
        non_rows = self.db.query("""SELECT p.id,p.name,p.sku,p.unit,c.name as category,m.name as manufacturer,p.sale_price as price,
            COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id WHERE p.is_medical=0 GROUP BY p.id HAVING stock<=5 ORDER BY p.name;""")
        self._nonmed_tree.delete(*self._nonmed_tree.get_children())
        for r in non_rows: self._nonmed_tree.insert('', 'end', iid=r['id'], values=(r['id'], r['name'], r['sku'] or '', r.get('unit','') or '', r.get('category') or '', r.get('manufacturer') or '', f"{r['price']:.2f}", r['stock']))
        self._open_tab_by_name('Inventory')
        try:
            for child in self.tab_inventory.winfo_children():
                if isinstance(child, ttk.Notebook):
                    child.select(0); break
        except Exception:
            pass

    def _filter_inventory_near_expiry(self):
        self._batch_tree.delete(*self._batch_tree.get_children())
        rows = self.db.query("""SELECT b.id, p.name as product, b.batch_no, b.quantity, b.expiry_date, s.name as supplier FROM batches b LEFT JOIN products p ON p.id=b.product_id LEFT JOIN suppliers s ON s.id=b.supplier_id WHERE b.expiry_date IS NOT NULL AND julianday(b.expiry_date)-julianday('now')<=30 AND b.quantity>0 ORDER BY b.expiry_date ASC;""")
        for r in rows: self._batch_tree.insert('', 'end', values=(r['id'], r['product'], r['batch_no'] or '', r['quantity'], r['expiry_date'] or '', r['supplier'] or ''))
        self._open_tab_by_name('Inventory')
        try:
            for child in self.tab_inventory.winfo_children():
                if isinstance(child, ttk.Notebook):
                    nb = child; nb.select(nb.index('end')-1); break
        except Exception:
            pass

    def _open_low_stock(self):
        try:
            self._filter_inventory_low_stock()
        except Exception:
            rows = self.db.query("""SELECT p.id,p.name,COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p GROUP BY p.id HAVING stock<=5;""")
            dlg = tk.Toplevel(self.root); dlg.title('Low Stock Items'); dlg.geometry('700x400')
            tree = ttk.Treeview(dlg, columns=('id','name','stock'), show='headings'); tree.pack(fill='both', expand=True, padx=8, pady=8)
            for c in ('id','name','stock'): tree.heading(c, text=c.capitalize()); tree.column(c,width=200,anchor='w')
            for r in rows: tree.insert('', 'end', values=(r['id'], r['name'], r['stock']))
            ttk.Button(dlg, text='Close', command=dlg.destroy).pack(pady=6)

    def _open_near_expiry(self):
        try:
            self._filter_inventory_near_expiry()
        except Exception:
            rows = self.db.query("""SELECT b.id, p.name AS product, b.batch_no, b.quantity, b.expiry_date FROM batches b JOIN products p ON p.id=b.product_id WHERE b.expiry_date IS NOT NULL AND julianday(b.expiry_date)-julianday('now')<=30 AND b.quantity>0 ORDER BY b.expiry_date ASC;""")
            dlg = tk.Toplevel(self.root); dlg.title('Near Expiry (<=30 days)'); dlg.geometry('800x420')
            tree = ttk.Treeview(dlg, columns=('id','product','batch_no','quantity','expiry'), show='headings'); tree.pack(fill='both', expand=True, padx=8, pady=8)
            for c in ('id','product','batch_no','quantity','expiry'): tree.heading(c, text=c.capitalize()); tree.column(c,width=150,anchor='w')
            for r in rows: tree.insert('', 'end', values=(r['id'], r['product'], r['batch_no'] or '', r['quantity'], r['expiry_date'] or ''))
            ttk.Button(dlg, text='Close', command=dlg.destroy).pack(pady=6)

    # autocomplete helpers
    def _supplier_suggestions(self, term):
        rows = self.db.query('SELECT name FROM suppliers WHERE name LIKE ? ORDER BY name LIMIT 10;', (f'%{term}%',))
        return [r['name'] for r in rows]

    def _manufacturer_suggestions(self, term):
        rows = self.db.query('SELECT name FROM manufacturers WHERE name LIKE ? ORDER BY name LIMIT 10;', (f'%{term}%',))
        return [r['name'] for r in rows]

    def _product_suggestions(self, term):
        rows = self.db.query('SELECT name, sale_price FROM products WHERE name LIKE ? ORDER BY name LIMIT 12;', (f'%{term}%',))
        return [r['name'] for r in rows]


    # ---------------- Seeder ----------------
    def _seed_test_data(self):
        try:
            cnt = self.db.query('SELECT COUNT(*) AS c FROM products;')[0]['c']
            if cnt > 0: return
            man1 = self.db.execute('INSERT OR IGNORE INTO manufacturers(name,contact) VALUES(?,?);', ('GoodPharma','contact1'))
            sup1 = self.db.execute('INSERT OR IGNORE INTO suppliers(name,phone) VALUES(?,?);', ('SupplyCo','1234567890'))
            self.db.execute('INSERT OR IGNORE INTO formulas(name,composition) VALUES(?,?);', ('Paracetamol 500mg','Paracetamol'))
            self.db.execute('INSERT OR IGNORE INTO categories(name) VALUES(?);', ('Analgesics',))
            pid1 = self.db.execute('INSERT INTO products(name,sku,is_medical,unit,sale_price) VALUES(?,?,?,?,?);', ('Paracetamol 500mg','PARA500',1,'tablet',0.50))
            pid2 = self.db.execute('INSERT INTO products(name,sku,is_medical,unit,sale_price) VALUES(?,?,?,?,?);', ('Ibuprofen 200mg','IBU200',1,'tablet',0.75))
            pid3 = self.db.execute('INSERT INTO products(name,sku,is_medical,unit,sale_price) VALUES(?,?,?,?,?);', ('Cough Syrup 100ml','COUGH100',1,'ml',3.50))
            self.db.execute('INSERT INTO batches(product_id,quantity,expiry_date,created_at) VALUES(?,?,?,?);', (pid1,50, (datetime.now()+timedelta(days=20)).strftime('%Y-%m-%d'), now_str()))
            self.db.execute('INSERT INTO batches(product_id,quantity,expiry_date,created_at) VALUES(?,?,?,?);', (pid2,10, (datetime.now()+timedelta(days=200)).strftime('%Y-%m-%d'), now_str()))
            self.db.execute('INSERT INTO batches(product_id,quantity,expiry_date,created_at) VALUES(?,?,?,?);', (pid3,5, (datetime.now()+timedelta(days=10)).strftime('%Y-%m-%d'), now_str()))
            cid = self.db.execute('INSERT INTO customers(name,phone) VALUES(?,?);', ('Alice','5551112222'))
            sale1 = self.db.execute('INSERT INTO sales(user_id,total,customer_id,customer_name,customer_phone,created_at) VALUES(?,?,?,?,?,?);', (1, 15.0, cid, 'Alice','5551112222', now_str()))
            si = self.db.execute('INSERT INTO sale_items(sale_id,product_id,quantity,price) VALUES(?,?,?,?);', (sale1, pid1, 2, 0.5))
            batches = self.db.query('SELECT id,quantity FROM batches WHERE product_id=? ORDER BY created_at ASC;', (pid1,))
            need = 2
            for b in batches:
                take = min(need, b['quantity'])
                if take>0:
                    self.db.execute('UPDATE batches SET quantity=quantity-? WHERE id=?;', (take, b['id']))
                    need -= take
                if need<=0: break
        except Exception as e:
            print('Seeder error', e)

    # ---------------- End App ----------------
    def run(self):
        self.root.mainloop()



if __name__ == '__main__':
    app = App()
   
    app.run()
