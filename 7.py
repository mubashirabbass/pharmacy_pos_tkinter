

import os
import sqlite3
import hashlib
import csv
import threading
import time
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# Optional libs
try:
    import ttkbootstrap as tb
    from ttkbootstrap.icons import Icon
    TTB_AVAILABLE = True
except Exception:
    tb = None
    Icon = None
    TTB_AVAILABLE = False
try:
    from tkcalendar import DateEntry as _TKDateEntry
    # wrapper to avoid ttkbootstrap conflicts
    class DateEntry(_TKDateEntry):
        def __init__(self, master=None, **kw):
            if 'bootstyle' in kw:
                kw.pop('bootstyle')
            super().__init__(master, **kw)
    TKCAL_AVAILABLE = True
except Exception:
    DateEntry = None
    TKCAL_AVAILABLE = False

try:
    import matplotlib
    # matplotlib.use('Agg')  # disabled to allow TkAgg
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_AVAILABLE = True
except Exception:
    MATPLOTLIB_AVAILABLE = False


try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as pdf_canvas
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# paths and DB
BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE_DIR, 'pharmacy.db')
BACKUP_FOLDER = os.path.join(BASE_DIR, 'backups')
os.makedirs(BACKUP_FOLDER, exist_ok=True)
RECEIPT_FOLDER = os.path.join(BASE_DIR, 'receipts')
os.makedirs(RECEIPT_FOLDER, exist_ok=True)


# -----------------------
# Utilities & DB Layer
# -----------------------
def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def now_str(): return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

class DB:

    def __init__(self, path=DB_PATH):
        self.path = path
        self._ensure()

    def connect(self):
        con = sqlite3.connect(self.path, detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
        con.row_factory = sqlite3.Row
        con.execute('PRAGMA foreign_keys = ON;')
        return con

    def _ensure(self):
        con = self.connect(); cur = con.cursor()

        # users - FIXED: Added 'staff' to the CHECK constraint
        cur.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT CHECK(role IN ('admin','staff','cashier')) NOT NULL
        );''')
        cur.execute('SELECT COUNT(*) as c FROM users;')
        if cur.fetchone()['c'] == 0:
            cur.executemany('INSERT INTO users(username,password_hash,role) VALUES(?,?,?);', [
                ('admin', hash_pw('admin123'), 'admin'),
                ('cashier', hash_pw('cashier123'), 'cashier'),
            ])

        # customers
        cur.execute('''CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            phone TEXT UNIQUE,
            notes TEXT
        );''')

        # categories/manufacturers/suppliers/formulas
        cur.execute('''CREATE TABLE IF NOT EXISTS categories (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, notes TEXT);''')
        cur.execute('''CREATE TABLE IF NOT EXISTS manufacturers (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, contact TEXT, notes TEXT);''')
        cur.execute('''CREATE TABLE IF NOT EXISTS suppliers (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, phone TEXT, email TEXT, address TEXT);''')
        cur.execute('''CREATE TABLE IF NOT EXISTS formulas (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, composition TEXT);''')

        # products
        cur.execute('''CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            sku TEXT UNIQUE,
            is_medical INTEGER DEFAULT 1,
            category_id INTEGER,
            manufacturer_id INTEGER,
            formula_id INTEGER,
            unit TEXT,
            sale_price REAL DEFAULT 0,
            notes TEXT
        );''')

        # batches
        cur.execute('''CREATE TABLE IF NOT EXISTS batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            supplier_id INTEGER,
            batch_no TEXT,
            quantity INTEGER NOT NULL,
            expiry_date TEXT,
            cost_price REAL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
            FOREIGN KEY(supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL
        );''')

        # --- Ensure manufacturer_id exists in batches for existing DB ---
        cur.execute("PRAGMA table_info(batches);")
        columns = [c['name'] for c in cur.fetchall()]
        if 'manufacturer_id' not in columns:
            # Add manufacturer_id column
            cur.execute("ALTER TABLE batches ADD COLUMN manufacturer_id INTEGER;")
            # Update table with proper foreign key
            cur.execute("PRAGMA foreign_keys=off;")
            cur.execute("""CREATE TABLE IF NOT EXISTS batches_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                supplier_id INTEGER,
                manufacturer_id INTEGER,
                batch_no TEXT,
                quantity INTEGER NOT NULL,
                expiry_date TEXT,
                cost_price REAL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE,
                FOREIGN KEY(supplier_id) REFERENCES suppliers(id) ON DELETE SET NULL,
                FOREIGN KEY(manufacturer_id) REFERENCES manufacturers(id) ON DELETE SET NULL
            );""")
            cur.execute("INSERT INTO batches_new SELECT id, product_id, supplier_id, NULL, batch_no, quantity, expiry_date, cost_price, created_at FROM batches;")
            cur.execute("DROP TABLE batches;")
            cur.execute("ALTER TABLE batches_new RENAME TO batches;")
            cur.execute("PRAGMA foreign_keys=on;")

        # sales & items
        cur.execute('''CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            total REAL NOT NULL,
            customer_id INTEGER,
            customer_name TEXT,
            customer_phone TEXT,
            discount REAL DEFAULT 0,
            tax REAL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL,
            FOREIGN KEY(customer_id) REFERENCES customers(id) ON DELETE SET NULL
        );''')
        cur.execute('''CREATE TABLE IF NOT EXISTS sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            price REAL NOT NULL,
            FOREIGN KEY(sale_id) REFERENCES sales(id) ON DELETE CASCADE,
            FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
        );''')
        cur.execute('''CREATE TABLE IF NOT EXISTS sale_item_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_item_id INTEGER NOT NULL,
            batch_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            FOREIGN KEY(sale_item_id) REFERENCES sale_items(id) ON DELETE CASCADE,
            FOREIGN KEY(batch_id) REFERENCES batches(id) ON DELETE CASCADE
        );''')
        cur.execute('''CREATE TABLE IF NOT EXISTS returns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_item_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            reason TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(sale_item_id) REFERENCES sale_items(id) ON DELETE CASCADE
        );''')

        # settings
        cur.execute('''CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);''')
        def set_if_missing(k,v):
            cur.execute('SELECT value FROM settings WHERE key=?;',(k,))
            if not cur.fetchone(): cur.execute('INSERT INTO settings(key,value) VALUES(?,?);',(k,str(v)))
        set_if_missing('tax_percent','0.0')
        set_if_missing('default_discount','0.0')
        set_if_missing('pharmacy_name','Pharmacy Receipt')
        set_if_missing('pharmacy_address','123 Main Street, City')
        set_if_missing('auto_backup_enabled','0')

        con.commit(); con.close()


    def query(self, sql, params=()):
        with self.connect() as con:
            cur = con.execute(sql, params)
            return [dict(r) for r in cur.fetchall()]

    def execute(self, sql, params=()):
        with self.connect() as con:
            cur = con.execute(sql, params)
            con.commit()
            return cur.lastrowid

db = DB()


# ------------------------
# Widgets: Autocomplete, FormDialog
# ------------------------

# ------------------------
# Widgets: Autocomplete, FormDialog
# ------------------------

# ------------------------
# Widgets: Autocomplete, FormDialog
# ------------------------

class AutocompleteEntry(ttk.Entry):
    def __init__(self, master, suggestions_getter=None, width=30, **kwargs):
        super().__init__(master, width=width, **kwargs)
        self.suggestions_getter = suggestions_getter
        self.var = tk.StringVar()
        self.config(textvariable=self.var)
        self.var.trace_add('write', self._on_change)
        self.listbox = None
        self.bind('<Down>', self._focus_listbox)
        self.bind('<Escape>', lambda e: self._hide())
        self.bind('<Return>', self._select_first)

    def _on_change(self, *args):
        term = self.var.get().strip()
        if not term:
            self._hide(); return
        try:
            suggestions = self.suggestions_getter(term) if self.suggestions_getter else []
        except Exception:
            suggestions = []
        if not suggestions:
            self._hide(); return
        if not self.listbox:
            self.listbox = tk.Listbox(self.master, height=6)
            self.listbox.bind('<<ListboxSelect>>', self._on_select)
            self.listbox.bind('<Return>', self._on_select)
            self.listbox.bind('<Escape>', lambda e: self._hide())
        self.listbox.delete(0, 'end')
        for s in suggestions:
            self.listbox.insert('end', s)
        x = self.winfo_rootx() - self.master.winfo_rootx()
        y = self.winfo_rooty() - self.master.winfo_rooty() + self.winfo_height()
        self.listbox.place(x=x, y=y, width=self.winfo_width())

    def _on_select(self, event=None):
        if not self.listbox: return
        sel = self.listbox.curselection()
        if not sel: return
        val = self.listbox.get(sel[0])
        self.var.set(val)
        self._hide()

    def _focus_listbox(self, event=None):
        if self.listbox:
            self.listbox.focus_set()
            self.listbox.selection_set(0)

    def _hide(self):
        if self.listbox:
            self.listbox.destroy()
            self.listbox = None

    def _select_first(self, event=None):
        if self.listbox and self.listbox.size() > 0:
            val = self.listbox.get(0)
            self.var.set(val)
            self._hide()


class PlaceholderEntry(ttk.Entry):
    def __init__(self, master, placeholder='', textvariable=None, width=30, **kw):
        if textvariable is None:
            self.var = tk.StringVar()
            kw['textvariable'] = self.var
        else:
            self.var = textvariable
            kw['textvariable'] = self.var
        super().__init__(master, width=width, **kw)
        self.placeholder = placeholder
        self._has_placeholder = False
        self._placeholder_color = '#6c757d'
        try:
            self._normal_color = self.cget('foreground') or 'black'
        except Exception:
            self._normal_color = 'black'
        self.bind('<FocusIn>', self._clear_placeholder)
        self.bind('<FocusOut>', self._set_placeholder)
        self._set_placeholder()

    def _set_placeholder(self, event=None):
        if not self.var.get():
            self._has_placeholder = True
            try:
                self.config(foreground=self._placeholder_color)
            except Exception:
                pass
            self.var.set(self.placeholder)

    def _clear_placeholder(self, event=None):
        if self._has_placeholder:
            self._has_placeholder = False
            try:
                self.config(foreground=self._normal_color)
            except Exception:
                pass
            self.var.set('')





class FormDialog(tk.Toplevel):
    def __init__(self, master, title, fields, on_submit=None, initial=None):
        super().__init__(master)
        self.title(title)
        self.on_submit = on_submit
        self.result = None
        self.widgets = {}

        frm = ttk.Frame(self, padding=8)
        frm.pack(fill='both', expand=True)

        for i, f in enumerate(fields):
            ttk.Label(frm, text=f.get('label', f['key'])).grid(row=i, column=0, sticky='w', pady=4)

            widget_type = f.get('widget', 'entry')
            if widget_type == 'autocomplete':
                w = AutocompleteEntry(frm, suggestions_getter=f.get('suggestions_getter'))
            elif widget_type == 'combobox':
                state = f.get('state', 'normal')
                w = ttk.Combobox(frm, values=f.get('values', []), state=state)
            elif widget_type == 'spinbox':
                w = ttk.Spinbox(frm, from_=f.get('from', 0), to=f.get('to', 999999), increment=f.get('inc', 1))
            elif widget_type == 'text':
                w = tk.Text(frm, height=f.get('height', 3), width=f.get('width', 40))
            else:
                w = ttk.Entry(frm)

            w.grid(row=i, column=1, sticky='we', pady=4)

            if initial and f['key'] in initial and initial[f['key']] is not None:
                val = initial[f['key']]
                if widget_type == 'text':
                    try:
                        w.insert('1.0', str(val))
                    except Exception:
                        pass
                else:
                    try:
                        w.insert(0, str(val))
                    except Exception:
                        pass

            self.widgets[f['key']] = (w, f)

        btns = ttk.Frame(frm)
        btns.grid(row=len(fields), column=0, columnspan=2, pady=8)
        ttk.Button(btns, text='Save', command=self._save).pack(side='left', padx=6)
        ttk.Button(btns, text='Cancel', command=self.destroy).pack(side='left')
        self.bind('<Return>', lambda e: self._save())
        self.bind('<Escape>', lambda e: self.destroy())

    def _save(self):
        data = {}
        for key, (w, f) in self.widgets.items():
            widget = f.get('widget', 'entry')
            if widget == 'text':
                try:
                    data[key] = w.get('1.0', 'end').strip()
                except Exception:
                    data[key] = ''
            else:
                try:
                    data[key] = w.get().strip()
                except Exception:
                    data[key] = ''
        self.result = data
        if self.on_submit:
            self.on_submit(data)
        self.destroy()

class NewSaleTab(ttk.Frame):
    def __init__(self, master, db, user):
        super().__init__(master)
        self.db, self.user = db, user
        self.cart = []
        self.selected_product = None
        self._build()

    def _build(self):
        cust = ttk.Frame(self); cust.pack(fill='x', padx=10, pady=6)
        ttk.Label(cust, text="Customer Name").pack(side='left')
        self.customer_name_e = ttk.Entry(cust, width=30); self.customer_name_e.pack(side='left', padx=6)
        ttk.Label(cust, text="Phone").pack(side='left')
        self.customer_phone_e = ttk.Entry(cust, width=20); self.customer_phone_e.pack(side='left', padx=6)

        top = ttk.Frame(self); top.pack(fill='x', padx=10, pady=6)
        ttk.Label(top, text="Search by name or ID").pack(side='left')
        self.search_e = AutocompleteEntry(top, suggestions_getter=lambda term: [f"{r['id']} - {r['name']} - {r['sale_price']}" for r in self.db.query("SELECT id,name,sale_price FROM products WHERE name LIKE ? OR CAST(id AS TEXT) LIKE ? ORDER BY name LIMIT 50;", (f"%{term}%",f"%{term}%"))]); self.search_e.pack(side='left', padx=6)
        # keep update_suggestions compatibility
        self.search_e.var.trace_add('write', lambda *a: self.update_suggestions())
        ttk.Label(top, text="Qty").pack(side='left', padx=(10,0))
        self.qty_e = ttk.Entry(top, width=6); self.qty_e.pack(side='left', padx=6)
        ttk.Button(top, text="Add", command=self.add_to_cart).pack(side='left')

        self.suggestions = tk.Listbox(self, height=6)
        self.suggestions.pack(fill='x', padx=10)
        self.suggestions.bind("<Double-Button-1>", self._on_suggestion_double)

        self.tree = ttk.Treeview(self, columns=['product','qty','price','subtotal'], show='headings')
        for c in ['product','qty','price','subtotal']:
            self.tree.heading(c, text=c.capitalize())
            self.tree.column(c, anchor='center')
        self.tree.pack(fill='both', expand=True, padx=10, pady=6)

        btns = ttk.Frame(self); btns.pack(fill='x', padx=10, pady=4)
        ttk.Button(btns, text="Remove Selected", command=self.remove_selected).pack(side='left')

        self.lbl_total = ttk.Label(self, text="Total: 0.00", font=('Segoe UI', 12, 'bold'))
        self.lbl_total.pack(anchor='e', padx=10)
        ttk.Button(self, text="Checkout", command=self.checkout).pack(anchor='e', padx=10, pady=6)

    def update_suggestions(self, event=None):
        term = self.search_e.get().strip()
        self.suggestions.delete(0, 'end')
        if not term:
            return
        rows = self.db.query("SELECT id, name, sale_price FROM products WHERE name LIKE ? OR CAST(id AS TEXT) LIKE ? ORDER BY name LIMIT 50;", (f"%{term}%", f"%{term}%"))
        for r in rows:
            self.suggestions.insert('end', f"{r['id']} - {r['name']} - {r['sale_price']}")

    def _on_suggestion_double(self, event=None):
        sel = self.suggestions.curselection()
        if not sel:
            return
        val = self.suggestions.get(sel[0])
        pid = int(val.split(' - ')[0])
        row = self.db.query("SELECT * FROM products WHERE id=?;", (pid,))
        if row:
            self.selected_product = row[0]
            self.search_e.delete(0, 'end')
            self.search_e.insert(0, f"{self.selected_product['name']}")

    def add_to_cart(self):
        term = self.search_e.get().strip()
        try:
            qty = int(self.qty_e.get() or 0)
        except ValueError:
            qty = 0
        prod = None
        if term.isdigit():
            rows = self.db.query("SELECT * FROM products WHERE id=?;", (int(term),))
            if rows: prod = rows[0]
        if not prod:
            rows = self.db.query("SELECT * FROM products WHERE name=? LIMIT 1;", (term,))
            if rows: prod = rows[0]
        if not prod:
            messagebox.showwarning("Not found", "Product not found. Use search box and double-click a suggestion.")
            return
        if qty <= 0:
            messagebox.showwarning("Invalid qty", "Enter quantity > 0")
            return
        self.cart.append({'id': prod['id'], 'name': prod['name'], 'qty': qty, 'price': prod['sale_price'], 'subtotal': prod['sale_price']*qty})
        self.search_e.delete(0, 'end'); self.qty_e.delete(0, 'end'); self.refresh()

    def remove_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.cart[idx]
        self.refresh()

    def refresh(self):
        self.tree.delete(*self.tree.get_children())
        total = 0
        for item in self.cart:
            self.tree.insert('', 'end', values=[item['name'], item['qty'], item['price'], item['subtotal']])
            total += item['subtotal']
        self.lbl_total.config(text=f"Total: {total:.2f}")

    def checkout(self):
        if not self.cart:
            messagebox.showwarning("Empty", "Cart is empty")
            return
        cust_name = self.customer_name_e.get().strip()
        cust_phone = self.customer_phone_e.get().strip()
        total = sum(i['subtotal'] for i in self.cart)
        
        # Get current timestamp
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        sid = self.db.execute(
            "INSERT INTO sales(user_id,total,customer_name,customer_phone,created_at) VALUES(?,?,?,?,?);", 
            (self.user['id'], total, cust_name, cust_phone, created_at)
        )
        
        for i in self.cart:
            # Insert sale item
            sale_item_id = self.db.execute(
                "INSERT INTO sale_items(sale_id,product_id,quantity,price) VALUES(?,?,?,?);", 
                (sid, i['id'], i['qty'], i['price'])
            )
            
            # Update batches and create sale_item_batches entries
            self._fifo_deduct_with_batch_tracking(i['id'], i['qty'], sale_item_id, i['name'])
        
        if messagebox.askyesno("Print Receipt", "Do you want to print a receipt?"):
            self.generate_receipt(sid, total, cust_name, cust_phone)
        
        messagebox.showinfo("Sale Complete", f"Sale #{sid} completed.")
        self.cart.clear()
        self.refresh()

    def _fifo_deduct_with_batch_tracking(self, product_id, qty_needed, sale_item_id, pname):
        batches = self.db.query(
            "SELECT id, quantity FROM batches WHERE product_id=? AND quantity>0 ORDER BY created_at ASC;", 
            (product_id,)
        )
        remain = qty_needed
        
        for b in batches:
            if remain <= 0:
                break
            take = min(remain, b['quantity'])
            
            # Update batch quantity
            self.db.execute("UPDATE batches SET quantity=quantity-? WHERE id=?;", (take, b['id']))
            
            # Create sale_item_batches entry
            self.db.execute(
                "INSERT INTO sale_item_batches(sale_item_id, batch_id, quantity) VALUES(?,?,?);",
                (sale_item_id, b['id'], take)
            )
            
            remain -= take
        
        if remain > 0:
            messagebox.showwarning("Stock Warning", f"Product {pname} had insufficient stock. Short by {remain}.")
    

    def generate_receipt(self, sale_id, total, cust_name, cust_phone):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas as pdf_canvas
        except Exception:
            messagebox.showerror("Missing Package", "reportlab not installed; cannot generate PDF.")
            return
        settings = {r['key']:r['value'] for r in self.db.query("SELECT key,value FROM settings;")}
        pharmacy_name = settings.get('pharmacy_name','Pharmacy Receipt')
        pharmacy_address = settings.get('pharmacy_address','')
        items = self.db.query("SELECT si.quantity, si.price, p.name FROM sale_items si JOIN products p ON si.product_id=p.id WHERE si.sale_id=?;", (sale_id,))
        folder = os.path.join(os.path.dirname(__file__), "receipts"); os.makedirs(folder, exist_ok=True)
        filepath = os.path.join(folder, f"receipt_{sale_id}.pdf")
        c = pdf_canvas.Canvas(filepath, pagesize=A4); width, height = A4
        y = height - 60
        c.setFont("Helvetica-Bold", 16); c.drawCentredString(width/2, y, pharmacy_name); y -= 18
        if pharmacy_address:
            c.setFont("Helvetica", 10); c.drawCentredString(width/2, y, pharmacy_address); y -= 16
        c.line(40, y, width-40, y); y -= 14
        c.setFont("Helvetica", 10)
        c.drawString(40, y, f"Sale ID: {sale_id}"); c.drawRightString(width-40, y, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"); y -= 12
        c.drawString(40, y, f"Cashier: {self.user['username']}"); y -= 12
        if cust_name: c.drawString(40, y, f"Customer: {cust_name}"); y -= 12
        if cust_phone: c.drawString(40, y, f"Phone: {cust_phone}"); y -= 12
        c.line(40, y, width-40, y); y -= 14
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, "Item"); c.drawRightString(width-200, y, "Qty"); c.drawRightString(width-120, y, "Price"); c.drawRightString(width-40, y, "Subtotal"); y -= 12
        c.setFont("Helvetica", 10)
        for it in items:
            c.drawString(40, y, str(it['name']))
            c.drawRightString(width-200, y, str(it['quantity']))
            c.drawRightString(width-120, y, f"{it['price']:.2f}")
            c.drawRightString(width-40, y, f"{it['price']*it['quantity']:.2f}")
            y -= 12
            if y < 80:
                c.showPage(); y = height - 60
        c.line(40, y, width-40, y); y -= 16
        c.setFont("Helvetica-Bold", 12); c.drawRightString(width-40, y, f"TOTAL: {total:.2f}"); y -= 24
        c.setFont("Helvetica", 10); c.drawCentredString(width/2, y, "Thank you for shopping!")
        c.save()
        try:
            os.startfile(filepath)
        except Exception:
            pass


class App:
    def __init__(self):
        if TTB_AVAILABLE:
            self.root = tb.Window(themename='flatly')
        else:
            self.root = tk.Tk()
        self.root.title('Pharmacy Management System')
        self.root.geometry('1200x780')
        self.root.protocol('WM_DELETE_WINDOW', self._on_close)

        self.db = db
        self.user = None
        self._auto_job = None
        self._build_login()

    def _on_close(self):
        try:
            if getattr(self, '_auto_job', None):
                self.root.after_cancel(self._auto_job)
        except Exception:
            pass
        self.root.quit()
        self.root.destroy()

    def _logout(self):
        """Log out and return to login screen."""
        self.user = None
        self._build_login()

    def _open_profile(self):
        def save(d):
            pw = d.get('new_password','').strip()
            if pw:
                self.db.execute(
                    'UPDATE users SET password_hash=? WHERE id=?;',
                    (hash_pw(pw), self.user['id'])
                )
                messagebox.showinfo('Profile','Password updated.')
        FormDialog(
            self.root, 'Profile - Change Password',
            [
                {'key':'username','label':'Username','widget':'entry'},
                {'key':'role','label':'Role','widget':'entry'},
                {'key':'new_password','label':'New Password','widget':'entry'},
            ],
            initial={'username':self.user['username'], 'role':self.user['role']},
            on_submit=save
        )

    # ---------------- Main ----------------
    def _build_main(self):
        for w in self.root.winfo_children():
            w.destroy()

        top = ttk.Frame(self.root)
        top.pack(fill='x')
        ttk.Label(top, text=f"Welcome, {self.user['username'].title()}",
                  font=('Segoe UI',14,'bold')).pack(side='left', padx=10, pady=8)
        ttk.Button(top, text='Profile', command=self._open_profile).pack(side='right', padx=6)
        ttk.Button(top, text='Logout', command=self._logout).pack(side='right')

        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill='both', expand=True, padx=8, pady=8)

        if self.user['role'] == 'admin':
            self.tab_dashboard = ttk.Frame(self.nb); self.nb.add(self.tab_dashboard, text='Dashboard')
            self.tab_inventory = ttk.Frame(self.nb); self.nb.add(self.tab_inventory, text='Inventory')
            self.tab_pos = ttk.Frame(self.nb); self.nb.add(self.tab_pos, text='POS')
            self.tab_manage_staff = ttk.Frame(self.nb); self.nb.add(self.tab_manage_staff, text='Manage Staff')
            self.tab_settings = ttk.Frame(self.nb); self.nb.add(self.tab_settings, text='Settings')

        elif self.user['role'] == 'staff':
            self.tab_inventory = ttk.Frame(self.nb); self.nb.add(self.tab_inventory, text='Inventory')
            self.tab_pos = ttk.Frame(self.nb); self.nb.add(self.tab_pos, text='POS')

        elif self.user['role'] == 'cashier':
            self.tab_pos = ttk.Frame(self.nb); self.nb.add(self.tab_pos, text='POS')

        # Build relevant tabs
        if self.user['role'] == 'admin':
            self._build_dashboard_tab()
            self._build_inventory_tab()
            self._build_pos_tab()
            self._build_sale_history_tab()
            self._build_return_history_tab()
            self._build_manage_staff_tab()

        elif self.user['role'] == 'staff':
            self._build_inventory_tab()
            self._build_pos_tab()

        elif self.user['role'] == 'cashier':
            self._build_pos_tab()

    # ---------------- Staff Add Fix ----------------
    def _add_staff(self):
        def save(d):
            if not d.get('username') or not d.get('password'):
                return messagebox.showerror('Error','Username and password required')
            existing = self.db.query("SELECT id FROM users WHERE username=?;", (d['username'],))
            if existing:
                return messagebox.showerror('Error','Username already exists')
            self.db.execute(
                "INSERT INTO users(username,password_hash,role) VALUES(?,?,?)",
                (d['username'], hash_pw(d['password']), d['role'])
            )
            messagebox.showinfo('Saved','Staff added successfully')
            self._build_manage_staff_tab()

        FormDialog(self.root, 'Add Staff', [
            {'key':'username','label':'Username'},
            {'key':'password','label':'Password'},
            {'key':'role','label':'Role','widget':'combobox','values':['staff','cashier']}
        ], on_submit=save)


    # ---------------- Login ----------------
    def _build_login(self):
        for w in self.root.winfo_children(): w.destroy()
        frm = ttk.Frame(self.root, padding=20); frm.pack(expand=True)

        # Pharmacy Name + Logo
        top = ttk.Frame(frm); top.grid(row=0, column=0, columnspan=2, pady=(0,20))
        try:
            logo_img = tk.PhotoImage(file=os.path.join(BASE_DIR, "logo.png"))
            logo_lbl = ttk.Label(top, image=logo_img)
            logo_lbl.image = logo_img
            logo_lbl.pack()
        except Exception:
            ttk.Label(top, text='🏥', font=('Segoe UI', 40)).pack()

        ttk.Label(top, text='بایو فینِکس فارمیسی', font=('Nori Nastaleeq', 22, 'bold')).pack()
        ttk.Label(top, text='[-Login Portal-]', font=('Segoe UI', 9,'bold')).pack()
        from tkinter import font
    
        bold_font = font.Font(weight="bold")
        # Login Form
        ttk.Label(frm, text='Login As:', font=bold_font).grid(row=1, column=0, sticky='e')
        role_cb = ttk.Combobox(frm, values=['admin','staff','cashier'], state='readonly')
        role_cb.set('admin')
        role_cb.grid(row=1, column=1, sticky='w', pady=4)

        ttk.Label(frm, text='Username', font=bold_font).grid(row=2, column=0, sticky='e')
        user_e = ttk.Entry(frm); user_e.grid(row=2, column=1, sticky='w', pady=4)

        ttk.Label(frm, text='Password', font=bold_font).grid(row=3, column=0, sticky='e')
        pw_e = ttk.Entry(frm, show='•'); pw_e.grid(row=3, column=1, sticky='w', pady=4)


        def try_login():

            u = user_e.get().strip(); p = pw_e.get().strip(); r = role_cb.get().strip()
            if not u or not p: return messagebox.showerror('Error','Enter username & password')
            rows = self.db.query('SELECT * FROM users WHERE username=?;',(u,))
            if not rows or rows[0]['password_hash'] != hash_pw(p) or rows[0]['role'] != r:
                return messagebox.showerror('Error','Invalid credentials or role')
            self.user = {'id':rows[0]['id'],'username':u,'role':rows[0]['role']}
            self._build_main()
        ttk.Button(frm, text='Login', command=try_login).grid(row=4, column=0, columnspan=2, pady=8)
        user_e.focus_set()
        self.root.bind('<Return>', lambda e: try_login())

    # ---------------- Main ----------------
    def _build_main(self):
        # Clear
        for w in self.root.winfo_children():
            w.destroy()

        # Top bar
        top = ttk.Frame(self.root); top.pack(fill='x')
        ttk.Label(top, text=f"Welcome, {self.user['username'].title()}", font=('Segoe UI',14,'bold')).pack(side='left', padx=10, pady=8)
        ttk.Button(top, text='Profile', command=self._open_profile).pack(side='right', padx=6)
        ttk.Button(top, text='Logout', command=self._logout).pack(side='right')

        # Notebook (tabs)
        self.nb = ttk.Notebook(self.root); self.nb.pack(fill='both', expand=True, padx=8, pady=8)

        role = self.user.get('role')
        # Create tabs based on role
        if role == 'admin':
            self.tab_dashboard = ttk.Frame(self.nb); self.nb.add(self.tab_dashboard, text='Dashboard')
            self.tab_inventory = ttk.Frame(self.nb); self.nb.add(self.tab_inventory, text='Inventory')
            self.tab_pos = ttk.Frame(self.nb); self.nb.add(self.tab_pos, text='POS')
            self.tab_manage_staff = ttk.Frame(self.nb); self.nb.add(self.tab_manage_staff, text='Manage Staff')
            self.tab_settings = ttk.Frame(self.nb); self.nb.add(self.tab_settings, text='Settings')
        elif role == 'staff':
            self.tab_inventory = ttk.Frame(self.nb); self.nb.add(self.tab_inventory, text='Inventory')
            self.tab_pos = ttk.Frame(self.nb); self.nb.add(self.tab_pos, text='POS')
        elif role == 'cashier':
            self.tab_pos = ttk.Frame(self.nb); self.nb.add(self.tab_pos, text='POS')

        # Build tab contents
        if role == 'admin':
            self._build_dashboard_tab()
            self._build_inventory_tab()
            self._build_pos_tab()
            self._build_sale_history_tab()
            self._build_return_history_tab()
            self._build_manage_staff_tab()
            self._build_settings_tab()
        elif role == 'staff':
            self._build_inventory_tab()
            self._build_pos_tab()
        elif role == 'cashier':
            self._build_pos_tab()

# ---------------- Dashboard ----------------
    def _build_dashboard_tab(self):
        """Modern admin dashboard with interactive stat cards."""
        for w in self.tab_dashboard.winfo_children():
            w.destroy()

        frame = self.tab_dashboard

        # Header
        header = ttk.Frame(frame, padding=(12, 8))
        header.pack(fill='x')
        ttk.Label(header, text=f"👋 Welcome back, {self.user['username'].title()}!", font=('Segoe UI', 20, 'bold'), foreground='#0f5132').pack(anchor='w')
        ttk.Label(header, text="Here's a quick snapshot of your store performance", font=('Segoe UI', 11)).pack(anchor='w', pady=(2,8))

        # Cards container
        cards_row = ttk.Frame(frame)
        cards_row.pack(fill='x', padx=12, pady=6)

        def make_card(parent, title, value, color='#198754', onclick=None):
            card = ttk.Frame(parent, padding=12, relief='raised')
            card.pack(side='left', expand=True, fill='both', padx=8, pady=8)

            top = ttk.Frame(card)
            top.pack(fill='x')
            ttk.Label(top, text=title, font=('Segoe UI', 11, 'bold')).pack(side='left', anchor='w')

            val_lbl = ttk.Label(card, text='0', font=('Segoe UI', 24, 'bold'), foreground=color)
            val_lbl.pack(anchor='center', pady=(8,4))

            ttk.Label(card, text='Click to view details', font=('Segoe UI', 9)).pack(anchor='w')

            # click handler to jump to relevant tab
            def _on_click(e=None):
                if onclick:
                    try:
                        onclick()
                    except Exception:
                        pass
            for w in (card, top, val_lbl):
                w.bind('<Button-1>', _on_click)

            # animate value
            try:
                self._animate_number(val_lbl, value)
            except Exception:
                val_lbl.config(text=str(value))

            return card

        # Date range for current month
        from datetime import datetime, timedelta
        today = datetime.now().date()
        first_day = today.replace(day=1)
        if first_day.month == 12:
            next_month_first = first_day.replace(year=first_day.year + 1, month=1, day=1)
        else:
            next_month_first = first_day.replace(month=first_day.month + 1, day=1)

        # --- Queries ---
        try:
            rows = self.db.query(
                "SELECT COALESCE(SUM(total), 0) AS s FROM sales "
                "WHERE created_at >= ? AND created_at < ?;",
                (first_day.isoformat(), next_month_first.isoformat())
            )
            total_sales = int(rows[0]['s']) if rows else 0
        except Exception:
            total_sales = 0

        try:
            rows = self.db.query(
                "SELECT COUNT(*) AS c FROM ("
                "  SELECT p.id FROM products p "
                "  LEFT JOIN batches b ON b.product_id=p.id "
                "  GROUP BY p.id"
                ") t;"
            )
            total_products = int(rows[0]['c']) if rows else 0
        except Exception:
            total_products = 0

        try:
            in_30 = (today + timedelta(days=30)).isoformat()
            rows = self.db.query(
                "SELECT COUNT(*) AS c FROM batches "
                "WHERE expiry_date IS NOT NULL "
                "AND expiry_date <= ? "
                "AND quantity > 0;",
                (in_30,)
            )
            near_expiry = int(rows[0]['c']) if rows else 0
        except Exception:
            near_expiry = 0

        try:
            rows = self.db.query(
                "SELECT COUNT(*) AS c FROM ("
                "  SELECT p.id, COALESCE(SUM(b.quantity),0) AS stock "
                "  FROM products p "
                "  LEFT JOIN batches b ON b.product_id=p.id "
                "  GROUP BY p.id HAVING stock <= 5"
                ") t;"
            )
            low_stock = int(rows[0]['c']) if rows else 0
        except Exception:
            low_stock = 0

        # Render cards
        make_card(cards_row, '💰 Sales (This Month)', total_sales, color='#198754', onclick=lambda: self.nb.select(self.tab_pos) if hasattr(self, 'tab_pos') else None)
        make_card(cards_row, '📦 Products in Inventory', total_products, color='#0d6efd', onclick=lambda: self.nb.select(self.tab_inventory) if hasattr(self, 'tab_inventory') else None)
        make_card(cards_row, '⏳ Products Near Expiry', near_expiry, color='#ffc107', onclick=lambda: self.nb.select(self.tab_inventory) if hasattr(self, 'tab_inventory') else None)
        make_card(cards_row, '⚠️ Low Stock Items', low_stock, color='#dc3545', onclick=lambda: self.nb.select(self.tab_inventory) if hasattr(self, 'tab_inventory') else None)

        # Quick action buttons
        quick = ttk.Frame(frame)
        quick.pack(fill='x', padx=12, pady=12)
        ttk.Button(quick, text='Open POS', command=lambda: self.nb.select(self.tab_pos)).pack(side='left', padx=6)
        ttk.Button(quick, text='Open Inventory', command=lambda: self.nb.select(self.tab_inventory)).pack(side='left', padx=6)
        if hasattr(self, 'tab_manage_staff'):
            ttk.Button(quick, text='Manage Staff', command=lambda: self.nb.select(self.tab_manage_staff)).pack(side='left', padx=6)

            if MATPLOTLIB_AVAILABLE:
                try:
                    fig = Figure(figsize=(8,2.2), dpi=90); ax = fig.add_subplot(111)
                    days = []; totals = []
                    for i in range(6,-1,-1):
                        d = (datetime.now().date() - timedelta(days=i)).strftime('%Y-%m-%d')
                        days.append(d[5:])
                        r = self.db.query('SELECT COALESCE(SUM(total),0) AS s FROM sales WHERE substr(created_at,1,10)=?;', (d,))
                        totals.append(float(r[0]['s']))
                    ax.plot(days, totals, marker='o'); ax.set_title('Sales — Last 7 days'); ax.grid(True)
                    canvas = FigureCanvasTkAgg(fig, master=frame); canvas.draw(); canvas.get_tk_widget().pack(fill='x', padx=12, pady=10)
                except Exception as e:
                    print('graph error', e)

    # ---------------- Inventory with nested tabs ----------------
    
    def _import_inventory(self, inv_type):
        path = filedialog.askopenfilename(filetypes=[('CSV Files','*.csv'),('Excel Files','*.xlsx')])
        if not path:
            return
        rows = []
        try:
            if path.lower().endswith('.csv'):
                with open(path, newline='', encoding='utf-8') as f:
                    rows = list(csv.DictReader(f))
            elif path.lower().endswith('.xlsx') and OPENPYXL_AVAILABLE:
                from openpyxl import load_workbook
                wb = load_workbook(path); ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                messagebox.showerror('Error','Unsupported file type')
                return

            for r in rows:
                if inv_type in ('medical','nonmedical'):
                    self.db.execute(
                        """INSERT OR IGNORE INTO products(name,is_medical,sku,unit,sale_price)
                        VALUES(?,?,?,?,?)""",
                        (r.get('name'), 1 if inv_type=='medical' else 0, r.get('sku'), r.get('unit'), float(r.get('price') or 0))
                    )
                elif inv_type == 'suppliers':
                    self.db.execute("INSERT OR IGNORE INTO suppliers(name,phone,email,address) VALUES(?,?,?,?)",
                        (r.get('name'), r.get('phone'), r.get('email'), r.get('address')))
                elif inv_type == 'manufacturers':
                    self.db.execute("INSERT OR IGNORE INTO manufacturers(name,contact,notes) VALUES(?,?,?)",
                        (r.get('name'), r.get('contact'), r.get('notes')))
                elif inv_type == 'categories':
                    self.db.execute("INSERT OR IGNORE INTO categories(name,notes) VALUES(?,?)",
                        (r.get('name'), r.get('notes')))
                elif inv_type == 'formulas':
                    self.db.execute("INSERT OR IGNORE INTO formulas(name,composition) VALUES(?,?)",
                        (r.get('name'), r.get('composition')))
                elif inv_type == 'batches':
                    pid = self.db.query("SELECT id FROM products WHERE name=?", (r.get('product'),))
                    sid = self.db.query("SELECT id FROM suppliers WHERE name=?", (r.get('supplier'),))
                    pid = pid[0]['id'] if pid else None
                    sid = sid[0]['id'] if sid else None
                    if pid:
                        self.db.execute(
                            """INSERT INTO batches(product_id,supplier_id,batch_no,quantity,expiry_date,cost_price,created_at)
                            VALUES(?,?,?,?,?,?,?)""",
                            (pid, sid, r.get('batch_no'), int(r.get('quantity') or 0), r.get('expiry'), float(r.get('cost_price') or 0), now_str())
                        )

            messagebox.showinfo('Import','Data imported successfully!')
            try:
                self._inv_refresh_all()
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror('Import Error', str(e))
    
    def _build_inventory_tab(self):
        for w in self.tab_inventory.winfo_children(): w.destroy()
        frame = self.tab_inventory
        ttk.Label(frame, text='Inventory', font=('Segoe UI',14,'bold')).pack(anchor='w', padx=10, pady=(6,0))
        inv_nb = ttk.Notebook(frame); inv_nb.pack(fill='both', expand=True, padx=8, pady=8)

        med_tab = ttk.Frame(inv_nb); inv_nb.add(med_tab, text='Medical Products')
        # search frame for Medical Products
        med_tab_search_fr = ttk.Frame(med_tab)
        med_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(med_tab_search_fr, text='🔍').pack(side='left')
        self.med_tab_search_var = tk.StringVar()
        med_tab_search_entry = PlaceholderEntry(med_tab_search_fr, placeholder='Search...', textvariable=self.med_tab_search_var, width=36)
        med_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.med_tab_search_var.trace_add('write', lambda *a, t='Medical Products': self._filter_tab(t, self.med_tab_search_var.get()))

        nonmed_tab = ttk.Frame(inv_nb); inv_nb.add(nonmed_tab, text='Non-Medical Products')
        # search frame for Non-Medical Products
        nonmed_tab_search_fr = ttk.Frame(nonmed_tab)
        nonmed_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(nonmed_tab_search_fr, text='🔍').pack(side='left')
        self.nonmed_tab_search_var = tk.StringVar()
        nonmed_tab_search_entry = PlaceholderEntry(nonmed_tab_search_fr, placeholder='Search...', textvariable=self.nonmed_tab_search_var, width=36)
        nonmed_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.nonmed_tab_search_var.trace_add('write', lambda *a, t='Non-Medical Products': self._filter_tab(t, self.nonmed_tab_search_var.get()))

        suppliers_tab = ttk.Frame(inv_nb); inv_nb.add(suppliers_tab, text='Suppliers')
        # search frame for Suppliers
        suppliers_tab_search_fr = ttk.Frame(suppliers_tab)
        suppliers_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(suppliers_tab_search_fr, text='🔍').pack(side='left')
        self.suppliers_tab_search_var = tk.StringVar()
        suppliers_tab_search_entry = PlaceholderEntry(suppliers_tab_search_fr, placeholder='Search...', textvariable=self.suppliers_tab_search_var, width=36)
        suppliers_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.suppliers_tab_search_var.trace_add('write', lambda *a, t='Suppliers': self._filter_tab(t, self.suppliers_tab_search_var.get()))

        manufacturers_tab = ttk.Frame(inv_nb); inv_nb.add(manufacturers_tab, text='Manufacturers')
        # search frame for Manufacturers
        manufacturers_tab_search_fr = ttk.Frame(manufacturers_tab)
        manufacturers_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(manufacturers_tab_search_fr, text='🔍').pack(side='left')
        self.manufacturers_tab_search_var = tk.StringVar()
        manufacturers_tab_search_entry = PlaceholderEntry(manufacturers_tab_search_fr, placeholder='Search...', textvariable=self.manufacturers_tab_search_var, width=36)
        manufacturers_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.manufacturers_tab_search_var.trace_add('write', lambda *a, t='Manufacturers': self._filter_tab(t, self.manufacturers_tab_search_var.get()))

        categories_tab = ttk.Frame(inv_nb); inv_nb.add(categories_tab, text='Categories')
        # search frame for Categories
        categories_tab_search_fr = ttk.Frame(categories_tab)
        categories_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(categories_tab_search_fr, text='🔍').pack(side='left')
        self.categories_tab_search_var = tk.StringVar()
        categories_tab_search_entry = PlaceholderEntry(categories_tab_search_fr, placeholder='Search...', textvariable=self.categories_tab_search_var, width=36)
        categories_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.categories_tab_search_var.trace_add('write', lambda *a, t='Categories': self._filter_tab(t, self.categories_tab_search_var.get()))

        formulas_tab = ttk.Frame(inv_nb); inv_nb.add(formulas_tab, text='Formulas')
        # search frame for Formulas
        formulas_tab_search_fr = ttk.Frame(formulas_tab)
        formulas_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(formulas_tab_search_fr, text='🔍').pack(side='left')
        self.formulas_tab_search_var = tk.StringVar()
        formulas_tab_search_entry = PlaceholderEntry(formulas_tab_search_fr, placeholder='Search...', textvariable=self.formulas_tab_search_var, width=36)
        formulas_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.formulas_tab_search_var.trace_add('write', lambda *a, t='Formulas': self._filter_tab(t, self.formulas_tab_search_var.get()))

        batches_tab = ttk.Frame(inv_nb); inv_nb.add(batches_tab, text='Batches')
        # search frame for Batches
        batches_tab_search_fr = ttk.Frame(batches_tab)
        batches_tab_search_fr.pack(fill='x', padx=8, pady=(6,0))
        ttk.Label(batches_tab_search_fr, text='🔍').pack(side='left')
        self.batches_tab_search_var = tk.StringVar()
        batches_tab_search_entry = PlaceholderEntry(batches_tab_search_fr, placeholder='Search...', textvariable=self.batches_tab_search_var, width=36)
        batches_tab_search_entry.pack(side='left', padx=6)
        # wire trace
        self.batches_tab_search_var.trace_add('write', lambda *a, t='Batches': self._filter_tab(t, self.batches_tab_search_var.get()))


        cols = ('id','name','sku','unit','category','manufacturer','price','stock')

        def make_prod_tree(parent):
            tree = ttk.Treeview(parent, columns=cols, show='headings', height=18)
            # headings
            for c in cols:
                tree.heading(c, text=c.capitalize())
            # column alignments & widths
            tree.column('id', width=60, anchor='center')
            tree.column('name', width=180, anchor='w')
            tree.column('sku', width=100, anchor='center')
            tree.column('unit', width=100, anchor='center')
            tree.column('category', width=140, anchor='w')
            tree.column('manufacturer', width=140, anchor='w')
            tree.column('price', width=100, anchor='e')
            tree.column('stock', width=80, anchor='center')
            tree.pack(fill='both', expand=True, padx=8, pady=8)
            return tree

        self._med_tree = make_prod_tree(med_tab)
        btns_med = ttk.Frame(med_tab); btns_med.pack(fill='x', padx=8)
        ttk.Button(btns_med, text='Add', command=lambda: self._inv_add_product(is_medical=1)).pack(side='left', padx=6)
        ttk.Button(btns_med, text='Edit', command=lambda: self._inv_edit_product(self._med_tree)).pack(side='left', padx=6)
        ttk.Button(btns_med, text='Delete', command=lambda: self._inv_delete_product(self._med_tree)).pack(side='left', padx=6)
        ttk.Button(btns_med, text='Add by CSV/Excel', command=lambda: self._import_inventory('medical')).pack(side='left', padx=6)
        ttk.Button(btns_med, text='Clear Filter', command=self._inv_refresh_all).pack(side='right', padx=6)

        self._nonmed_tree = make_prod_tree(nonmed_tab)
        btns_non = ttk.Frame(nonmed_tab); btns_non.pack(fill='x', padx=8)
        ttk.Button(btns_non, text='Add', command=lambda: self._inv_add_product(is_medical=0)).pack(side='left', padx=6)
        ttk.Button(btns_non, text='Edit', command=lambda: self._inv_edit_product(self._nonmed_tree)).pack(side='left', padx=6)
        ttk.Button(btns_non, text='Delete', command=lambda: self._inv_delete_product(self._nonmed_tree)).pack(side='left', padx=6)
        ttk.Button(btns_non, text='Add by CSV/Excel', command=lambda: self._import_inventory('nonmedical')).pack(side='left', padx=6)
        ttk.Button(btns_non, text='Clear Filter', command=self._inv_refresh_all).pack(side='right', padx=6)

        self._sup_tree = ttk.Treeview(suppliers_tab, columns=('id','name','phone','email','address'), show='headings')
        for c in ('id','name','phone','email','address'):
            self._sup_tree.heading(c, text=c.capitalize())
        self._sup_tree.column('id', width=60, anchor='center')
        self._sup_tree.column('name', width=180, anchor='w')
        self._sup_tree.column('phone', width=120, anchor='center')
        self._sup_tree.column('email', width=180, anchor='w')
        self._sup_tree.column('address', width=240, anchor='w')
        self._sup_tree.pack(fill='both', expand=True, padx=8, pady=8)
        sup_btns = ttk.Frame(suppliers_tab); sup_btns.pack(fill='x', padx=8)
        ttk.Button(sup_btns, text='Add', command=self._add_supplier).pack(side='left', padx=6)
        ttk.Button(sup_btns, text='Edit', command=self._edit_supplier).pack(side='left', padx=6)
        ttk.Button(sup_btns, text='Delete', command=self._delete_supplier).pack(side='left', padx=6)
        ttk.Button(sup_btns, text='Add by CSV/Excel', command=lambda: self._import_inventory('suppliers')).pack(side='left', padx=6)

        self._man_tree = ttk.Treeview(manufacturers_tab, columns=('id','name','contact','notes'), show='headings')
        for c in ('id','name','contact','notes'):
            self._man_tree.heading(c, text=c.capitalize())
        self._man_tree.column('id', width=60, anchor='center')
        self._man_tree.column('name', width=180, anchor='w')
        self._man_tree.column('contact', width=160, anchor='w')
        self._man_tree.column('notes', width=240, anchor='w')
        self._man_tree.pack(fill='both', expand=True, padx=8, pady=8)
        man_btns = ttk.Frame(manufacturers_tab); man_btns.pack(fill='x', padx=8)
        ttk.Button(man_btns, text='Add', command=self._add_manufacturer).pack(side='left', padx=6)
        ttk.Button(man_btns, text='Edit', command=self._edit_manufacturer).pack(side='left', padx=6)
        ttk.Button(man_btns, text='Delete', command=self._delete_manufacturer).pack(side='left', padx=6)
        ttk.Button(man_btns, text='Add by CSV/Excel', command=lambda: self._import_inventory('manufacturers')).pack(side='left', padx=6)

        self._cat_tree = ttk.Treeview(categories_tab, columns=('id','name','notes'), show='headings')
        for c in ('id','name','notes'):
            self._cat_tree.heading(c, text=c.capitalize())
        self._cat_tree.column('id', width=60, anchor='center')
        self._cat_tree.column('name', width=200, anchor='w')
        self._cat_tree.column('notes', width=300, anchor='w')
        self._cat_tree.pack(fill='both', expand=True, padx=8, pady=8)
        cat_btns = ttk.Frame(categories_tab); cat_btns.pack(fill='x', padx=8)
        ttk.Button(cat_btns, text='Add', command=self._add_category).pack(side='left', padx=6)
        ttk.Button(cat_btns, text='Edit', command=self._edit_category).pack(side='left', padx=6)
        ttk.Button(cat_btns, text='Delete', command=self._delete_category).pack(side='left', padx=6)
        ttk.Button(cat_btns, text='Add by CSV/Excel', command=lambda: self._import_inventory('categories')).pack(side='left', padx=6)

        self._form_tree = ttk.Treeview(formulas_tab, columns=('id','name','composition'), show='headings')
        for c in ('id','name','composition'):
            self._form_tree.heading(c, text=c.capitalize())
        self._form_tree.column('id', width=60, anchor='center')
        self._form_tree.column('name', width=200, anchor='w')
        self._form_tree.column('composition', width=320, anchor='w')
        self._form_tree.pack(fill='both', expand=True, padx=8, pady=8)
        form_btns = ttk.Frame(formulas_tab); form_btns.pack(fill='x', padx=8)
        ttk.Button(form_btns, text='Add', command=self._add_formula).pack(side='left', padx=6)
        ttk.Button(form_btns, text='Edit', command=self._edit_formula).pack(side='left', padx=6)
        ttk.Button(form_btns, text='Delete', command=self._delete_formula).pack(side='left', padx=6)
        ttk.Button(form_btns, text='Add by CSV/Excel', command=lambda: self._import_inventory('formulas')).pack(side='left', padx=6)

        self._batch_tree = ttk.Treeview(batches_tab, columns=('id','product','batch_no','quantity','expiry','supplier'), show='headings')
        for c in ('id','product','batch_no','quantity','expiry','supplier'): self._batch_tree.heading(c, text=c.capitalize()); self._batch_tree.column(c, width=140, anchor='w')
        self._batch_tree.pack(fill='both', expand=True, padx=8, pady=8)
        batch_btns = ttk.Frame(batches_tab); batch_btns.pack(fill='x', padx=8)
        ttk.Button(batch_btns, text='Add', command=self._add_batch).pack(side='left', padx=6)
        ttk.Button(batch_btns, text='Edit', command=self._edit_batch).pack(side='left', padx=6)
        ttk.Button(batch_btns, text='Delete', command=self._delete_batch).pack(side='left', padx=6)
        ttk.Button(batch_btns, text='Clear Filter', command=self._inv_refresh_all).pack(side='right', padx=6)

        self._inv_refresh_all()

    
    def _filter_tab(self, tab_name, term):
        try:
            mapping = {
                'Medical Products': ('_med_tree', self._get_medical_data),
                'Non-Medical Products': ('_nonmed_tree', self._get_nonmedical_data),
                'Suppliers': ('_sup_tree', self._get_suppliers_data),
                'Manufacturers': ('_man_tree', self._get_manufacturers_data),
                'Categories': ('_cat_tree', self._get_categories_data),
                'Formulas': ('_form_tree', self._get_formulas_data),
                'Batches': ('_batch_tree', self._get_batches_data)
            }
            
            term = (term or '').strip().lower()
            tree_attr, data_getter = mapping.get(tab_name, (None, None))
            
            if not tree_attr or not data_getter:
                return
                
            tree = getattr(self, tree_attr, None)
            if not tree:
                return
            
            # Clear the tree
            tree.delete(*tree.get_children())
            
            # If empty term, load all data
            if not term:
                data = data_getter()
                for item_data in data:
                    tree.insert('', 'end', values=item_data)
                return
            
            # Filter: show only matching items
            data = data_getter()
            for item_data in data:
                if any(term in str(v).lower() for v in item_data):
                    tree.insert('', 'end', values=item_data)
                    
        except Exception as e:
            print(f"Filter error in {tab_name}: {e}")

    def _get_medical_data(self):
        rows = self.db.query('''SELECT p.id,p.name,p.sku,p.unit,c.name as category,m.name as manufacturer,p.sale_price as price,
            COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p
            LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id WHERE p.is_medical=1 ORDER BY p.name;''')
        return [(r['id'], r['name'], r['sku'] or '', r.get('unit','') or '', r.get('category') or '', r.get('manufacturer') or '', f"{r['price']:.2f}", r['stock']) for r in rows]

    def _get_nonmedical_data(self):
        rows = self.db.query('''SELECT p.id,p.name,p.sku,p.unit,c.name as category,m.name as manufacturer,p.sale_price as price,
            COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p
            LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id WHERE p.is_medical=0 ORDER BY p.name;''')
        return [(r['id'], r['name'], r['sku'] or '', r.get('unit','') or '', r.get('category') or '', r.get('manufacturer') or '', f"{r['price']:.2f}", r['stock']) for r in rows]

    def _get_suppliers_data(self):
        rows = self.db.query('SELECT id,name,phone,email,address FROM suppliers ORDER BY name;')
        return [(r['id'], r['name'], r['phone'] or '', r['email'] or '', r['address'] or '') for r in rows]

    def _get_manufacturers_data(self):
        rows = self.db.query('SELECT id,name,contact,notes FROM manufacturers ORDER BY name;')
        return [(r['id'], r['name'], r['contact'] or '', r['notes'] or '') for r in rows]

    def _get_categories_data(self):
        rows = self.db.query('SELECT id,name,notes FROM categories ORDER BY name;')
        return [(r['id'], r['name'], r['notes'] or '') for r in rows]

    def _get_formulas_data(self):
        rows = self.db.query('SELECT id,name,composition FROM formulas ORDER BY name;')
        return [(r['id'], r['name'], r['composition'] or '') for r in rows]

    def _get_batches_data(self):
        rows = self.db.query('SELECT b.id, p.name as product, b.batch_no, b.quantity, b.expiry_date, s.name as supplier FROM batches b LEFT JOIN products p ON p.id=b.product_id LEFT JOIN suppliers s ON s.id=b.supplier_id ORDER BY b.id DESC;')
        return [(r['id'], r['product'], r['batch_no'] or '', r['quantity'], r['expiry_date'] or '', r['supplier'] or '') for r in rows]
    
    def _inv_refresh_all(self):
        # Medical Products
        self._med_tree.delete(*self._med_tree.get_children())
        for item_data in self._get_medical_data():
            self._med_tree.insert('', 'end', iid=item_data[0], values=item_data)

        # Non-Medical Products
        self._nonmed_tree.delete(*self._nonmed_tree.get_children())
        for item_data in self._get_nonmedical_data():
            self._nonmed_tree.insert('', 'end', iid=item_data[0], values=item_data)

        # Suppliers
        self._sup_tree.delete(*self._sup_tree.get_children())
        for item_data in self._get_suppliers_data():
            self._sup_tree.insert('', 'end', values=item_data)

        # Manufacturers
        self._man_tree.delete(*self._man_tree.get_children())
        for item_data in self._get_manufacturers_data():
            self._man_tree.insert('', 'end', values=item_data)

        # Categories
        self._cat_tree.delete(*self._cat_tree.get_children())
        for item_data in self._get_categories_data():
            self._cat_tree.insert('', 'end', values=item_data)

        # Formulas
        self._form_tree.delete(*self._form_tree.get_children())
        for item_data in self._get_formulas_data():
            self._form_tree.insert('', 'end', values=item_data)

        # Batches
        self._batch_tree.delete(*self._batch_tree.get_children())
        for item_data in self._get_batches_data():
            self._batch_tree.insert('', 'end', values=item_data)
            
    def _inv_add_product(self, is_medical=1):
        cats = [r['name'] for r in self.db.query('SELECT name FROM categories ORDER BY name;')]
        mans = [r['name'] for r in self.db.query('SELECT name FROM manufacturers ORDER BY name;')]
        forms = [r['name'] for r in self.db.query('SELECT name FROM formulas ORDER BY name;')]
        units = ['mg','ml','g','IU','tablet','capsule','bottle','strip','box']

        def save(d):
            if not d.get('name'):
                return messagebox.showerror('Error','Name required')
            cid = mid = fid = None
            if d.get('category'):
                row = self.db.query('SELECT id FROM categories WHERE name=?;',(d['category'],))
                if row: cid = row[0]['id']
                else: cid = self.db.execute('INSERT INTO categories(name) VALUES(?);',(d['category'],))
            if d.get('manufacturer'):
                row = self.db.query('SELECT id FROM manufacturers WHERE name=?;',(d['manufacturer'],))
                if row: mid = row[0]['id']
                else: mid = self.db.execute('INSERT INTO manufacturers(name) VALUES(?);',(d['manufacturer'],))
            if d.get('formula'):
                row = self.db.query('SELECT id FROM formulas WHERE name=?;',(d['formula'],))
                if row: fid = row[0]['id']
                else: fid = self.db.execute('INSERT INTO formulas(name) VALUES(?);',(d['formula'],))
            try:
                self.db.execute('INSERT INTO products(name,sku,is_medical,category_id,manufacturer_id,formula_id,unit,sale_price,notes) VALUES(?,?,?,?,?,?,?,?,?);',
                                (d.get('name'), d.get('sku') or None, 1 if is_medical else 0, cid, mid, fid, d.get('unit') or '', float(d.get('price') or 0), d.get('notes') or ''))
                messagebox.showinfo('Saved','Product added'); self._inv_refresh_all()
            except Exception as e:
                messagebox.showerror('Error', str(e))

        fields = [
            {'key':'name','label':'Name'},
            {'key':'sku','label':'SKU'},
            {'key':'unit','label':'Unit','widget':'combobox','values':units,'state':'normal'},
            {'key':'category','label':'Category','widget':'combobox','values':cats,'state':'normal'},
            {'key':'manufacturer','label':'Manufacturer','widget':'combobox','values':mans,'state':'normal'},
            {'key':'formula','label':'Formula','widget':'combobox','values':forms,'state':'normal'},
            {'key':'price','label':'Sale Price'},
            {'key':'notes','label':'Notes','widget':'text'}
        ]

        dlg = FormDialog(self.root, 'Add Product', fields, on_submit=save)

        # Make comboboxes searchable
        self.make_searchable(dlg.widgets['unit'][0], units)
        self.make_searchable(dlg.widgets['category'][0], cats)
        self.make_searchable(dlg.widgets['manufacturer'][0], mans)
        self.make_searchable(dlg.widgets['formula'][0], forms)


   
    def _inv_edit_product(self, tree):
        sel = tree.selection()
        if not sel: return messagebox.showwarning('Select','Select a product to edit')
        values = tree.item(sel[0], 'values')
        pid = int(values[0])
        row = self.db.query('SELECT p.*, c.name as category_name, m.name as manufacturer_name, f.name as formula_name FROM products p LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id LEFT JOIN formulas f ON p.formula_id=f.id WHERE p.id=?;',(pid,))
        if not row: return messagebox.showerror('Error','Not found')
        row = row[0]
        initial = {'name':row['name'],'sku':row['sku'] or '','unit':row['unit'] or '','category':row.get('category_name') or '','manufacturer':row.get('manufacturer_name') or '','formula':row.get('formula_name') or '','price':row['sale_price'],'notes':row['notes']}
        
        cats = [r['name'] for r in self.db.query('SELECT name FROM categories ORDER BY name;')]
        mans = [r['name'] for r in self.db.query('SELECT name FROM manufacturers ORDER BY name;')]
        forms = [r['name'] for r in self.db.query('SELECT name FROM formulas ORDER BY name;')]
        units = ['mg','ml','g','IU','tablet','capsule','bottle','strip','box']

        def save(d):
            cid = mid = fid = None
            if d.get('category'):
                r = self.db.query('SELECT id FROM categories WHERE name=?;',(d['category'],))
                if r: cid = r[0]['id']
                else: cid = self.db.execute('INSERT INTO categories(name) VALUES(?);',(d['category'],))
            if d.get('manufacturer'):
                r = self.db.query('SELECT id FROM manufacturers WHERE name=?;',(d['manufacturer'],))
                if r: mid = r[0]['id']
                else: mid = self.db.execute('INSERT INTO manufacturers(name) VALUES(?);',(d['manufacturer'],))
            if d.get('formula'):
                r = self.db.query('SELECT id FROM formulas WHERE name=?;',(d['formula'],))
                if r: fid = r[0]['id']
                else: fid = self.db.execute('INSERT INTO formulas(name) VALUES(?);',(d['formula'],))
            try:
                self.db.execute('UPDATE products SET name=?,sku=?,category_id=?,manufacturer_id=?,formula_id=?,unit=?,sale_price=?,notes=? WHERE id=?;',
                                (d.get('name'), d.get('sku') or None, cid, mid, fid, d.get('unit') or '', float(d.get('price') or 0), d.get('notes') or '', pid))
                messagebox.showinfo('Saved','Product updated'); self._inv_refresh_all()
            except Exception as e:
                messagebox.showerror('Error', str(e))

        fields = [
            {'key':'name','label':'Name'},
            {'key':'sku','label':'SKU'},
            {'key':'unit','label':'Unit','widget':'combobox','values':units,'state':'normal'},
            {'key':'category','label':'Category','widget':'combobox','values':cats,'state':'normal'},
            {'key':'manufacturer','label':'Manufacturer','widget':'combobox','values':mans,'state':'normal'},
            {'key':'formula','label':'Formula','widget':'combobox','values':forms,'state':'normal'},
            {'key':'price','label':'Sale Price'},
            {'key':'notes','label':'Notes','widget':'text'}
        ]

        dlg = FormDialog(self.root, 'Edit Product', fields, initial=initial, on_submit=save)

        # Make comboboxes searchable
        self.make_searchable(dlg.widgets['unit'][0], units)
        self.make_searchable(dlg.widgets['category'][0], cats)
        self.make_searchable(dlg.widgets['manufacturer'][0], mans)
        self.make_searchable(dlg.widgets['formula'][0], forms)


    def _inv_delete_product(self, tree):
        sel = tree.selection()
        if not sel: return messagebox.showwarning('Select','Select a product to delete')
        values = tree.item(sel[0], 'values')
        pid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete selected product?'): return
        try:
            self.db.execute('DELETE FROM products WHERE id=?;',(pid,))
            messagebox.showinfo('Deleted','Product deleted'); self._inv_refresh_all()
        except Exception as e:
            messagebox.showerror('Error', str(e))

    # Suppliers CRUD
    def _add_supplier(self):
        def save(d):
            if not d.get('name'): return messagebox.showerror('Error','Name required')
            try:
                self.db.execute('INSERT INTO suppliers(name,phone,email,address) VALUES(?,?,?,?);',(d.get('name'), d.get('phone'), d.get('email'), d.get('address')))
                messagebox.showinfo('Saved','Supplier added'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Add Supplier', [{'key':'name','label':'Name'},{'key':'phone','label':'Phone'},{'key':'email','label':'Email'},{'key':'address','label':'Address'}], on_submit=save)

    def _edit_supplier(self):
        sel = self._sup_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select supplier to edit')
        values = self._sup_tree.item(sel[0], 'values')
        sid = int(values[0]); r = self.db.query('SELECT * FROM suppliers WHERE id=?;',(sid,))[0]
        initial = {'name':r['name'],'phone':r['phone'],'email':r['email'],'address':r['address']}
        def save(d):
            try:
                self.db.execute('UPDATE suppliers SET name=?,phone=?,email=?,address=? WHERE id=?;',(d.get('name'),d.get('phone'),d.get('email'),d.get('address'),sid))
                messagebox.showinfo('Saved','Supplier updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Edit Supplier', [{'key':'name','label':'Name'},{'key':'phone','label':'Phone'},{'key':'email','label':'Email'},{'key':'address','label':'Address'}], initial=initial, on_submit=save)

    def _delete_supplier(self):
        sel = self._sup_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select supplier to delete')
        values = self._sup_tree.item(sel[0], 'values')
        sid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete supplier?'): return
        try:
            self.db.execute('DELETE FROM suppliers WHERE id=?;',(sid,)); messagebox.showinfo('Deleted','Supplier deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))

    # Manufacturers CRUD
    def _add_manufacturer(self):
        def save(d):
            if not d.get('name'): return messagebox.showerror('Error','Name required')
            try:
                self.db.execute('INSERT INTO manufacturers(name,contact,notes) VALUES(?,?,?);',(d.get('name'),d.get('contact'),d.get('notes')))
                messagebox.showinfo('Saved','Manufacturer added'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Add Manufacturer', [{'key':'name','label':'Name'},{'key':'contact','label':'Contact'},{'key':'notes','label':'Notes','widget':'text'}], on_submit=save)

    def _edit_manufacturer(self):
        sel = self._man_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select manufacturer to edit')
        values = self._man_tree.item(sel[0], 'values')
        mid = int(values[0]); r = self.db.query('SELECT * FROM manufacturers WHERE id=?;',(mid,))[0]
        initial = {'name':r['name'],'contact':r['contact'],'notes':r['notes']}
        def save(d):
            try:
                self.db.execute('UPDATE manufacturers SET name=?,contact=?,notes=? WHERE id=?;',(d.get('name'),d.get('contact'),d.get('notes'),mid))
                messagebox.showinfo('Saved','Manufacturer updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Edit Manufacturer', [{'key':'name','label':'Name'},{'key':'contact','label':'Contact'},{'key':'notes','label':'Notes','widget':'text'}], initial=initial, on_submit=save)

    def _delete_manufacturer(self):
        sel = self._man_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select manufacturer to delete')
        values = self._man_tree.item(sel[0], 'values')
        mid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete manufacturer?'): return
        try:
            self.db.execute('DELETE FROM manufacturers WHERE id=?;',(mid,)); messagebox.showinfo('Deleted','Manufacturer deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))

    # Categories CRUD
    def _add_category(self):
        def save(d):
            if not d.get('name'): return messagebox.showerror('Error','Name required')
            try:
                self.db.execute('INSERT INTO categories(name,notes) VALUES(?,?);',(d.get('name'), d.get('notes'))); messagebox.showinfo('Saved','Category added'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Add Category', [{'key':'name','label':'Name'},{'key':'notes','label':'Notes','widget':'text'}], on_submit=save)

    def _edit_category(self):
        sel = self._cat_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select category to edit')
        values = self._cat_tree.item(sel[0], 'values')
        cid = int(values[0]); r = self.db.query('SELECT * FROM categories WHERE id=?;',(cid,))[0]
        initial = {'name':r['name'],'notes':r['notes']}
        def save(d):
            try:
                self.db.execute('UPDATE categories SET name=?,notes=? WHERE id=?;',(d.get('name'),d.get('notes'),cid)); messagebox.showinfo('Saved','Category updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Edit Category', [{'key':'name','label':'Name'},{'key':'notes','label':'Notes','widget':'text'}], initial=initial, on_submit=save)

    def _delete_category(self):
        sel = self._cat_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select category to delete')
        values = self._cat_tree.item(sel[0], 'values')
        cid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete category?'): return
        try:
            self.db.execute('DELETE FROM categories WHERE id=?;',(cid,)); messagebox.showinfo('Deleted','Category deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))

    # Formulas CRUD
    def _add_formula(self):
        def save(d):
            if not d.get('name'): return messagebox.showerror('Error','Name required')
            try:
                self.db.execute('INSERT INTO formulas(name,composition) VALUES(?,?);',(d.get('name'),d.get('composition'))); messagebox.showinfo('Saved','Formula added'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Add Formula', [{'key':'name','label':'Name'},{'key':'composition','label':'Composition','widget':'text'}], on_submit=save)

    def _edit_formula(self):
        sel = self._form_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select formula to edit')
        values = self._form_tree.item(sel[0], 'values')
        fid = int(values[0]); r = self.db.query('SELECT * FROM formulas WHERE id=?;',(fid,))[0]
        initial = {'name':r['name'],'composition':r['composition']}
        def save(d):
            try:
                self.db.execute('UPDATE formulas SET name=?,composition=? WHERE id=?;',(d.get('name'),d.get('composition'),fid)); messagebox.showinfo('Saved','Formula updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        FormDialog(self.root, 'Edit Formula', [{'key':'name','label':'Name'},{'key':'composition','label':'Composition','widget':'text'}], initial=initial, on_submit=save)

    def _delete_formula(self):
        sel = self._form_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select formula to delete')
        values = self._form_tree.item(sel[0], 'values')
        fid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete formula?'): return
        try:
            self.db.execute('DELETE FROM formulas WHERE id=?;',(fid,)); messagebox.showinfo('Deleted','Formula deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))


    def make_searchable(self, cb_widget, original_values):
        def on_keyrelease(event):
            typed = cb_widget.get()
            if typed == '':
                cb_widget['values'] = original_values  # show all if empty
            else:
                # Only items starting with typed string (case-insensitive)
                filtered = [v for v in original_values if v.lower().startswith(typed.lower())]
                cb_widget['values'] = filtered
            if cb_widget['values']:
                cb_widget.event_generate('<Down>')  # open dropdown automatically

        cb_widget.bind('<KeyRelease>', on_keyrelease)


    def _add_batch(self):
        # Fetch data from database
        products = [r['name'] for r in self.db.query('SELECT name FROM products ORDER BY name;')]
        suppliers = [r['name'] for r in self.db.query('SELECT name FROM suppliers ORDER BY name;')]
        manufacturers = [r['name'] for r in self.db.query('SELECT name FROM manufacturers ORDER BY name;')]

        # Function to save batch
        def save(d):
            pid = sid = mid = None

            p = self.db.query('SELECT id FROM products WHERE name=?;', (d.get('product'),))
            if p: pid = p[0]['id']

            if d.get('supplier'):
                s = self.db.query('SELECT id FROM suppliers WHERE name=?;', (d.get('supplier'),))
                if s: sid = s[0]['id']

            if d.get('manufacturer'):
                m = self.db.query('SELECT id FROM manufacturers WHERE name=?;', (d.get('manufacturer'),))
                if m: mid = m[0]['id']

            if not pid:
                return messagebox.showerror('Error', 'Product is required and must exist.')

            try:
                self.db.execute(
                    'INSERT INTO batches(product_id, supplier_id, manufacturer_id, batch_no, quantity, expiry_date, cost_price, created_at) '
                    'VALUES(?,?,?,?,?,?,?,?);',
                    (pid, sid, mid, d.get('batch_no') or '', int(d.get('quantity') or 0),
                    d.get('expiry') or None, float(d.get('cost_price') or 0), now_str())
                )
                messagebox.showinfo('Saved', 'Batch added successfully.')
                self._inv_refresh_all()
            except Exception as e:
                messagebox.showerror('Error', str(e))

        # Form fields — state='normal' is required for typing
        fields = [
            {'key': 'product', 'label': 'Product', 'widget': 'combobox', 'values': products, 'state': 'normal'},
            {'key': 'supplier', 'label': 'Supplier', 'widget': 'combobox', 'values': suppliers, 'state': 'normal'},
            {'key': 'manufacturer', 'label': 'Manufacturer', 'widget': 'combobox', 'values': manufacturers, 'state': 'normal'},
            {'key': 'batch_no', 'label': 'Batch No'},
            {'key': 'quantity', 'label': 'Quantity'},
            {'key': 'expiry', 'label': 'Expiry (YYYY-MM-DD)'},
            {'key': 'cost_price', 'label': 'Cost Price'}
        ]

        # Open form dialog
        dlg = FormDialog(self.root, 'Add Batch', fields, on_submit=save)

        # Make Product, Supplier, Manufacturer comboboxes smooth-searchable
        self.make_searchable(dlg.widgets['product'][0], products)
        self.make_searchable(dlg.widgets['supplier'][0], suppliers)
        self.make_searchable(dlg.widgets['manufacturer'][0], manufacturers)

    # Find the _edit_batch method and update it to include manufacturer:
    def _edit_batch(self):
        sel = self._batch_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select batch to edit')
        values = self._batch_tree.item(sel[0], 'values')
        bid = int(values[0]); r = self.db.query('SELECT * FROM batches WHERE id=?;',(bid,))[0]
        prodname = self.db.query('SELECT name FROM products WHERE id=?;',(r['product_id'],))[0]['name']
        supname = None
        if r['supplier_id']:
            s = self.db.query('SELECT name FROM suppliers WHERE id=?;',(r['supplier_id'],))
            supname = s[0]['name'] if s else None
        
        # Get manufacturer name
        manname = None
        if r['product_id']:
            m = self.db.query('SELECT m.name FROM products p LEFT JOIN manufacturers m ON m.id=p.manufacturer_id WHERE p.id=?;',(r['product_id'],))
            manname = m[0]['name'] if m and m[0]['name'] else None
        
        initial = {'product':prodname,'supplier':supname,'manufacturer':manname,'batch_no':r['batch_no'],'quantity':r['quantity'],'expiry':r['expiry_date'],'cost_price':r['cost_price']}
        
        def save(d):
            p = self.db.query('SELECT id FROM products WHERE name=?;',(d.get('product'),))
            if not p: return messagebox.showerror('Error','Product required and must exist')
            pid = p[0]['id']
            sid = None
            if d.get('supplier'):
                s = self.db.query('SELECT id FROM suppliers WHERE name=?;',(d.get('supplier'),))
                if s: sid = s[0]['id']
            
            # Get manufacturer ID
            mid = None
            if d.get('manufacturer'):
                m = self.db.query('SELECT id FROM manufacturers WHERE name=?;',(d.get('manufacturer'),))
                if m: mid = m[0]['id']
            
            try:
                self.db.execute('UPDATE batches SET product_id=?,supplier_id=?,batch_no=?,quantity=?,expiry_date=?,cost_price=? WHERE id=?;',
                                (pid, sid, d.get('batch_no') or '', int(d.get('quantity') or 0), d.get('expiry') or None, float(d.get('cost_price') or 0), bid))
                messagebox.showinfo('Saved','Batch updated'); self._inv_refresh_all()
            except Exception as e: messagebox.showerror('Error',str(e))
        
        fields = [
            {'key':'product','label':'Product','widget':'entry'},
            {'key':'supplier','label':'Supplier','widget':'entry'},
            {'key':'manufacturer','label':'Manufacturer','widget':'entry'},
            {'key':'batch_no','label':'Batch No'},
            {'key':'quantity','label':'Quantity'},
            {'key':'expiry','label':'Expiry (YYYY-MM-DD)'},
            {'key':'cost_price','label':'Cost Price'}
        ]
        
        dlg = FormDialog(self.root, 'Edit Batch', fields, initial=initial, on_submit=save)
        
        # Add autocomplete functionality after creating the widgets
        if 'product' in dlg.widgets:
            product_entry = dlg.widgets['product'][0]
            product_entry.suggestions_getter = lambda term: [r['name'] for r in self.db.query(
                "SELECT name FROM products WHERE name LIKE ? ORDER BY name LIMIT 10", (f"%{term}%",))]
        
        if 'supplier' in dlg.widgets:
            supplier_entry = dlg.widgets['supplier'][0]
            supplier_entry.suggestions_getter = lambda term: [r['name'] for r in self.db.query(
                "SELECT name FROM suppliers WHERE name LIKE ? ORDER BY name LIMIT 10", (f"%{term}%",))]
        
        if 'manufacturer' in dlg.widgets:
            manufacturer_entry = dlg.widgets['manufacturer'][0]
            manufacturer_entry.suggestions_getter = lambda term: [r['name'] for r in self.db.query(
                "SELECT name FROM manufacturers WHERE name LIKE ? ORDER BY name LIMIT 10", (f"%{term}%",))]
        
        # Add functionality to auto-fill manufacturer when product is selected
        def on_product_select(event=None):
            product_name = dlg.widgets['product'][0].get().strip()
            if product_name:
                # Get manufacturer for this product
                result = self.db.query('''
                    SELECT m.name 
                    FROM products p 
                    LEFT JOIN manufacturers m ON p.manufacturer_id = m.id 
                    WHERE p.name = ? LIMIT 1
                ''', (product_name,))
                
                if result and result[0]['name']:
                    dlg.widgets['manufacturer'][0].delete(0, 'end')
                    dlg.widgets['manufacturer'][0].insert(0, result[0]['name'])
        
        # Bind the product entry to auto-fill manufacturer
        dlg.widgets['product'][0].bind('<FocusOut>', on_product_select)
        dlg.widgets['product'][0].bind('<Return>', on_product_select)

    def _delete_batch(self):
        sel = self._batch_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select batch to delete')
        values = self._batch_tree.item(sel[0], 'values')
        bid = int(values[0])
        if not messagebox.askyesno('Confirm','Delete batch?'): return
        try:
            self.db.execute('DELETE FROM batches WHERE id=?;',(bid,)); messagebox.showinfo('Deleted','Batch deleted'); self._inv_refresh_all()
        except Exception as e: messagebox.showerror('Error',str(e))

     # ---------------- POS with nested tabs ----------------
    def _build_pos_tab(self):
        # Clear POS tab
        for w in self.tab_pos.winfo_children():
            w.destroy()
        pos_nb = ttk.Notebook(self.tab_pos)
        pos_nb.pack(fill='both', expand=True, padx=8, pady=8)

        # --- New Sale ---
        new_sale_tab = ttk.Frame(pos_nb)
        pos_nb.add(new_sale_tab, text='New Sale')
        NewSaleTab(new_sale_tab, self.db, self.user).pack(fill='both', expand=True)

        # --- Sale History ---
        history_tab = ttk.Frame(pos_nb)
        pos_nb.add(history_tab, text='Sale History')
        self._sale_history_tree = ttk.Treeview(
            history_tab,
            columns=('sale_id','date','customer','product','qty','price','expiry','supplier','manufacturer','subtotal'),
            show='headings', height=18
        )
        for c in ('sale_id','date','customer','product','qty','price','expiry','supplier','manufacturer','subtotal'):
            self._sale_history_tree.heading(c, text=c.capitalize())
            self._sale_history_tree.column(c, width=120, anchor='w')
        self._sale_history_tree.pack(fill='both', expand=True, padx=8, pady=8)
        btns = ttk.Frame(history_tab)
        btns.pack(fill='x')
        ttk.Button(btns, text='Refresh', command=self._sale_history_refresh).pack(side='left', padx=6)
        ttk.Button(btns, text='Print Receipt (Selected)', command=self._sale_history_print_selected).pack(side='left', padx=6)
        try:
            self._sale_history_refresh()
        except Exception:
            pass
        # --- Return Item ---
        return_item_tab = ttk.Frame(pos_nb)
        pos_nb.add(return_item_tab, text='Return Item')
        self._build_return_item_tab(return_item_tab)

        # --- Return History ---
        returns_tab = ttk.Frame(pos_nb)
        pos_nb.add(returns_tab, text='Return History')
        self._return_tree = ttk.Treeview(
            returns_tab,
            columns=('id','sale_item','product','qty','reason','created','expiry'),
            show='headings'
        )
        for c in ('id','sale_item','product','qty','reason','created','expiry'):
            self._return_tree.heading(c, text=c.capitalize())
            self._return_tree.column(c, width=120, anchor='w')
        self._return_tree.pack(fill='both', expand=True, padx=8, pady=8)
        ttk.Button(returns_tab, text='Refresh', command=self._return_refresh).pack(anchor='e', padx=8, pady=6)
        try:
            self._return_refresh()
        except Exception:
            pass

        # --- Sale Reports ---
        reports_tab = ttk.Frame(pos_nb)
        pos_nb.add(reports_tab, text='Sale Reports')
        try:
            # Use the fixed reports implementation
            self._build_reports_in_frame(reports_tab)
        except Exception as e:
            print(f"Error building reports: {e}")
            # Simple fallback
            ttk.Label(reports_tab, text="Reports functionality not available").pack(pady=20)


    def _build_return_item_tab(self, frame):
        # --- Search by Sale ID ---
        search_frame = ttk.Frame(frame)
        search_frame.pack(fill="x", padx=10, pady=6)

        ttk.Label(search_frame, text="Sale ID:").pack(side="left")
        self.return_sale_id = ttk.Entry(search_frame, width=15)
        self.return_sale_id.pack(side="left", padx=6)

        ttk.Button(search_frame, text="Fetch Sale",
            command=lambda: self._load_sale_for_return(self.return_sale_id.get())
        ).pack(side="left")

        # --- Treeview for sale items ---
        self.return_tree = ttk.Treeview(
            frame,
            columns=("id", "product", "qty", "price", "total"),
            show="headings"
        )
        for col in ("id", "product", "qty", "price", "total"):
            self.return_tree.heading(col, text=col.capitalize())
            self.return_tree.column(col, width=120, anchor="w")
        self.return_tree.pack(fill="both", expand=True, padx=10, pady=6)

        # --- Return controls ---
        control_frame = ttk.Frame(frame)
        control_frame.pack(fill="x", padx=10, pady=6)

        ttk.Label(control_frame, text="Quantity to Return:").pack(side="left")
        self.return_qty = ttk.Entry(control_frame, width=10)
        self.return_qty.pack(side="left", padx=6)

        ttk.Label(control_frame, text="Reason:").pack(side="left")
        self.return_reason = ttk.Entry(control_frame, width=30)
        self.return_reason.pack(side="left", padx=6)

        ttk.Button(control_frame, text="Process Return",
            command=self._process_return
        ).pack(side="left", padx=6)

    def _load_sale_for_return(self, sale_id):
        rows = self.db.query("""
            SELECT si.id, p.name AS product, si.quantity, si.price,
                (si.quantity * si.price) as total
            FROM sale_items si
            JOIN products p ON si.product_id = p.id
            WHERE si.sale_id = ?;
        """, (sale_id,))
        self.return_tree.delete(*self.return_tree.get_children())
        for r in rows:
            self.return_tree.insert("", "end", values=(r["id"], r["product"], r["quantity"], r["price"], r["total"]))

    def _process_return(self):
        selected = self.return_tree.selection()
        if not selected:
            messagebox.showerror("Error", "Select an item to return.")
            return

        item_id, product, qty, price, total = self.return_tree.item(selected[0], "values")
        return_qty = int(self.return_qty.get() or 0)
        reason = self.return_reason.get()

        if return_qty <= 0 or return_qty > int(qty):
            messagebox.showerror("Error", "Invalid return quantity.")
            return

        # 1. Log the return
        self.db.execute("""
            INSERT INTO returns (sale_item_id, quantity, reason)
            VALUES (?, ?, ?)
        """, (item_id, return_qty, reason))

        # 2. Restore stock to batches used in this sale
        batches = self.db.query("""
            SELECT sib.batch_id, sib.quantity
            FROM sale_item_batches sib
            WHERE sib.sale_item_id = ?
            ORDER BY sib.id ASC;
        """, (item_id,))

        remain = return_qty
        for b in batches:
            if remain <= 0:
                break
            give_back = min(remain, b["quantity"])
            self.db.execute(
                "UPDATE batches SET quantity = quantity + ? WHERE id = ?;",
                (give_back, b["batch_id"])
            )
            remain -= give_back

        # 3. Update sale_items to reflect reduced sold qty
        self.db.execute("""
            UPDATE sale_items
            SET quantity = quantity - ?
            WHERE id = ?
        """, (return_qty, item_id))

        messagebox.showinfo("Success", f"Returned {return_qty} x {product}")
        # Refresh items for this sale
        self._load_sale_for_return(self.return_sale_id.get())

    def _build_sale_history_tab(self):
        # kept for backward compatibility (not used as top-level)
        pass

    def _sale_history_refresh(self):
        tree = getattr(self, '_sale_history_tree', None)
        if not tree:
            return
        
        tree.delete(*tree.get_children())
        
        # Simplified query that doesn't rely on sale_item_batches
        rows = self.db.query('''
            SELECT 
                s.id AS sale_id, 
                s.created_at AS date, 
                s.customer_name AS customer,
                p.name AS product, 
                si.quantity AS qty, 
                si.price AS price,
                (si.quantity * si.price) AS subtotal
            FROM sales s
            JOIN sale_items si ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            ORDER BY s.created_at DESC LIMIT 200;
        ''')
        
        for r in rows:
            tree.insert('', 'end', values=(
                r['sale_id'], 
                r['date'], 
                r['customer'] or '', 
                r['product'], 
                r['qty'], 
                f"{r['price']:.2f}", 
                "",  # expiry (empty for now)
                "",  # supplier (empty for now)
                "",  # manufacturer (empty for now)
                f"{r['subtotal']:.2f}"
            ))

    def _sale_history_print_selected(self):
        sel = self._sale_history_tree.selection()
        if not sel: return messagebox.showwarning('Select','Select a row (sale item) to print its sale receipt')
        item = self._sale_history_tree.item(sel[0])['values']
        sale_id = item[0]
        if REPORTLAB_AVAILABLE:
            self._print_receipt(sale_id)
        else:
            messagebox.showwarning('Missing','reportlab required for PDF receipt')

    def _print_receipt(self, sale_id):
        if not REPORTLAB_AVAILABLE:
            return messagebox.showerror('Missing','reportlab is required to generate PDF receipts')
        items = self.db.query('SELECT si.quantity,si.price,p.name FROM sale_items si JOIN products p ON si.product_id=p.id WHERE si.sale_id=?;', (sale_id,))
        sale = self.db.query('SELECT * FROM sales WHERE id=?;', (sale_id,))
        if not sale:
            return messagebox.showerror('Error','Sale not found')
        sale = sale[0]
        fp = os.path.join(RECEIPT_FOLDER, f'receipt_{sale_id}.pdf')
        c = pdf_canvas.Canvas(fp, pagesize=A4); width, height = A4; y = height - 60
        c.setFont('Helvetica-Bold', 14); c.drawString(50, y, 'Pharmacy Receipt'); y -= 25
        c.setFont('Helvetica', 10); c.drawString(50, y, f'Sale ID: {sale_id}'); c.drawString(300, y, f'Date: {sale["created_at"]}'); y -= 20
        if sale['customer_name']: c.drawString(50, y, f'Customer: {sale["customer_name"]}'); c.drawString(300, y, f'Phone: {sale["customer_phone"] or ""}'); y -= 20
        c.drawString(50, y, f'Cashier ID: {sale["user_id"]}'); y -= 25
        c.setFont('Helvetica-Bold', 10); c.drawString(50, y, 'Product'); c.drawString(300, y, 'Qty'); c.drawString(360, y, 'Price'); c.drawString(430, y, 'Subtotal'); y -= 15
        c.setFont('Helvetica', 10)
        for it in items:
            c.drawString(50, y, str(it['name'])); c.drawString(300, y, str(it['quantity'])); c.drawString(360, y, f"{it['price']:.2f}"); c.drawString(430, y, f"{it['price']*it['quantity']:.2f}"); y -= 15
            if y < 80:
                c.showPage(); y = height - 60
        c.setFont('Helvetica-Bold', 12); c.drawString(50, y-20, f'Total: {sale["total"]:.2f}')
        c.save()
        try:
            os.startfile(fp)
        except Exception:
            pass
        messagebox.showinfo('Receipt Saved', f'Saved to {fp}')

    # ---------------- Return History ----------------
    def _build_return_history_tab(self):
        # kept for backward compatibility (not used as top-level)
        pass

    def _return_refresh(self):
        tree = getattr(self, '_return_tree', None)
        if not tree: return
        tree.delete(*tree.get_children())
        rows = self.db.query('''SELECT r.id, r.sale_item_id AS sale_item, p.name AS product, r.quantity AS qty, r.reason, r.created_at AS created, b.expiry_date AS expiry
            FROM returns r JOIN sale_items si ON si.id=r.sale_item_id JOIN products p ON p.id=si.product_id
            LEFT JOIN sale_item_batches sib ON sib.sale_item_id=si.id LEFT JOIN batches b ON b.id=sib.batch_id
            ORDER BY r.id DESC LIMIT 500;''')
        for r in rows:
            tree.insert('', 'end', values=(r['id'], r['sale_item'], r['product'], r['qty'], r['reason'], r['created'], r['expiry'] or ''))

   
        # ---------------- Reports ----------------
    def _build_reports_in_frame(self, frame):
        for w in frame.winfo_children(): 
            w.destroy()
        
        # ---------------- Filters ----------------
        filter_frame = ttk.Frame(frame, padding=10)
        filter_frame.pack(fill='x', pady=5)
        
        # Date range
        ttk.Label(filter_frame, text="From Date:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        from_date = ttk.Entry(filter_frame, width=12)
        from_date.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(filter_frame, text="To Date:").grid(row=0, column=2, padx=5, pady=5, sticky='e')
        to_date = ttk.Entry(filter_frame, width=12)
        to_date.grid(row=0, column=3, padx=5, pady=5)
        
        # Product filter with autocomplete
        ttk.Label(filter_frame, text="Product:").grid(row=0, column=4, padx=5, pady=5, sticky='e')
        product_filter = AutocompleteEntry(filter_frame, width=20, suggestions_getter=self._product_suggestions)
        product_filter.grid(row=0, column=5, padx=5, pady=5)
        
        # Customer filter with autocomplete
        ttk.Label(filter_frame, text="Customer:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
        customer_filter = AutocompleteEntry(filter_frame, width=20, suggestions_getter=self._customer_suggestions)
        customer_filter.grid(row=1, column=1, padx=5, pady=5)
        
        # Supplier filter with autocomplete
        ttk.Label(filter_frame, text="Supplier:").grid(row=1, column=2, padx=5, pady=5, sticky='e')
        supplier_filter = AutocompleteEntry(filter_frame, width=20, suggestions_getter=self._supplier_suggestions)
        supplier_filter.grid(row=1, column=3, padx=5, pady=5)
        
        # Filter button
        filter_btn = ttk.Button(filter_frame, text="Apply Filters", 
                                command=lambda: self._apply_report_filters(
                                    from_date.get(), to_date.get(),
                                    product_filter.get(), customer_filter.get(), supplier_filter.get()
                                ))
        filter_btn.grid(row=2, column=0, columnspan=4, pady=10)
        
        # ---------------- Report Treeview ----------------
        columns = ('sale_id', 'date', 'customer', 'product', 'quantity', 'price', 'subtotal', 'supplier')
        tree = ttk.Treeview(frame, columns=columns, show='headings', height=15)
        
        col_widths = {
            'sale_id': 80, 'date': 150, 'customer': 150, 'product': 200,
            'quantity': 80, 'price': 80, 'subtotal': 100, 'supplier': 150
        }
        for col in columns:
            tree.heading(col, text=col.replace('_', ' ').title())
            tree.column(col, width=col_widths[col], anchor='center')
        
        scrollbar = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        scrollbar.pack(side='right', fill='y', padx=(0,10), pady=10)
        
        # ---------------- Export Buttons ----------------
        export_frame = ttk.Frame(frame)
        export_frame.pack(fill='x', pady=5, padx=10)
        
        ttk.Button(export_frame, text="Refresh", command=self._load_all_sales).pack(side='left', padx=5)
        ttk.Button(export_frame, text="Export CSV", command=lambda: self._export_report(tree, "csv")).pack(side='left', padx=5)
        ttk.Button(export_frame, text="Export Excel", command=lambda: self._export_report(tree, "excel")).pack(side='left', padx=5)
        
        # Save reference
        self._report_tree = tree
        self._load_all_sales()

    def _apply_report_filters(self, from_date, to_date, product_filter, customer_filter, supplier_filter):
        if not hasattr(self, '_report_tree') or self._report_tree is None:
            return
        
        query = '''
            SELECT s.id AS sale_id, s.created_at AS date, s.customer_name AS customer,
                p.name AS product, si.quantity AS quantity, si.price AS price,
                (si.quantity * si.price) AS subtotal, sup.name AS supplier
            FROM sales s
            JOIN sale_items si ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            LEFT JOIN sale_item_batches sib ON sib.sale_item_id = si.id
            LEFT JOIN batches b ON b.id = sib.batch_id
            LEFT JOIN suppliers sup ON sup.id = b.supplier_id
            WHERE 1=1
        '''
        params = []
        if from_date: params += [from_date]; query += " AND date(s.created_at) >= ?"
        if to_date:   params += [to_date];   query += " AND date(s.created_at) <= ?"
        if product_filter:  params += [f'%{product_filter}%']; query += " AND p.name LIKE ?"
        if customer_filter: params += [f'%{customer_filter}%']; query += " AND s.customer_name LIKE ?"
        if supplier_filter: params += [f'%{supplier_filter}%']; query += " AND sup.name LIKE ?"
        
        query += " ORDER BY s.created_at DESC"
        
        try:
            rows = self.db.query(query, tuple(params))
            self._report_tree.delete(*self._report_tree.get_children())
            for r in rows:
                self._report_tree.insert('', 'end', values=(
                    r['sale_id'], r['date'], r['customer'] or 'N/A',
                    r['product'], r['quantity'], f"{r['price']:.2f}",
                    f"{r['subtotal']:.2f}", r['supplier'] or 'N/A'
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load report data: {str(e)}")

    def _product_suggestions(self, term):
            rows = self.db.query('SELECT name FROM products WHERE name LIKE ? ORDER BY name LIMIT 10;', (f'%{term}%',))
            return [r['name'] for r in rows]

    def _customer_suggestions(self, term):
        rows = self.db.query('SELECT DISTINCT customer_name FROM sales WHERE customer_name LIKE ? ORDER BY customer_name LIMIT 10;', (f'%{term}%',))
        return [r['customer_name'] for r in rows if r['customer_name']]

    def _supplier_suggestions(self, term):
        rows = self.db.query('SELECT name FROM suppliers WHERE name LIKE ? ORDER BY name LIMIT 10;', (f'%{term}%',))
        return [r['name'] for r in rows if r['name']]
    
    def _load_all_sales(self):
        if not hasattr(self, '_report_tree') or self._report_tree is None:
            return
        try:
            rows = self.db.query('''
                SELECT s.id AS sale_id, s.created_at AS date, s.customer_name AS customer,
                    p.name AS product, si.quantity AS quantity, si.price AS price,
                    (si.quantity * si.price) AS subtotal,
                    sup.name AS supplier
                FROM sales s
                JOIN sale_items si ON si.sale_id = s.id
                JOIN products p ON p.id = si.product_id
                LEFT JOIN sale_item_batches sib ON sib.sale_item_id = si.id
                LEFT JOIN batches b ON b.id = sib.batch_id
                LEFT JOIN suppliers sup ON sup.id = b.supplier_id
                ORDER BY s.created_at DESC LIMIT 500
            ''')
            self._report_tree.delete(*self._report_tree.get_children())
            for r in rows:
                self._report_tree.insert('', 'end', values=(
                    r['sale_id'], r['date'], r['customer'] or 'N/A',
                    r['product'], r['quantity'], f"{r['price']:.2f}",
                    f"{r['subtotal']:.2f}", r['supplier'] or 'N/A'
                ))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sales data: {str(e)}")

    def _export_report(self, tree, fmt="csv"):
        try:
            rows = [tree.item(i, "values") for i in tree.get_children()]
            if not rows:
                messagebox.showwarning("No Data", "No report data to export.")
                return
            
            filetypes = {"csv": [("CSV Files", "*.csv")], "excel": [("Excel Files", "*.xlsx")]}
            ext = "xlsx" if fmt=="excel" else "csv"
            fpath = filedialog.asksaveasfilename(defaultextension=f".{ext}", filetypes=filetypes[fmt], title=f"Export Report as {fmt.upper()}")
            if not fpath: return
            
            if fmt=="csv":
                import csv
                with open(fpath, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(("Sale ID","Date","Customer","Product","Quantity","Price","Subtotal","Supplier"))
                    writer.writerows(rows)
            else:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Sales Report"
                ws.append(("Sale ID","Date","Customer","Product","Quantity","Price","Subtotal","Supplier"))
                for r in rows: ws.append(r)
                wb.save(fpath)
            
            messagebox.showinfo("Success", f"Report successfully exported to:\n{fpath}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export report: {str(e)}")

    

    # ---------------- Manage Staff ----------------
    def _build_manage_staff_tab(self):
        # Clear old widgets
        for w in self.tab_manage_staff.winfo_children():
            w.destroy()

        # Header
        header = ttk.Frame(self.tab_manage_staff)
        header.pack(fill='x', pady=6)
        ttk.Label(header, text='Manage Staff', font=('Segoe UI',14,'bold')).pack(side='left', padx=8)
        ttk.Button(header, text='Add Staff', command=self._add_staff).pack(side='right', padx=6)

        # Staff Table
        cols = ('id','username','role')
        tree = ttk.Treeview(self.tab_manage_staff, columns=cols, show='headings', height=15)
        tree.heading('id', text='ID'); tree.column('id', width=60, anchor='center')
        tree.heading('username', text='Username'); tree.column('username', width=200, anchor='w')
        tree.heading('role', text='Role'); tree.column('role', width=120, anchor='center')
        tree.pack(fill='both', expand=True, padx=10, pady=10)

        self._staff_tree = tree
        self._refresh_staff()

        # Buttons below table
        btns = ttk.Frame(self.tab_manage_staff)
        btns.pack(fill='x', pady=5)

        def edit_staff():
            sel = tree.selection()
            if not sel: return
            uid = tree.item(sel[0])['values'][0]
            username = tree.item(sel[0])['values'][1]
            role = tree.item(sel[0])['values'][2]
            
            def save(d):
                if not d.get('username'): 
                    return messagebox.showerror('Error','Username required')
                if d.get('password'):
                    self.db.execute(
                        "UPDATE users SET username=?, password_hash=?, role=? WHERE id=?",
                        (d['username'], hash_pw(d['password']), d['role'], uid)
                    )
                else:
                    self.db.execute(
                        "UPDATE users SET username=?, role=? WHERE id=?",
                        (d['username'], d['role'], uid)
                    )
                messagebox.showinfo('Saved','Staff updated successfully')
                self._refresh_staff()

            FormDialog(self.root, 'Edit Staff', [
                {'key':'username','label':'Username'},
                {'key':'password','label':'New Password (leave blank to keep current)'},
                {'key':'role','label':'Role','widget':'combobox','values':['staff','cashier']}
            ], initial={'username': username, 'role': role}, on_submit=save)

        def delete_staff():
            sel = tree.selection()
            if not sel: return
            uid = tree.item(sel[0])['values'][0]
            if uid == self.user['id']:
                return messagebox.showerror('Error', 'Cannot delete your own account')
            if messagebox.askyesno('Confirm','Delete this staff member?'):
                try:
                    self.db.execute("DELETE FROM users WHERE id=?;", (uid,))
                    self._refresh_staff()
                except Exception as e:
                    messagebox.showerror('Error', f'Failed to delete staff: {str(e)}')

        ttk.Button(btns, text='Edit Staff', command=edit_staff).pack(side='left', padx=6)
        ttk.Button(btns, text='Delete Staff', command=delete_staff).pack(side='left', padx=6)

    def _refresh_staff(self):
        tree = getattr(self, '_staff_tree', None)
        if not tree: return
        tree.delete(*tree.get_children())
        rows = self.db.query("SELECT id, username, role FROM users WHERE role IN ('staff', 'cashier') ORDER BY username;")
        for r in rows:
            tree.insert('', 'end', values=(r['id'], r['username'], r['role']))

    def _add_staff(self):
        def save(d):
            if not d.get('username') or not d.get('password'):
                return messagebox.showerror('Error','Username and password required')
            existing = self.db.query("SELECT id FROM users WHERE username=?;", (d['username'],))
            if existing:
                return messagebox.showerror('Error','Username already exists')
            try:
                self.db.execute(
                    "INSERT INTO users(username,password_hash,role) VALUES(?,?,?)",
                    (d['username'], hash_pw(d['password']), d['role'])
                )
                messagebox.showinfo('Saved','Staff added successfully')
                self._refresh_staff()
            except Exception as e:
                messagebox.showerror('Error', f'Failed to add staff: {str(e)}')

        FormDialog(self.root, 'Add Staff', [
            {'key':'username','label':'Username'},
            {'key':'password','label':'Password'},
            {'key':'role','label':'Role','widget':'combobox','values':['staff','cashier']}
        ], on_submit=save)

    # ---------------- Settings ----------------
    def _build_settings_tab(self):
        
        for w in self.tab_settings.winfo_children(): w.destroy()
        f = ttk.Frame(self.tab_settings); f.pack(fill='x', padx=8, pady=8)
        # Pharmacy identity
        ttk.Label(f, text='Pharmacy Name').grid(row=0, column=0, sticky='w', padx=4, pady=4)
        name_e = ttk.Entry(f, width=40); name_e.grid(row=0, column=1, padx=4, pady=4)
        ttk.Label(f, text='Pharmacy Address').grid(row=1, column=0, sticky='w', padx=4, pady=4)
        addr_e = ttk.Entry(f, width=60); addr_e.grid(row=1, column=1, padx=4, pady=4)
        # Existing
        ttk.Label(f, text='Default Tax Percent (%)').grid(row=2, column=0, sticky='w', padx=4, pady=4)
        tax_e = ttk.Entry(f, width=8); tax_e.grid(row=2, column=1, padx=4)
        ttk.Label(f, text='Default Discount (%)').grid(row=3, column=0, sticky='w', padx=4, pady=4)
        disc_e = ttk.Entry(f, width=8); disc_e.grid(row=3, column=1, padx=4)
        def _get(key, default=''):
            rows = self.db.query('SELECT value FROM settings WHERE key=?;', (key,))
            return rows[0]['value'] if rows else default
        name_e.insert(0, _get('pharmacy_name','Pharmacy Receipt'))
        addr_e.insert(0, _get('pharmacy_address','123 Main Street, City'))
        tax_e.insert(0, _get('tax_percent','0'))
        disc_e.insert(0, _get('default_discount','0'))
        def save():
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('pharmacy_name', name_e.get().strip()))
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('pharmacy_address', addr_e.get().strip()))
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('tax_percent', tax_e.get().strip()))
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('default_discount', disc_e.get().strip()))
            messagebox.showinfo('Saved','Settings saved')
        ttk.Button(f, text='Save Settings', command=save).grid(row=4, column=0, columnspan=2, pady=8)

        def save():
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('tax_percent', tax_e.get().strip()))
            self.db.execute('INSERT OR REPLACE INTO settings(key,value) VALUES(?,?);', ('default_discount', disc_e.get().strip()))
            messagebox.showinfo('Saved','Settings saved')
        ttk.Button(f, text='Save Settings', command=save).grid(row=3, column=0, columnspan=2, pady=8)

    # ---------------- Helpers ----------------
    def _open_tab_by_name(self, name):
        for i in range(self.nb.index('end')):
            if self.nb.tab(i, option='text') == name:
                self.nb.select(i); return
        messagebox.showinfo('Info', f'Tab {name} not found')

    def _filter_inventory_low_stock(self):
        med_rows = self.db.query("""SELECT p.id,p.name,p.sku,p.unit,c.name as category,m.name as manufacturer,p.sale_price as price,
            COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id WHERE p.is_medical=1 GROUP BY p.id HAVING stock<=5 ORDER BY p.name;""")
        self._med_tree.delete(*self._med_tree.get_children())
        for r in med_rows: self._med_tree.insert('', 'end', iid=r['id'], values=(r['id'], r['name'], r['sku'] or '', r.get('unit','') or '', r.get('category') or '', r.get('manufacturer') or '', f"{r['price']:.2f}", r['stock']))
        non_rows = self.db.query("""SELECT p.id,p.name,p.sku,p.unit,c.name as category,m.name as manufacturer,p.sale_price as price,
            COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p LEFT JOIN categories c ON p.category_id=c.id LEFT JOIN manufacturers m ON p.manufacturer_id=m.id WHERE p.is_medical=0 GROUP BY p.id HAVING stock<=5 ORDER BY p.name;""")
        self._nonmed_tree.delete(*self._nonmed_tree.get_children())
        for r in non_rows: self._nonmed_tree.insert('', 'end', iid=r['id'], values=(r['id'], r['name'], r['sku'] or '', r.get('unit','') or '', r.get('category') or '', r.get('manufacturer') or '', f"{r['price']:.2f}", r['stock']))
        self._open_tab_by_name('Inventory')
        try:
            for child in self.tab_inventory.winfo_children():
                if isinstance(child, ttk.Notebook):
                    child.select(0); break
        except Exception:
            pass

    def _filter_inventory_near_expiry(self):
        self._batch_tree.delete(*self._batch_tree.get_children())
        rows = self.db.query("""SELECT b.id, p.name as product, b.batch_no, b.quantity, b.expiry_date, s.name as supplier FROM batches b LEFT JOIN products p ON p.id=b.product_id LEFT JOIN suppliers s ON s.id=b.supplier_id WHERE b.expiry_date IS NOT NULL AND julianday(b.expiry_date)-julianday('now')<=30 AND b.quantity>0 ORDER BY b.expiry_date ASC;""")
        for r in rows: self._batch_tree.insert('', 'end', values=(r['id'], r['product'], r['batch_no'] or '', r['quantity'], r['expiry_date'] or '', r['supplier'] or ''))
        self._open_tab_by_name('Inventory')
        try:
            for child in self.tab_inventory.winfo_children():
                if isinstance(child, ttk.Notebook):
                    nb = child; nb.select(nb.index('end')-1); break
        except Exception:
            pass

    def _open_low_stock(self):
        try:
            self._filter_inventory_low_stock()
        except Exception:
            rows = self.db.query("""SELECT p.id,p.name,COALESCE((SELECT SUM(quantity) FROM batches b WHERE b.product_id=p.id),0) AS stock FROM products p GROUP BY p.id HAVING stock<=5;""")
            dlg = tk.Toplevel(self.root); dlg.title('Low Stock Items'); dlg.geometry('700x400')
            tree = ttk.Treeview(dlg, columns=('id','name','stock'), show='headings'); tree.pack(fill='both', expand=True, padx=8, pady=8)
            for c in ('id','name','stock'): tree.heading(c, text=c.capitalize()); tree.column(c,width=200,anchor='w')
            for r in rows: tree.insert('', 'end', values=(r['id'], r['name'], r['stock']))
            ttk.Button(dlg, text='Close', command=dlg.destroy).pack(pady=6)

    def _open_near_expiry(self):
        try:
            self._filter_inventory_near_expiry()
        except Exception:
            rows = self.db.query("""SELECT b.id, p.name AS product, b.batch_no, b.quantity, b.expiry_date FROM batches b JOIN products p ON p.id=b.product_id WHERE b.expiry_date IS NOT NULL AND julianday(b.expiry_date)-julianday('now')<=30 AND b.quantity>0 ORDER BY b.expiry_date ASC;""")
            dlg = tk.Toplevel(self.root); dlg.title('Near Expiry (<=30 days)'); dlg.geometry('800x420')
            tree = ttk.Treeview(dlg, columns=('id','product','batch_no','quantity','expiry'), show='headings'); tree.pack(fill='both', expand=True, padx=8, pady=8)
            for c in ('id','product','batch_no','quantity','expiry'): tree.heading(c, text=c.capitalize()); tree.column(c,width=150,anchor='w')
            for r in rows: tree.insert('', 'end', values=(r['id'], r['product'], r['batch_no'] or '', r['quantity'], r['expiry_date'] or ''))
            ttk.Button(dlg, text='Close', command=dlg.destroy).pack(pady=6)

    # autocomplete helpers
    def _supplier_suggestions(self, term):
        rows = self.db.query('SELECT name FROM suppliers WHERE name LIKE ? ORDER BY name LIMIT 10;', (f'%{term}%',))
        return [r['name'] for r in rows]

    def _manufacturer_suggestions(self, term):
        rows = self.db.query('SELECT name FROM manufacturers WHERE name LIKE ? ORDER BY name LIMIT 10;', (f'%{term}%',))
        return [r['name'] for r in rows]

    def _product_suggestions(self, term):
        rows = self.db.query('SELECT name, sale_price FROM products WHERE name LIKE ? ORDER BY name LIMIT 12;', (f'%{term}%',))
        return [r['name'] for r in rows]


    # ---------------- Seeder ----------------
    def _seed_test_data(self):
        try:
            cnt = self.db.query('SELECT COUNT(*) AS c FROM products;')[0]['c']
            if cnt > 0: return
            man1 = self.db.execute('INSERT OR IGNORE INTO manufacturers(name,contact) VALUES(?,?);', ('GoodPharma','contact1'))
            sup1 = self.db.execute('INSERT OR IGNORE INTO suppliers(name,phone) VALUES(?,?);', ('SupplyCo','1234567890'))
            self.db.execute('INSERT OR IGNORE INTO formulas(name,composition) VALUES(?,?);', ('Paracetamol 500mg','Paracetamol'))
            self.db.execute('INSERT OR IGNORE INTO categories(name) VALUES(?);', ('Analgesics',))
            pid1 = self.db.execute('INSERT INTO products(name,sku,is_medical,unit,sale_price) VALUES(?,?,?,?,?);', ('Paracetamol 500mg','PARA500',1,'tablet',0.50))
            pid2 = self.db.execute('INSERT INTO products(name,sku,is_medical,unit,sale_price) VALUES(?,?,?,?,?);', ('Ibuprofen 200mg','IBU200',1,'tablet',0.75))
            pid3 = self.db.execute('INSERT INTO products(name,sku,is_medical,unit,sale_price) VALUES(?,?,?,?,?);', ('Cough Syrup 100ml','COUGH100',1,'ml',3.50))
            self.db.execute('INSERT INTO batches(product_id,quantity,expiry_date,created_at) VALUES(?,?,?,?);', (pid1,50, (datetime.now()+timedelta(days=20)).strftime('%Y-%m-%d'), now_str()))
            self.db.execute('INSERT INTO batches(product_id,quantity,expiry_date,created_at) VALUES(?,?,?,?);', (pid2,10, (datetime.now()+timedelta(days=200)).strftime('%Y-%m-%d'), now_str()))
            self.db.execute('INSERT INTO batches(product_id,quantity,expiry_date,created_at) VALUES(?,?,?,?);', (pid3,5, (datetime.now()+timedelta(days=10)).strftime('%Y-%m-%d'), now_str()))
            cid = self.db.execute('INSERT INTO customers(name,phone) VALUES(?,?);', ('Alice','5551112222'))
            sale1 = self.db.execute('INSERT INTO sales(user_id,total,customer_id,customer_name,customer_phone,created_at) VALUES(?,?,?,?,?,?);', (1, 15.0, cid, 'Alice','5551112222', now_str()))
            si = self.db.execute('INSERT INTO sale_items(sale_id,product_id,quantity,price) VALUES(?,?,?,?);', (sale1, pid1, 2, 0.5))
            batches = self.db.query('SELECT id,quantity FROM batches WHERE product_id=? ORDER BY created_at ASC;', (pid1,))
            need = 2
            for b in batches:
                take = min(need, b['quantity'])
                if take>0:
                    self.db.execute('UPDATE batches SET quantity=quantity-? WHERE id=?;', (take, b['id']))
                    need -= take
                if need<=0: break
        except Exception as e:
            print('Seeder error', e)

    # ---------------- End App ----------------
    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    app = App()
    app.run()
